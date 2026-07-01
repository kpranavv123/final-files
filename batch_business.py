import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
BATCH_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\Batch\Batch.tab"
OUTPUT_FILE       = r"C:\Users\SW526XH\Downloads\Go Live-2\Batch\Validated_Batch_Business.xlsx"


# ─────────────────────────────────────────────
#  DATE FORMAT
# ─────────────────────────────────────────────
DATE_PATTERN = re.compile(r'^\d{4}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$')  # YYYYMMDD


# ─────────────────────────────────────────────
#  Colours / Styles  — matched to Batch technical template
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUB_FILL        = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  FIELD ORDER & METADATA  (only fields with business rules)
# ─────────────────────────────────────────────
FIELD_ORDER = ["DATEOFMANUFACTURE", "SHELFLIFEEXPIRATION"]

# Both fields use sub-rows in the summary (blank-mismatch vs date-order)
FIELDS_WITH_SUB_ROWS = {"DATEOFMANUFACTURE", "SHELFLIFEEXPIRATION"}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class BatchBusinessRuleEngine:
    """
    Rules:
      DATEOFMANUFACTURE
        a) Manufacturing date must not be greater than expiry date (SHELFLIFEEXPIRATION).
        b) If manufacturing date exists, expiry date must not be blank.
      SHELFLIFEEXPIRATION
        a) Expiry date must not be lesser than manufacturing date (DATEOFMANUFACTURE).
        b) If expiry date exists, manufacturing date must not be blank.
    Both fields are validated together since the rules are cross-field.
    """

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    @staticmethod
    def _parse_date(value):
        s = str(value).strip()
        if not DATE_PATTERN.match(s):
            return None
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except ValueError:
            return None

    def validate_row(self, row) -> dict:
        """Returns {field_name: reason} for whichever of DATEOFMANUFACTURE /
        SHELFLIFEEXPIRATION failed a business rule on this row."""
        reasons = {}

        dom_raw = row.get("DATEOFMANUFACTURE", "")
        sle_raw = row.get("SHELFLIFEEXPIRATION", "")
        dom_blank = self._is_blank(dom_raw)
        sle_blank = self._is_blank(sle_raw)

        # ── Rule 1b / 2b: cross-field blank checks ──
        if not dom_blank and sle_blank:
            reasons["DATEOFMANUFACTURE"] = (
                "DATEOFMANUFACTURE: Exists but SHELFLIFEEXPIRATION is blank"
            )
            reasons["SHELFLIFEEXPIRATION"] = (
                "SHELFLIFEEXPIRATION: Field is blank while DATEOFMANUFACTURE is not blank"
            )
        elif not sle_blank and dom_blank:
            reasons["SHELFLIFEEXPIRATION"] = (
                "SHELFLIFEEXPIRATION: Exists but DATEOFMANUFACTURE is blank"
            )
            reasons["DATEOFMANUFACTURE"] = (
                "DATEOFMANUFACTURE: Field is blank while SHELFLIFEEXPIRATION is not blank"
            )

        # ── Rule 1a / 2a: date order check (only when both present) ──
        if not dom_blank and not sle_blank:
            dom_date = self._parse_date(dom_raw)
            sle_date = self._parse_date(sle_raw)
            if dom_date and sle_date and dom_date > sle_date:
                reasons["DATEOFMANUFACTURE"] = (
                    f"DATEOFMANUFACTURE: '{str(dom_raw).strip()}' is greater than "
                    f"SHELFLIFEEXPIRATION '{str(sle_raw).strip()}'"
                )
                reasons["SHELFLIFEEXPIRATION"] = (
                    f"SHELFLIFEEXPIRATION: '{str(sle_raw).strip()}' is lesser than "
                    f"DATEOFMANUFACTURE '{str(dom_raw).strip()}'"
                )

        return reasons


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class BatchBusinessTableValidator:

    def __init__(self, batch_path: str):
        self.batch_path = batch_path
        self.df         = pd.DataFrame()
        self.error_map  = {}   # row_idx -> [failed field names]
        self.reason_map = {}   # row_idx -> {field: reason}

    def load(self):
        self.df = pd.read_csv(self.batch_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        engine = BatchBusinessRuleEngine()

        for idx, row in self.df.iterrows():
            try:
                reasons = engine.validate_row(row)
            except Exception:
                reasons = {}

            if reasons:
                self.error_map[idx]  = list(reasons.keys())
                self.reason_map[idx] = reasons

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors

    # ── Sub-count helpers ─────────────────────
    def get_dateofmanufacture_error_subcounts(self) -> dict:
        counts = {"blank_mismatch": 0, "date_order": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("DATEOFMANUFACTURE", "")
            if not reason:
                continue
            if "greater than" in reason.lower():
                counts["date_order"] += 1
            else:
                counts["blank_mismatch"] += 1
        return counts

    def get_shelflife_error_subcounts(self) -> dict:
        counts = {"blank_mismatch": 0, "date_order": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("SHELFLIFEEXPIRATION", "")
            if not reason:
                continue
            if "lesser than" in reason.lower():
                counts["date_order"] += 1
            else:
                counts["blank_mismatch"] += 1
        return counts


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class BatchBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "DATEOFMANUFACTURE": [
            "Manufacturing date must not be greater than the expiry date (SHELFLIFEEXPIRATION).",
            "If Manufacturing date exists, Expiry date (SHELFLIFEEXPIRATION) must not be blank.",
        ],
        "SHELFLIFEEXPIRATION": [
            "Expiry date must not be lesser than the Manufacturing date (DATEOFMANUFACTURE).",
            "If Expiry date exists, Manufacturing date (DATEOFMANUFACTURE) must not be blank.",
        ],
    }

    def __init__(self, validator: BatchBusinessTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    # ══════════════════════════════════════════
    #  Summary sheet  (only fields that have business rules)
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="Batch Business Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        # ── Build per-field error counts ──
        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        dom_subcounts = self.validator.get_dateofmanufacture_error_subcounts()
        sle_subcounts = self.validator.get_shelflife_error_subcounts()

        SUB_ROW_DEFS = {
            "DATEOFMANUFACTURE": [
                ("  ↳ Blank / Cross-field mismatch", "blank_mismatch",
                 "DATEOFMANUFACTURE: Blank while SHELFLIFEEXPIRATION is not blank (or vice versa)"),
                ("  ↳ Manufacturing date after Expiry date", "date_order",
                 "DATEOFMANUFACTURE: Greater than SHELFLIFEEXPIRATION"),
            ],
            "SHELFLIFEEXPIRATION": [
                ("  ↳ Blank / Cross-field mismatch", "blank_mismatch",
                 "SHELFLIFEEXPIRATION: Blank while DATEOFMANUFACTURE is not blank (or vice versa)"),
                ("  ↳ Expiry date before Manufacturing date", "date_order",
                 "SHELFLIFEEXPIRATION: Lesser than DATEOFMANUFACTURE"),
            ],
        }
        SUBCOUNTS_BY_FIELD = {
            "DATEOFMANUFACTURE": dom_subcounts,
            "SHELFLIFEEXPIRATION": sle_subcounts,
        }

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value="")

            self._style_summary_data_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1

            if col_name in FIELDS_WITH_SUB_ROWS and has_errors:
                subcounts = SUBCOUNTS_BY_FIELD[col_name]
                for sub_label, sub_key, sub_reason in SUB_ROW_DEFS[col_name]:
                    sub_count      = subcounts.get(sub_key, 0)
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            field_num += 1

        # ── TOTAL row ──
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2   # blank spacer

        # ── Quick-glance stats block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 42, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets  (ALL source columns, not just rule columns) ──
    def _write_field_error_sheets(self, wb, df: pd.DataFrame, all_cols: list):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            # NOTE: business error sheets show ALL extract columns (not just KEEP_COLS)
            keep_here = all_cols + ["ERROR_COLUMNS"]
            subset    = df.loc[row_indices, keep_here].copy()

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER

                # Highlight every field involved in this row's cross-field
                # failure (both DATEOFMANUFACTURE and SHELFLIFEEXPIRATION
                # when relevant), not just the sheet's own field.
                row_reasons = self.validator.reason_map.get(orig_idx, {})
                for involved_field in row_reasons:
                    if involved_field in col_idx_map:
                        target_cell      = ws.cell(row=excel_row, column=col_idx_map[involved_field])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Batch Table – Business Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 75

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        # All source columns, in original order — used for per-field error sheets
        all_cols = list(df.columns)

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df, all_cols)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class BatchBusinessTableProcessor:

    def __init__(self, batch_path: str, output_path: str):
        self.validator = BatchBusinessTableValidator(batch_path)
        self.writer    = BatchBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Batch columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = BatchBusinessTableProcessor(
        batch_path  = BATCH_INPUT_FILE,
        output_path = OUTPUT_FILE,
    )
    processor.run()
