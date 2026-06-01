import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
CUSTOMER_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Cutomer_2026-05-20-1205.tab"
OUTPUT_FILE         = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Validated_Customer_Business2.xlsx"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL  = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL  = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
HDR_FONT  = Font(bold=True, name="Arial", size=9)
BODY_FONT = Font(name="Arial", size=9)
ERR_FONT  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
RULE_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Business Ruleset Info
# ══════════════════════════════════════════════
BUSINESS_RULESET_INFO = {
    # ── UPDATED: REGION_CODE rule replaced with CHANNEL+REGION_CODE combo ──
    "CHANNEL_REGION_CODE": [
        "One CLUSTER should be mapped to a single CHANNEL-REGION_CODE combination.",
    ],
    "AREA_CODE": [
        "One area is mapped to multiple regions.",
    ],
    "SALESHIERARCHY": [
        "One territory is mapped to multiple areas.",
    ],
    "SUB_CHANNEL_CODE_JDA_REPORTING": [
        "One sub-channel is mapped to multiple channels.",
    ],
    "CUSTOMER": [
        "One customer is mapped to multiple territories",
        "One customer is mapped to multiple sub-channels",
        "One customer is mapped to multiple states",
    ],
}


# ─────────────────────────────────────────────
#  SPECIAL KEY for the CHANNEL+REGION_CODE rule
# ─────────────────────────────────────────────
CHANNEL_REGION_RULE_KEY  = "CHANNEL_REGION_CODE"
CHANNEL_REGION_RULE_TEXT = (
    "One CLUSTER should be mapped to a single CHANNEL-REGION_CODE combination."
)

# Format:
# Summary Field, Child Column Candidates, Parent Column Candidates, Rule Text
# NOTE: CHANNEL_REGION_CODE rule is handled separately in validate(); not listed here.
PARENT_CHILD_RULESET_INFO = [
    (
        "AREA_CODE",
        ["AREA_CODE"],
        ["REGION_CODE"],
        "One area is mapped to multiple regions.",
    ),
    (
        "SALESHIERARCHY",
        ["SALESHIERARCHY"],
        ["AREA_CODE"],
        "One territory is mapped to multiple areas.",
    ),
    (
        "CUSTOMER",
        ["CUSTOMER"],
        ["SALESHIERARCHY"],
        "One customer is mapped to multiple territories",
    ),
    (
        "SUB_CHANNEL_CODE_JDA_REPORTING",
        ["SUB_CHANNEL_CODE_JDA_REPORTING"],
        ["CHANNEL"],
        "One sub-channel is mapped to multiple channels.",
    ),
    (
        "CUSTOMER",
        ["CUSTOMER"],
        ["SUB_CHANNEL_CODE_JDA_REPORTING"],
        "One customer is mapped to multiple sub-channels",
    ),
    (
        "CUSTOMER",
        ["CUSTOMER"],
        ["STATE_CODE"],
        "One customer is mapped to multiple states",
    ),
]


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class CustomerBusinessValidator:
    """Loads Customer file and validates business rules only."""

    def __init__(self, customer_path: str):
        self.customer_path = customer_path
        self.df = pd.DataFrame()
        self.error_map = {}
        self.parent_child_error_rows = []
        self.parent_child_error_detail_rows_by_sheet = {}
        self.duplicate_error_rows = []
        self.duplicate_error_indices = set()
        self.duplicate_summary_count = 0

    @staticmethod
    def _clean(value) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip()

    @staticmethod
    def _find_column(df: pd.DataFrame, candidates):
        if isinstance(candidates, str):
            candidates = [candidates]

        lookup = {str(c).strip().upper(): c for c in df.columns}

        for candidate in candidates:
            key = str(candidate).strip().upper()
            if key in lookup:
                return lookup[key]

        return None

    def _add_error(self, idx, field_name: str, reason: str):
        self.error_map.setdefault(idx, {})

        if field_name in self.error_map[idx] and self.error_map[idx][field_name] != reason:
            existing = self.error_map[idx][field_name]
            if reason not in existing.split(" | "):
                self.error_map[idx][field_name] = existing + " | " + reason
        else:
            self.error_map[idx][field_name] = reason

    def load(self):
        self.df = pd.read_csv(self.customer_path, sep="\t", dtype=str)
        self.df.columns = [str(c).strip().upper() for c in self.df.columns]
        print(f"    Customer rows loaded: {len(self.df)}")

    def validate(self):
        print("🔍 Validating customer business rules …")
        self.validate_channel_region_cluster()   # NEW rule replacing old REGION_CODE rule
        self.validate_parent_child_rules()
        # self.validate_duplicate_records()

    # ── NEW: CHANNEL-REGION_CODE combination uniqueness per CLUSTER ──────────
    def validate_channel_region_cluster(self):
        """
        Rule: Each CLUSTER must map to exactly one CHANNEL-REGION_CODE combination.
        Steps:
          1. Concatenate CHANNEL + '-' + REGION_CODE → CHANNEL_REGION_CODE (temp column)
          2. For each CLUSTER, count distinct CHANNEL_REGION_CODE values.
          3. Flag all rows where the CLUSTER has more than one distinct combo.
        """
        channel_col = self._find_column(self.df, ["CHANNEL"])
        region_col  = self._find_column(self.df, ["REGION_CODE"])
        cluster_col = self._find_column(self.df, ["CLUSTER"])

        if channel_col is None or region_col is None or cluster_col is None:
            missing = [
                name for name, col in
                [("CHANNEL", channel_col), ("REGION_CODE", region_col), ("CLUSTER", cluster_col)]
                if col is None
            ]
            print(f"    ⚠️  Skipping CHANNEL_REGION_CODE rule — missing columns: {missing}")
            return

        work = self.df[[channel_col, region_col, cluster_col]].copy()
        work[channel_col] = work[channel_col].apply(self._clean)
        work[region_col]  = work[region_col].apply(self._clean)
        work[cluster_col] = work[cluster_col].apply(self._clean)

        # Build the concatenated combo column (used only internally for comparison)
        work["_CHANNEL_REGION_CODE"] = (
            work[channel_col] + "-" + work[region_col]
        )

        # Filter out rows where any of the three key fields are blank
        work_filtered = work[
            (work[channel_col] != "") &
            (work[region_col]  != "") &
            (work[cluster_col] != "")
        ]

        # Count distinct CHANNEL_REGION_CODE combos per CLUSTER
        combo_count_by_cluster = (
            work_filtered
            .groupby(cluster_col)["_CHANNEL_REGION_CODE"]
            .nunique()
        )
        bad_clusters = set(combo_count_by_cluster[combo_count_by_cluster > 1].index)

        if not bad_clusters:
            print("    ✅ CHANNEL_REGION_CODE rule: no violations found.")
            return

        print(f"    ⚠️  CHANNEL_REGION_CODE rule: {len(bad_clusters)} CLUSTER(s) with multiple combos.")

        sheet_name = f"{cluster_col}_{CHANNEL_REGION_RULE_KEY}"

        for cluster_value in sorted(bad_clusters):
            affected_indices = self.df.index[
                self.df[cluster_col].apply(self._clean) == cluster_value
            ].tolist()

            # Collect all distinct CHANNEL_REGION_CODE values for this cluster
            mapped_combos = sorted(
                set(
                    work_filtered.loc[
                        work_filtered[cluster_col] == cluster_value,
                        "_CHANNEL_REGION_CODE",
                    ]
                ) - {""}
            )

            self.parent_child_error_rows.append({
                "Ruleset":              CHANNEL_REGION_RULE_KEY,
                "Rule":                 CHANNEL_REGION_RULE_TEXT,
                "Child Column":         cluster_col,
                "Child Value":          cluster_value,
                "Parent Column":        "CHANNEL_REGION_CODE (concatenated)",
                "Mapped Parent Values": ", ".join(mapped_combos),
                "Parent Count":         len(mapped_combos),
                "Excel Row Numbers":    ", ".join(str(i + 2) for i in affected_indices),
            })

            for row_index in affected_indices:
                self.parent_child_error_detail_rows_by_sheet.setdefault(sheet_name, []).append({
                    "row_index":  row_index,
                    "ruleset":    CHANNEL_REGION_RULE_KEY,
                    "rule":       CHANNEL_REGION_RULE_TEXT,
                    "child_col":  cluster_col,
                    "parent_col": region_col,      # highlight REGION_CODE in red
                    "extra_highlight_cols": [channel_col],  # also highlight CHANNEL
                })

                self._add_error(row_index, CHANNEL_REGION_RULE_KEY, CHANNEL_REGION_RULE_TEXT)

    def validate_parent_child_rules(self):
        for ruleset, child_candidates, parent_candidates, rule in PARENT_CHILD_RULESET_INFO:
            child_col  = self._find_column(self.df, child_candidates)
            parent_col = self._find_column(self.df, parent_candidates)

            if child_col is None or parent_col is None:
                continue

            sheet_name = f"{child_col}_{parent_col}"

            hierarchy_df = self.df[[child_col, parent_col]].copy()
            hierarchy_df[child_col]  = hierarchy_df[child_col].apply(self._clean)
            hierarchy_df[parent_col] = hierarchy_df[parent_col].apply(self._clean)

            hierarchy_df = hierarchy_df[
                (hierarchy_df[child_col]  != "") &
                (hierarchy_df[parent_col] != "")
            ]

            parent_count_by_child = hierarchy_df.groupby(child_col)[parent_col].nunique(dropna=True)
            invalid_children      = set(parent_count_by_child[parent_count_by_child > 1].index)

            for child_value in sorted(invalid_children):
                affected_rows = self.df.index[
                    self.df[child_col].apply(self._clean) == child_value
                ].tolist()

                mapped_parents = sorted(
                    set(hierarchy_df.loc[hierarchy_df[child_col] == child_value, parent_col]) - {""}
                )

                self.parent_child_error_rows.append({
                    "Ruleset":              ruleset,
                    "Rule":                 rule,
                    "Child Column":         child_col,
                    "Child Value":          child_value,
                    "Parent Column":        parent_col,
                    "Mapped Parent Values": ", ".join(mapped_parents),
                    "Parent Count":         len(mapped_parents),
                    "Excel Row Numbers":    ", ".join(str(i + 2) for i in affected_rows),
                })

                for row_index in affected_rows:
                    self.parent_child_error_detail_rows_by_sheet.setdefault(sheet_name, []).append({
                        "row_index":            row_index,
                        "ruleset":              ruleset,
                        "rule":                 rule,
                        "child_col":            child_col,
                        "parent_col":           parent_col,
                        "extra_highlight_cols": [],
                    })

                    self._add_error(row_index, child_col, rule)

    def validate_duplicate_records(self):
        rule = "Duplicate records are not allowed"

        if self.df.empty:
            return

        normalized_df  = self.df.fillna("").astype(str).apply(lambda s: s.str.strip())
        duplicate_mask = normalized_df.duplicated(keep=False)
        duplicate_indices = self.df.index[duplicate_mask].tolist()

        if not duplicate_indices:
            return

        duplicate_df = normalized_df[duplicate_mask].copy()
        duplicate_df["__duplicate_group__"] = (
            duplicate_df.groupby(list(normalized_df.columns), dropna=False).ngroup() + 1
        )

        self.duplicate_error_indices  = set(duplicate_indices)
        self.duplicate_summary_count  = len(duplicate_indices)

        for row_index, row in duplicate_df.iterrows():
            self.duplicate_error_rows.append({
                "Duplicate Group":  int(row["__duplicate_group__"]),
                "Excel Row Number": int(row_index + 2),
                "Rule":             rule,
            })

    def get_error_series(self) -> pd.Series:
        details = {}

        for idx, errdict in self.error_map.items():
            messages = [f"{fld}: {msg}" for fld, msg in errdict.items()]
            details[idx] = "; ".join(messages)

        return pd.Series(details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class CustomerBusinessReportWriter:
    """Builds the customer business rules report."""

    SHEET_SUMMARY             = "Summary"
    SHEET_RULESETS            = "Rulesets"
    SHEET_PARENT_CHILD_ERRORS = "Parent Child Errors"
    SHEET_DUPLICATE_RECORDS   = "Duplicate Records"

    def __init__(self, validator: CustomerBusinessValidator, output_path: str):
        self.validator = validator
        self.output_path = output_path
        self._summary_fields_order = []

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)

        for ch in invalid_chars:
            name = name.replace(ch, "-")

        name = name.strip() or "Sheet"
        name = name[:31]

        if name not in wb.sheetnames:
            return name

        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = 0

            for c in col:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))

            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 4, 60))

    def _get_ruleset_columns(self):
        fields = [
            "CUSTOMER", "SUPPLYINGPLANT", "AREA_CODE", "AREA_NAME",
            "SALESHIERARCHY", "L1_GLOBAL_CHANNEL_CODE", "L1_GLOBAL_CHANNEL_DESC",
            "CHANNEL", "CHANNELDESC", "SUB_CHANNEL_CODE_JDA_REPORTING",
            "SUB_CHANNEL_DESC_JDA_REPORTING", "REGION_CODE", "REGION_NAME",
            "CLUSTER", "CLUSTER_NAME", "STATE_CODE", "STATE_NAME",
            "ADDITIONALCUSTOMERGROUP1", "ADDITIONALCUSTOMERGROUP1NAME",
            "COUNTRY", "CUSTOMERGROUP", "CUSTOMERGROUPNAME", "CUSTOMERNAME",
            "COUNTRYNAME", "DIVISION", "DIVISIONDESC", "DISTRIBUTIONCHANNEL",
        ]

        cols = [col for col in fields if col in self.validator.df.columns]
        cols.append("ERROR_FIELDS")
        return cols

    def _summary_order(self):
        return [
            CHANNEL_REGION_RULE_KEY,   # replaces old "REGION_CODE"
            "AREA_CODE",
            "SALESHIERARCHY",
            "SUB_CHANNEL_CODE_JDA_REPORTING",
            "CUSTOMER",
        ]

    def _write_ruleset_sheet(self, wb, summary_fields=None):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        title_cell           = ws.cell(row=1, column=1,
                                       value="Customer Table – Business Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:C1")

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        ruleset_info = {
            # ── UPDATED rule ──
            CHANNEL_REGION_RULE_KEY: (
                "Each CLUSTER must be mapped to exactly one CHANNEL-REGION_CODE combination. "
                "The CHANNEL_REGION_CODE is derived by concatenating CHANNEL and REGION_CODE. "
                "If a CLUSTER maps to more than one such combination, all affected rows are flagged."
            ),
            "AREA_CODE":                       "One Area should be mapped to one Region only.",
            "SALESHIERARCHY":                  "One Territory should be mapped to one Area only.",
            "SUB_CHANNEL_CODE_JDA_REPORTING":  "One Sub Channel should be mapped to one Channel only.",
            "CUSTOMER": (
                "One Customer should be mapped to one Territory only. "
                "One Customer should be mapped to one Sub Channel only. "
                "One Customer should be mapped to one State only."
            ),
        }

        ordered_fields = summary_fields or list(ruleset_info.keys())
        current_row    = 4

        for rule_num, field in enumerate(ordered_fields, start=1):
            if field not in ruleset_info:
                continue

            ws.cell(row=current_row, column=1, value=rule_num)
            ws.cell(row=current_row, column=2, value=field)
            ws.cell(row=current_row, column=3, value=ruleset_info[field])

            for c in range(1, 4):
                cell = ws.cell(row=current_row, column=c)
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if c == 1 else "left",
                    vertical="center",
                    wrap_text=True,
                )
                cell.font = Font(name="Arial", size=10, bold=(c in [1, 2]))

                if c in [1, 2]:
                    cell.fill = RULE_FILL
                else:
                    cell.font = BODY_FONT

            current_row += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 95

    def _write_summary_sheet(self, wb, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        title_cell           = ws.cell(row=1, column=1, value="Customer Business Rules Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        col_widths = [6, 36, 16, 16, 16, 16, 85]

        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        field_order      = self._summary_order()
        col_error_counts = {field: 0 for field in field_order}
        rule_error_counts = {}

        for bad_cols in self.validator.error_map.values():
            for col, reason in bad_cols.items():
                if col not in col_error_counts:
                    col_error_counts[col] = 0

                col_error_counts[col] += 1
                rule_error_counts[(col, reason)] = rule_error_counts.get((col, reason), 0) + 1

        sorted_fields              = [(field, col_error_counts.get(field, 0)) for field in field_order]
        self._summary_fields_order = [field for field, _ in sorted_fields]

        row_num      = 4
        item_counter = 1

        for col_name, field_total_errs in sorted_fields:
            reasons        = BUSINESS_RULESET_INFO.get(col_name, [])
            actual_reasons = set(r for (f, r) in rule_error_counts.keys() if f == col_name)
            all_reasons    = list(reasons) + list(actual_reasons - set(reasons))
            is_multi       = len(all_reasons) > 1

            if is_multi:
                ws.cell(row=row_num, column=1, value=item_counter).font = BODY_FONT
                ws.cell(row=row_num, column=2, value=col_name).font     = BODY_FONT
                ws.cell(row=row_num, column=3, value=field_total_errs).font = BODY_FONT
                ws.cell(row=row_num, column=4, value=total_rows).font   = BODY_FONT

                err_pct = field_total_errs / total_rows if total_rows else 0
                ws.cell(row=row_num, column=5, value=1 - err_pct).number_format = "0.00%"
                ws.cell(row=row_num, column=6, value=err_pct).number_format     = "0.00%"
                ws.cell(row=row_num, column=7, value="").font = BODY_FONT

                for c in range(1, 8):
                    ws.cell(row=row_num, column=c).border    = THIN_BORDER
                    ws.cell(row=row_num, column=c).alignment = Alignment(
                        horizontal="center" if c != 7 else "left"
                    )
                    ws.cell(row=row_num, column=c).font = BODY_FONT

                row_num += 1

                for reason in all_reasons:
                    count   = rule_error_counts.get((col_name, reason), 0)
                    sub_pct = count / total_rows if total_rows else 0

                    ws.cell(row=row_num, column=1, value="").font          = BODY_FONT
                    ws.cell(row=row_num, column=2, value=f"↳ {reason}").font = BODY_FONT
                    ws.cell(row=row_num, column=3, value=count).font       = BODY_FONT
                    ws.cell(row=row_num, column=4, value=total_rows).font  = BODY_FONT
                    ws.cell(row=row_num, column=5, value=1 - sub_pct).number_format = "0.00%"
                    ws.cell(row=row_num, column=6, value=sub_pct).number_format     = "0.00%"
                    ws.cell(row=row_num, column=7,
                            value=reason if count > 0 else "").font = BODY_FONT

                    for c in range(1, 8):
                        ws.cell(row=row_num, column=c).border    = THIN_BORDER
                        ws.cell(row=row_num, column=c).alignment = Alignment(
                            horizontal="center" if c != 7 else "left",
                            indent=(1 if c == 2 else 0),
                            wrap_text=(c == 7),
                        )
                        ws.cell(row=row_num, column=c).font = BODY_FONT

                    row_num += 1

            else:
                # Single row (CHANNEL_REGION_CODE falls here — no sub-rows)
                reason  = all_reasons[0] if all_reasons else ""
                err_pct = field_total_errs / total_rows if total_rows else 0

                values = [
                    item_counter,
                    col_name,
                    field_total_errs,
                    total_rows,
                    1 - err_pct,
                    err_pct,
                    reason if field_total_errs > 0 else "",
                ]

                for c_idx, value in enumerate(values, start=1):
                    cell           = ws.cell(row=row_num, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center" if c_idx != 7 else "left",
                        wrap_text=(c_idx == 7),
                    )

                    if c_idx in [5, 6]:
                        cell.number_format = "0.00%"

                row_num += 1

            item_counter += 1

        # ── TOTAL row ──────────────────────────────────────────────────────
        total_errors       = sum(col_error_counts.values())
        total_fill         = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        sum_record_counts  = len(sorted_fields) * total_rows
        total_error_percent = total_errors / sum_record_counts if sum_record_counts else 0

        ws.cell(row=row_num, column=2, value="TOTAL").font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=sum_record_counts).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=5, value=1 - total_error_percent).number_format = "0.00%"
        ws.cell(row=row_num, column=6, value=total_error_percent).number_format     = "0.00%"

        for c in range(1, 8):
            cell           = ws.cell(row=row_num, column=c)
            cell.fill      = total_fill
            cell.border    = THIN_BORDER
            cell.font      = Font(name="Arial", bold=True)
            cell.alignment = Alignment(horizontal="center" if c != 7 else "left")

        row_num += 2

        # ── Stats block ───────────────────────────────────────────────────
        records_with_errors = len(
            set(self.validator.error_map.keys()).union(self.validator.duplicate_error_indices)
        )
        records_passing = total_rows - records_with_errors

        stats      = [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]
        stats_fill = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

        for label, value in stats:
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = stats_fill
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left")

            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center")

            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    def _write_single_parent_child_detail_sheet(self, wb, sheet_name, detail_rows, df: pd.DataFrame):
        safe_name = self._safe_sheet_name(wb, sheet_name)
        ws        = wb.create_sheet(safe_name)

        row_indices    = [item["row_index"] for item in detail_rows]
        subset         = df.loc[row_indices].copy()
        detail_lookup  = {item["row_index"]: item for item in detail_rows}

        subset["ERROR_FIELDS"] = subset.index.map(
            lambda i: detail_lookup.get(i, {}).get("rule", "")
        )

        self._write_header(ws, subset.columns)

        col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

        for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
            detail    = detail_lookup.get(orig_idx, {})
            child_col = detail.get("child_col", "")
            parent_col = detail.get("parent_col", "")
            extra_cols = detail.get("extra_highlight_cols", [])

            for c_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                cell.font      = BODY_FONT
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.fill      = ROW_FILL

            # Highlight all relevant columns in red
            for error_col in [child_col, parent_col] + extra_cols:
                if error_col and error_col in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[error_col])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

        self._set_widths(ws)

        note_row = len(subset) + 3
        ws.cell(
            row=note_row,
            column=1,
            value=f"Total hierarchy error rows for '{safe_name}': {len(subset)}",
        ).font = Font(name="Arial", italic=True, size=9, bold=True)

    def _write_parent_child_errors_sheet(self, wb, df: pd.DataFrame):
        if self.validator.parent_child_error_detail_rows_by_sheet:
            for sheet_name, detail_rows in self.validator.parent_child_error_detail_rows_by_sheet.items():
                self._write_single_parent_child_detail_sheet(wb, sheet_name, detail_rows, df)
        else:
            ws      = wb.create_sheet(self.SHEET_PARENT_CHILD_ERRORS)
            headers = [
                "Ruleset", "Rule", "Child Column", "Child Value", "Parent Column",
                "Mapped Parent Values", "Parent Count", "Excel Row Numbers",
            ]
            self._write_header(ws, headers)
            ws.cell(row=2, column=1, value="No parent-child hierarchy errors found").font = BODY_FONT
            self._set_widths(ws)

    def _write_duplicate_records_sheet(self, wb):
        ws      = wb.create_sheet(self.SHEET_DUPLICATE_RECORDS)
        headers = ["Duplicate Group", "Excel Row Number", "Rule"]
        self._write_header(ws, headers)

        if not self.validator.duplicate_error_rows:
            ws.cell(row=2, column=1, value="No duplicate records found").font = BODY_FONT
        else:
            for r_idx, item in enumerate(self.validator.duplicate_error_rows, start=2):
                for c_idx, header in enumerate(headers, start=1):
                    cell           = ws.cell(row=r_idx, column=c_idx, value=item.get(header, ""))
                    cell.font      = BODY_FONT
                    cell.fill      = ROW_FILL
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self._set_widths(ws)

    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series       = v.get_error_series()
        df["ERROR_FIELDS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        ruleset_columns = self._get_ruleset_columns()
        filtered_cols   = [col for col in df.columns if col in ruleset_columns]
        df              = df[filtered_cols]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, len(df))
        self._write_ruleset_sheet(wb, self._summary_fields_order)
        self._write_parent_child_errors_sheet(wb, df)
        # self._write_duplicate_records_sheet(wb)

        wb.save(self.output_path)

        print(f"\n✅ Business output saved → {self.output_path}")
        print(f"   Total rows                    : {len(df)}")
        print(f"   Business error rows           : {len(set(v.error_map.keys()).union(v.duplicate_error_indices))}")
        print(f"   Parent-child hierarchy issues : {len(v.parent_child_error_rows)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class CustomerBusinessProcessor:
    """Ties together loading, business validation, and report writing."""

    def __init__(self, customer_path: str, output_path: str):
        self.validator = CustomerBusinessValidator(customer_path)
        self.writer    = CustomerBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂 Loading customer file …")
        self.validator.load()
        print(f"    Customer columns detected : {list(self.validator.df.columns)}")
        self.validator.validate()
        print("📝 Writing business report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = CustomerBusinessProcessor(
        customer_path=CUSTOMER_INPUT_FILE,
        output_path=OUTPUT_FILE,
    )
    processor.run()
