import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
CUSTOMER_INPUT_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Cutomer_2026-05-20-1205.tab"
SITE_INPUT_FILE          = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-20-1153.tab"
HDA_INPUT_FILE           = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\BillingDocument(HDA)_2026-05-22-1152.tab"
INDEPENDENT_DEMAND_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\ID\Independent Demand_2026-05-20-1754.tab"
OUTPUT_FILE              = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Validated_Customer_Technical3.xlsx"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=9)
ERR_FONT    = Font(name="Arial", size=9, bold=True, color="FFFFFF")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  FIELD ORDER — single source of truth
#  Used by Summary, Rulesets, and Error sheets.
# ══════════════════════════════════════════════
FIELD_ORDER = [
    "CUSTOMER",
    "CUSTOMERNAME",
    "SUPPLYINGPLANT",
    "CUSTOMERGROUP",
    "CUSTOMERGROUPNAME",
    "ADDITIONALCUSTOMERGROUP1",
    "ADDITIONALCUSTOMERGROUP1NAME",
    "COUNTRY",
    "COUNTRYNAME",
    "DISTRIBUTIONCHANNEL",
    "L1_GLOBAL_CHANNEL_CODE",
    "L1_GLOBAL_CHANNEL_DESC",
    "CHANNEL",
    "CHANNELDESC",
    "SUB_CHANNEL_CODE_JDA_REPORTING",
    "SUB_CHANNEL_DESC_JDA_REPORTING",
    "REGION_CODE",
    "REGION_NAME",
    "AREA_CODE",
    "AREA_NAME",
    "CLUSTER",
    "CLUSTER_NAME",
    "SALESHIERARCHY",
    "STATE_CODE",
    "STATE_NAME",
    "DUPLICATE_CHECK",   # always last
]


# ══════════════════════════════════════════════
#  Technical Ruleset Info
# ══════════════════════════════════════════════
SUMMARY_RULESET_INFO = {
    # Only the "is blank" sub-rule is active for CUSTOMER.
    # The "not found in HDA / IndependentDemand" sub-rule is intentionally
    # commented out below — remove the comment markers to re-enable it.
    "CUSTOMER": [
        "CUSTOMER is blank",
        # "CUSTOMER not found in either HDA or IndependentDemand",  # ← disabled
    ],
    "CUSTOMERNAME":                   ["CUSTOMERNAME is blank"],
    "SUPPLYINGPLANT": [
        "SUPPLYINGPLANT is blank",
        "SUPPLYINGPLANT not found in Site master",
    ],
    "CUSTOMERGROUP":                  ["CUSTOMERGROUP is blank"],
    "CUSTOMERGROUPNAME":              ["CUSTOMERGROUPNAME is blank"],
    "ADDITIONALCUSTOMERGROUP1":       ["ADDITIONALCUSTOMERGROUP1 is blank"],
    "ADDITIONALCUSTOMERGROUP1NAME":   ["ADDITIONALCUSTOMERGROUP1NAME is blank"],
    "COUNTRY":                        ["COUNTRY is blank"],
    "COUNTRYNAME":                    ["COUNTRYNAME is blank"],
    "DISTRIBUTIONCHANNEL":            ["DISTRIBUTIONCHANNEL is blank"],
    "L1_GLOBAL_CHANNEL_CODE":         ["L1_GLOBAL_CHANNEL_CODE is blank"],
    "L1_GLOBAL_CHANNEL_DESC":         ["L1_GLOBAL_CHANNEL_DESC is blank"],
    "CHANNEL":                        ["CHANNEL is blank"],
    "CHANNELDESC":                    ["CHANNELDESC is blank"],
    "SUB_CHANNEL_CODE_JDA_REPORTING": ["SUB_CHANNEL_CODE_JDA_REPORTING is blank"],
    "SUB_CHANNEL_DESC_JDA_REPORTING": ["SUB_CHANNEL_DESC_JDA_REPORTING is blank"],
    "REGION_CODE":                    ["REGION_CODE is blank"],
    "REGION_NAME":                    ["REGION_NAME is blank"],
    "AREA_CODE":                      ["AREA_CODE is blank"],
    "AREA_NAME":                      ["AREA_NAME is blank"],
    "CLUSTER":                        ["CLUSTER is blank"],
    "CLUSTER_NAME":                   ["CLUSTER_NAME is blank"],
    "SALESHIERARCHY":                 ["SALESHIERARCHY is blank"],
    "STATE_CODE":                     ["STATE_CODE is blank"],
    "STATE_NAME":                     ["STATE_NAME is blank"],
    "DUPLICATE_CHECK": [
        "Duplicate row: CUSTOMER + SUPPLYINGPLANT + DISTRIBUTIONCHANNEL combination appears more than once in the extract",
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class CustomerRuleEngine:

    def __init__(self, site_plants: set, hda_customers: set,
                 ind_demand_customers: set, duplicate_triples: set):
        self.site_plants          = set(str(p).strip() for p in site_plants)
        self.hda_customers        = set(str(p).strip() for p in hda_customers)
        self.ind_demand_customers = set(str(p).strip() for p in ind_demand_customers)
        # Set of (customer, supplyingplant, distributionchannel) tuples that are duplicates
        self.duplicate_triples    = duplicate_triples

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _check_not_blank(self, value) -> bool:
        return not self._is_blank(value)

    def validate_customer(self, row) -> tuple:
        val = row.get("CUSTOMER", None)
        if self._is_blank(val):
            return False, "CUSTOMER is blank"
        # ── DISABLED: cross-reference check against HDA / IndependentDemand ──
        # Uncomment the block below to re-enable this sub-rule.
        # val_str = str(val).strip()
        # if val_str in self.hda_customers or val_str in self.ind_demand_customers:
        #     return True, ""
        # return False, "CUSTOMER not found in either HDA or IndependentDemand"
        # ─────────────────────────────────────────────────────────────────────
        return True, ""

    def validate_supplyingplant(self, row) -> tuple:
        val = row.get("SUPPLYINGPLANT", None)
        if self._is_blank(val):
            return False, "SUPPLYINGPLANT is blank"
        if str(val).strip() in self.site_plants:
            return True, ""
        return False, "SUPPLYINGPLANT not found in Site master"

    def validate_duplicate_check(self, row) -> tuple:
        """
        Duplicate key: CUSTOMER + SUPPLYINGPLANT + DISTRIBUTIONCHANNEL.
        All three must be non-blank for the check to apply; blanks are
        already caught by their own individual rules.
        """
        cust    = str(row.get("CUSTOMER",           "")).strip()
        plant   = str(row.get("SUPPLYINGPLANT",     "")).strip()
        distch  = str(row.get("DISTRIBUTIONCHANNEL","")).strip()

        if (cust and cust != "nan"
                and plant and plant != "nan"
                and distch and distch != "nan"):
            if (cust, plant, distch) in self.duplicate_triples:
                return False, (
                    f"Duplicate row: CUSTOMER '{cust}' + SUPPLYINGPLANT '{plant}' "
                    f"+ DISTRIBUTIONCHANNEL '{distch}' combination appears more than once in the extract"
                )
        return True, ""

    def _validate_not_blank_field(self, row, field_name):
        if self._check_not_blank(row.get(field_name, None)):
            return True, ""
        return False, f"{field_name} is blank"

    def validate_customername(self, row):                   return self._validate_not_blank_field(row, "CUSTOMERNAME")
    def validate_customergroup(self, row):                  return self._validate_not_blank_field(row, "CUSTOMERGROUP")
    def validate_customergroupname(self, row):              return self._validate_not_blank_field(row, "CUSTOMERGROUPNAME")
    def validate_additionalcustomergroup1(self, row):       return self._validate_not_blank_field(row, "ADDITIONALCUSTOMERGROUP1")
    def validate_additionalcustomergroup1name(self, row):   return self._validate_not_blank_field(row, "ADDITIONALCUSTOMERGROUP1NAME")
    def validate_country(self, row):                        return self._validate_not_blank_field(row, "COUNTRY")
    def validate_countryname(self, row):                    return self._validate_not_blank_field(row, "COUNTRYNAME")
    def validate_distributionchannel(self, row):            return self._validate_not_blank_field(row, "DISTRIBUTIONCHANNEL")
    def validate_l1_global_channel_code(self, row):         return self._validate_not_blank_field(row, "L1_GLOBAL_CHANNEL_CODE")
    def validate_l1_global_channel_desc(self, row):         return self._validate_not_blank_field(row, "L1_GLOBAL_CHANNEL_DESC")
    def validate_channel(self, row):                        return self._validate_not_blank_field(row, "CHANNEL")
    def validate_channeldesc(self, row):                    return self._validate_not_blank_field(row, "CHANNELDESC")
    def validate_sub_channel_code_jda_reporting(self, row): return self._validate_not_blank_field(row, "SUB_CHANNEL_CODE_JDA_REPORTING")
    def validate_sub_channel_desc_jda_reporting(self, row): return self._validate_not_blank_field(row, "SUB_CHANNEL_DESC_JDA_REPORTING")
    def validate_region_code(self, row):                    return self._validate_not_blank_field(row, "REGION_CODE")
    def validate_region_name(self, row):                    return self._validate_not_blank_field(row, "REGION_NAME")
    def validate_area_code(self, row):                      return self._validate_not_blank_field(row, "AREA_CODE")
    def validate_area_name(self, row):                      return self._validate_not_blank_field(row, "AREA_NAME")
    def validate_cluster(self, row):                        return self._validate_not_blank_field(row, "CLUSTER")
    def validate_cluster_name(self, row):                   return self._validate_not_blank_field(row, "CLUSTER_NAME")
    def validate_saleshierarchy(self, row):                 return self._validate_not_blank_field(row, "SALESHIERARCHY")
    def validate_state_code(self, row):                     return self._validate_not_blank_field(row, "STATE_CODE")
    def validate_state_name(self, row):                     return self._validate_not_blank_field(row, "STATE_NAME")

    def get_rules(self) -> dict:
        """Returns rules in FIELD_ORDER so iteration is always consistent."""
        return {
            "CUSTOMER":                        self.validate_customer,
            "CUSTOMERNAME":                    self.validate_customername,
            "SUPPLYINGPLANT":                  self.validate_supplyingplant,
            "CUSTOMERGROUP":                   self.validate_customergroup,
            "CUSTOMERGROUPNAME":               self.validate_customergroupname,
            "ADDITIONALCUSTOMERGROUP1":        self.validate_additionalcustomergroup1,
            "ADDITIONALCUSTOMERGROUP1NAME":    self.validate_additionalcustomergroup1name,
            "COUNTRY":                         self.validate_country,
            "COUNTRYNAME":                     self.validate_countryname,
            "DISTRIBUTIONCHANNEL":             self.validate_distributionchannel,
            "L1_GLOBAL_CHANNEL_CODE":          self.validate_l1_global_channel_code,
            "L1_GLOBAL_CHANNEL_DESC":          self.validate_l1_global_channel_desc,
            "CHANNEL":                         self.validate_channel,
            "CHANNELDESC":                     self.validate_channeldesc,
            "SUB_CHANNEL_CODE_JDA_REPORTING":  self.validate_sub_channel_code_jda_reporting,
            "SUB_CHANNEL_DESC_JDA_REPORTING":  self.validate_sub_channel_desc_jda_reporting,
            "REGION_CODE":                     self.validate_region_code,
            "REGION_NAME":                     self.validate_region_name,
            "AREA_CODE":                       self.validate_area_code,
            "AREA_NAME":                       self.validate_area_name,
            "CLUSTER":                         self.validate_cluster,
            "CLUSTER_NAME":                    self.validate_cluster_name,
            "SALESHIERARCHY":                  self.validate_saleshierarchy,
            "STATE_CODE":                      self.validate_state_code,
            "STATE_NAME":                      self.validate_state_name,
            "DUPLICATE_CHECK":                 self.validate_duplicate_check,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class CustomerTableValidator:

    def __init__(self, customer_path: str, site_path: str, hda_path: str, ind_demand_path: str):
        self.customer_path        = customer_path
        self.site_path            = site_path
        self.hda_path             = hda_path
        self.ind_demand_path      = ind_demand_path
        self.df                   = pd.DataFrame()
        self.site_plants          = set()
        self.hda_customers        = set()
        self.ind_demand_customers = set()
        self.error_map            = {}

    def _add_error(self, idx, field_name: str, reason: str):
        self.error_map.setdefault(idx, {})
        if field_name in self.error_map[idx] and self.error_map[idx][field_name] != reason:
            existing = self.error_map[idx][field_name]
            if reason not in existing.split(" | "):
                self.error_map[idx][field_name] = existing + " | " + reason
        else:
            self.error_map[idx][field_name] = reason

    def _read_tab(self, path: str) -> pd.DataFrame:
        p = path.lower()
        if p.endswith(".csv"):
            return pd.read_csv(path, dtype=str)
        elif p.endswith((".tab", ".tsv")):
            return pd.read_csv(path, sep="\t", dtype=str,
                               encoding="utf-8", encoding_errors="replace", engine="python")
        elif p.endswith((".xlsx", ".xls")):
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {path}")

    def load(self):
        self.df = self._read_tab(self.customer_path)
        self.df.columns = [str(c).strip().upper() for c in self.df.columns]

        site_df = self._read_tab(self.site_path)
        site_df.columns = [str(c).strip().upper() for c in site_df.columns]
        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site table.")
        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site table – PLANT values loaded: {len(self.site_plants)}")

        hda_df = pd.read_csv(self.hda_path, sep="\t", dtype=str,
                              encoding="utf-8", encoding_errors="replace", engine="python")
        hda_df.columns = [str(c).strip().upper() for c in hda_df.columns]
        if "SOLDTOPARTY" not in hda_df.columns:
            raise ValueError("SOLDTOPARTY column not found in HDA.tab.")
        self.hda_customers = set(hda_df["SOLDTOPARTY"].dropna().str.strip().tolist())
        print(f"    HDA.tab – SOLDTOPARTY values loaded: {len(self.hda_customers)}")

        ind_df = pd.read_csv(self.ind_demand_path, sep="\t", dtype=str,
                             encoding="utf-8", encoding_errors="replace", engine="python")
        ind_df.columns = [str(c).strip().upper() for c in ind_df.columns]
        if "SOLDTOPARTY" not in ind_df.columns:
            raise ValueError("SOLDTOPARTY column not found in IndependentDemand.tab.")
        self.ind_demand_customers = set(ind_df["SOLDTOPARTY"].dropna().str.strip().tolist())
        print(f"    IndependentDemand.tab – SOLDTOPARTY values loaded: {len(self.ind_demand_customers)}")

    def validate(self):
        # ── Pre-compute duplicate triples: CUSTOMER + SUPPLYINGPLANT + DISTRIBUTIONCHANNEL ──
        has_cust   = "CUSTOMER"            in self.df.columns
        has_plant  = "SUPPLYINGPLANT"      in self.df.columns
        has_distch = "DISTRIBUTIONCHANNEL" in self.df.columns

        if has_cust and has_plant and has_distch:
            combo = (
                self.df["CUSTOMER"].fillna("").str.strip()
                + "|||"
                + self.df["SUPPLYINGPLANT"].fillna("").str.strip()
                + "|||"
                + self.df["DISTRIBUTIONCHANNEL"].fillna("").str.strip()
            )
            dup_mask   = combo.duplicated(keep=False)
            # Exclude rows where any part is blank (they have their own rules)
            valid_mask = (
                (self.df["CUSTOMER"].fillna("").str.strip() != "") &
                (self.df["SUPPLYINGPLANT"].fillna("").str.strip() != "") &
                (self.df["DISTRIBUTIONCHANNEL"].fillna("").str.strip() != "")
            )
            dup_combos_raw  = combo[dup_mask & valid_mask].tolist()
            duplicate_triples = set(
                tuple(c.split("|||", 2)) for c in dup_combos_raw
            )
        else:
            missing_cols = [c for c, f in [
                ("CUSTOMER", has_cust),
                ("SUPPLYINGPLANT", has_plant),
                ("DISTRIBUTIONCHANNEL", has_distch)
            ] if not f]
            print(f"⚠️  Duplicate check skipped — missing columns: {missing_cols}")
            duplicate_triples = set()

        engine = CustomerRuleEngine(
            self.site_plants,
            self.hda_customers,
            self.ind_demand_customers,
            duplicate_triples,
        )
        rules = engine.get_rules()

        for idx, row in self.df.iterrows():
            for col, rule_fn in rules.items():
                # DUPLICATE_CHECK is a virtual rule — no real column in the dataframe
                if col != "DUPLICATE_CHECK" and col not in self.df.columns:
                    continue
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {e}"

                if not passed:
                    self._add_error(idx, col, reason)

    def get_error_series(self) -> pd.Series:
        details = {}
        for idx, errdict in self.error_map.items():
            messages = [f"{fld}: {msg}" for fld, msg in errdict.items()]
            details[idx] = "; ".join(messages)
        return pd.Series(details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class CustomerReportWriter:

    SHEET_SUMMARY  = "Summary"
    SHEET_RULESETS = "Rulesets"

    def __init__(self, validator: CustomerTableValidator, output_path: str):
        self.validator             = validator
        self.output_path           = output_path
        self._summary_fields_order = list(FIELD_ORDER)   # driven by FIELD_ORDER constant

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)
        for ch in invalid_chars:
            name = name.replace(ch, "-")
        name = name.strip() or "Sheet"
        name = name[:31]
        if name not in wb.sheetnames:
            return name
        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = 0
            for c in col:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 4, 60))

    def _get_ruleset_columns(self):
        """Columns written to every error sheet — FIELD_ORDER (minus DUPLICATE_CHECK) + ERROR_FIELDS."""
        cols = [col for col in FIELD_ORDER
                if col != "DUPLICATE_CHECK" and col in self.validator.df.columns]
        cols.append("ERROR_FIELDS")
        return cols

    def _write_ruleset_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Customer Table – Technical Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        ruleset_info = {
            "CUSTOMER": (
                "Must not be blank."
                # Disabled: "Must also be found in HDA (SOLDTOPARTY) or IndependentDemand (SOLDTOPARTY)."
            ),
            "CUSTOMERNAME":                    "Field should not be blank.",
            "SUPPLYINGPLANT":                  "Must not be blank. Must be present in the Site master.",
            "CUSTOMERGROUP":                   "Field should not be blank.",
            "CUSTOMERGROUPNAME":               "Field should not be blank.",
            "ADDITIONALCUSTOMERGROUP1":        "Field should not be blank.",
            "ADDITIONALCUSTOMERGROUP1NAME":    "Field should not be blank.",
            "COUNTRY":                         "Field should not be blank.",
            "COUNTRYNAME":                     "Field should not be blank.",
            "DISTRIBUTIONCHANNEL":             "Field should not be blank.",
            "L1_GLOBAL_CHANNEL_CODE":          "Field should not be blank.",
            "L1_GLOBAL_CHANNEL_DESC":          "Field should not be blank.",
            "CHANNEL":                         "Field should not be blank.",
            "CHANNELDESC":                     "Field should not be blank.",
            "SUB_CHANNEL_CODE_JDA_REPORTING":  "Field should not be blank.",
            "SUB_CHANNEL_DESC_JDA_REPORTING":  "Field should not be blank.",
            "REGION_CODE":                     "Field should not be blank.",
            "REGION_NAME":                     "Field should not be blank.",
            "AREA_CODE":                       "Field should not be blank.",
            "AREA_NAME":                       "Field should not be blank.",
            "CLUSTER":                         "Field should not be blank.",
            "CLUSTER_NAME":                    "Field should not be blank.",
            "SALESHIERARCHY":                  "Field should not be blank.",
            "STATE_CODE":                      "Field should not be blank.",
            "STATE_NAME":                      "Field should not be blank.",
            "DUPLICATE_CHECK": (
                "The combination of CUSTOMER + SUPPLYINGPLANT + DISTRIBUTIONCHANNEL must be unique "
                "across the entire extract. If the same values for all three columns appear together "
                "in more than one row, all such rows are flagged as duplicates."
            ),
        }

        current_row = 4
        for rule_num, field in enumerate(FIELD_ORDER, start=1):
            if field not in ruleset_info:
                continue

            ws.cell(row=current_row, column=1, value=rule_num)
            ws.cell(row=current_row, column=2, value=field)
            ws.cell(row=current_row, column=3, value=ruleset_info[field])

            for c in range(1, 4):
                cell           = ws.cell(row=current_row, column=c)
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if c == 1 else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if c in [1, 2]:
                    cell.fill = RULE_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
                else:
                    cell.font = BODY_FONT

            current_row += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 95

    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        ws.merge_cells("A1:E1")
        title_cell           = ws.cell(row=1, column=1, value="Customer Technical Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        col_widths = [6, 36, 16, 16, 16, 16, 85]

        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_error_counts  = {col: 0 for col in FIELD_ORDER}
        rule_error_counts = {}

        for bad_cols in error_map.values():
            for col, reason in bad_cols.items():
                if col not in col_error_counts:
                    continue
                col_error_counts[col] += 1
                rule_error_counts[(col, reason)] = rule_error_counts.get((col, reason), 0) + 1

        row_num      = 4
        item_counter = 1

        for col_name in FIELD_ORDER:
            field_total_errs = col_error_counts.get(col_name, 0)
            reasons          = SUMMARY_RULESET_INFO.get(col_name, [])
            actual_reasons   = set(r for (f, r) in rule_error_counts.keys() if f == col_name)
            all_reasons      = list(reasons) + list(actual_reasons - set(reasons))
            is_multi         = len(all_reasons) > 1

            if col_name == "DUPLICATE_CHECK":
                all_reasons = [
                    "Duplicate row: CUSTOMER + SUPPLYINGPLANT + DISTRIBUTIONCHANNEL "
                    "combination appears more than once in extract"
                ]
                is_multi = False

            if is_multi:
                ws.cell(row=row_num, column=1, value=item_counter)
                ws.cell(row=row_num, column=2, value=col_name)
                ws.cell(row=row_num, column=3, value=field_total_errs)
                ws.cell(row=row_num, column=4, value=total_rows)

                err_pct = field_total_errs / total_rows if total_rows else 0
                ws.cell(row=row_num, column=5, value=1 - err_pct).number_format = "0.00%"
                ws.cell(row=row_num, column=6, value=err_pct).number_format     = "0.00%"
                ws.cell(row=row_num, column=7, value="")

                for c in range(1, 8):
                    ws.cell(row=row_num, column=c).border    = THIN_BORDER
                    ws.cell(row=row_num, column=c).alignment = Alignment(
                        horizontal="center" if c != 7 else "left"
                    )
                    ws.cell(row=row_num, column=c).font = BODY_FONT

                row_num += 1

                for reason in all_reasons:
                    count   = rule_error_counts.get((col_name, reason), 0)
                    sub_pct = count / total_rows if total_rows else 0

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=f"↳ {reason}")
                    ws.cell(row=row_num, column=3, value=count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=1 - sub_pct).number_format = "0.00%"
                    ws.cell(row=row_num, column=6, value=sub_pct).number_format     = "0.00%"
                    ws.cell(row=row_num, column=7, value=reason if count > 0 else "")

                    for c in range(1, 8):
                        ws.cell(row=row_num, column=c).border    = THIN_BORDER
                        ws.cell(row=row_num, column=c).alignment = Alignment(
                            horizontal="center" if c != 7 else "left",
                            indent=(1 if c == 2 else 0),
                            wrap_text=(c == 7),
                        )
                        ws.cell(row=row_num, column=c).font = BODY_FONT

                    row_num += 1

            else:
                reason  = all_reasons[0] if all_reasons else ""
                err_pct = field_total_errs / total_rows if total_rows else 0

                values = [
                    item_counter,
                    col_name,
                    field_total_errs,
                    total_rows,
                    1 - err_pct,
                    err_pct,
                    reason if field_total_errs > 0 else "",
                ]

                for c_idx, value in enumerate(values, start=1):
                    cell           = ws.cell(row=row_num, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center" if c_idx != 7 else "left",
                        wrap_text=(c_idx == 7),
                    )
                    if c_idx in [5, 6]:
                        cell.number_format = "0.00%"

                row_num += 1

            item_counter += 1

        # ── TOTAL row ──
        total_errors        = sum(col_error_counts.values())
        total_fill          = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        sum_record_counts   = len(FIELD_ORDER) * total_rows
        total_error_percent = total_errors / sum_record_counts if sum_record_counts else 0

        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=sum_record_counts)
        ws.cell(row=row_num, column=5, value=1 - total_error_percent).number_format = "0.00%"
        ws.cell(row=row_num, column=6, value=total_error_percent).number_format     = "0.00%"

        for c in range(1, 8):
            cell           = ws.cell(row=row_num, column=c)
            cell.fill      = total_fill
            cell.border    = THIN_BORDER
            cell.font      = Font(name="Arial", bold=True)
            cell.alignment = Alignment(horizontal="center" if c != 7 else "left")

        row_num += 2

        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors
        stats_fill          = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            lc           = ws.cell(row=row_num, column=1, value=label)
            lc.font      = Font(name="Arial", bold=True, size=10)
            lc.fill      = stats_fill
            lc.border    = THIN_BORDER
            lc.alignment = Alignment(horizontal="left")

            vc           = ws.cell(row=row_num, column=3, value=value)
            vc.font      = Font(name="Arial", size=10)
            vc.border    = THIN_BORDER
            vc.alignment = Alignment(horizontal="center")

            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v          = self.validator
        all_fields = set()
        for errdict in v.error_map.values():
            all_fields.update(errdict.keys())

        # Error sheets follow FIELD_ORDER
        fields_to_process = [f for f in FIELD_ORDER if f in all_fields]

        for field_name in fields_to_process:
            row_indices = [idx for idx, errdict in v.error_map.items() if field_name in errdict]
            if not row_indices:
                continue

            sheet_name = self._safe_sheet_name(wb, field_name)
            ws         = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()
            subset["ERROR_FIELDS"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(field_name, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.fill      = ROW_FILL

                # ── Highlight logic ───────────────────────────────────────
                if field_name == "DUPLICATE_CHECK":
                    # Highlight all three key columns in red
                    for highlight_col in ("CUSTOMER", "SUPPLYINGPLANT", "DISTRIBUTIONCHANNEL"):
                        if highlight_col in col_idx_map:
                            cell      = ws.cell(row=excel_row, column=col_idx_map[highlight_col])
                            cell.fill = RED_FILL
                            cell.font = ERR_FONT
                else:
                    if field_name in col_idx_map:
                        cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                        cell.fill = RED_FILL
                        cell.font = ERR_FONT
                # ─────────────────────────────────────────────────────────

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series       = v.get_error_series()
        df["ERROR_FIELDS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        # Keep only FIELD_ORDER columns (excluding the virtual DUPLICATE_CHECK) + ERROR_FIELDS
        ruleset_columns = self._get_ruleset_columns()
        filtered_cols   = [col for col in df.columns if col in ruleset_columns]
        df              = df[filtered_cols]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, len(df))
        self._write_ruleset_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        print(f"\n✅ Technical output saved → {self.output_path}")
        print(f"   Total rows : {len(df)}")
        print(f"   Error rows : {len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class CustomerTableProcessor:

    def __init__(self, customer_path: str, site_path: str, hda_path: str,
                 ind_demand_path: str, output_path: str):
        self.validator = CustomerTableValidator(customer_path, site_path, hda_path, ind_demand_path)
        self.writer    = CustomerReportWriter(self.validator, output_path)

    def run(self):
        print("📂 Loading files …")
        self.validator.load()
        print(f"    Customer columns detected : {list(self.validator.df.columns)}")
        print("🔍 Validating technical rules …")
        self.validator.validate()
        print("📝 Writing technical report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = CustomerTableProcessor(
        customer_path   = CUSTOMER_INPUT_FILE,
        site_path       = SITE_INPUT_FILE,
        hda_path        = HDA_INPUT_FILE,
        ind_demand_path = INDEPENDENT_DEMAND_FILE,
        output_path     = OUTPUT_FILE,
    )
    processor.run()
