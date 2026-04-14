import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ======================================================
# File paths
# ======================================================
HDA_FILE         = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\HistoricalDemandActuals.tab"
SUMMARY_HDA_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\HDA_Validated.tab"

PART_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\Part_site_FG 10.04.2026.csv"
CUSTOMER_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\Customer_2026-04-10-1807.xlsx"
SITE_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\Site_2026-04-09-1058.csv.xlsx"

OUTPUT_EXCEL = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\Validated_HDA.xlsx"

# ======================================================
# Helper: Universal loader
# ======================================================
def read_input(path: str) -> pd.DataFrame:
    if path.lower().endswith('.csv'):
        return pd.read_csv(path, dtype=str)
    return pd.read_excel(path, dtype=str)

# ======================================================
# Load master/reference data
# ======================================================
print("📂 Loading reference files...")

part_df = read_input(PART_FILE)
part_df.columns = part_df.columns.str.strip().str.upper()
if "MATERIALNUMBER_PLANT" not in part_df.columns:
    part_df["MATERIALNUMBER_PLANT"] = part_df["MATERIALNUMBER"].astype(str).str.strip() + "_" + part_df["PLANT"].astype(str).str.strip()
part_set = set(part_df["MATERIALNUMBER_PLANT"].dropna().str.strip())

customer_df = read_input(CUSTOMER_FILE)
customer_df.columns = customer_df.columns.str.strip().str.upper()
if "SUPPLYINGPLANT_CUSTOMER" not in customer_df.columns:
    customer_df["SUPPLYINGPLANT_CUSTOMER"] = customer_df["SUPPLYINGPLANT"].astype(str).str.strip() + "_" + customer_df["CUSTOMER"].astype(str).str.strip()
customer_set = set(customer_df["SUPPLYINGPLANT_CUSTOMER"].dropna().str.strip())

site_df = read_input(SITE_FILE)
site_df.columns = site_df.columns.str.strip().str.upper()
site_set = set(site_df["PLANT"].dropna().str.strip())

# ======================================================
# Validation rules
# ======================================================
rules = [
    ("PLANT",             "ERROR_PLANT",             "Plant is not present in site master."),
    ("BILLING_WEEK_START","ERROR_BILLING_WEEK_START", "Must not be blank and must be in YYYYMMDD format."),
    ("MATERIAL_PLANT",    "ERROR_MATERIAL_PLANT",     "Material-Plant combination not present in the Part master."),
    ("PLANT_SOLDTOPARTY", "ERROR_PLANT_SOLDTOPARTY",  "Plant-SoldToParty combination is not present in customer master."),
]

ERROR_MESSAGES = {col: reason for field, col, reason in rules}

ERROR_SHEETS = {
    "ERROR_PLANT":             ("PLANT",             ERROR_MESSAGES["ERROR_PLANT"]),
    "ERROR_BILLING_WEEK_START":("BILLING_WEEK_START", ERROR_MESSAGES["ERROR_BILLING_WEEK_START"]),
    "ERROR_MATERIAL_PLANT":    ("MATERIAL_PLANT",    ERROR_MESSAGES["ERROR_MATERIAL_PLANT"]),
    "ERROR_PLANT_SOLDTOPARTY": ("PLANT_SOLDTOPARTY", ERROR_MESSAGES["ERROR_PLANT_SOLDTOPARTY"]),
}

# ======================================================
# Constants
# ======================================================
CHUNK_SIZE     = 500_000
EXCEL_MAX_ROWS = 1_048_576
date_pattern   = re.compile(r"^\d{8}$")

sheet_tracker = {
    sheet: {"sheet_no": 1, "current_row": 0}
    for sheet, _ in ERROR_SHEETS.values()
}

error_counts = {field: 0 for field, _, _ in rules}
total_records = 0
records_with_errors = 0

# ======================================================
# WRITE ERROR DATA
# ======================================================
print("🔍 Validating HDA via streaming chunk ingestion...")
with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    # Guard: write a placeholder so the workbook is never empty
    pd.DataFrame({"_": ["placeholder"]}).to_excel(writer, sheet_name="_placeholder", index=False)

    for chunk in pd.read_csv(HDA_FILE, sep="\t", dtype=str, chunksize=CHUNK_SIZE):
        chunk = chunk.apply(lambda x: x.str.strip())
        chunk.columns = chunk.columns.str.strip().str.upper()
        total_records += len(chunk)

        # Build composite keys if they don't already exist in the .tab
        if "MATERIAL_PLANT" not in chunk.columns:
            chunk["MATERIAL_PLANT"] = chunk["MATERIAL"].astype(str).str.strip() + "_" + chunk["PLANT"].astype(str).str.strip()
        if "PLANT_SOLDTOPARTY" not in chunk.columns:
            chunk["PLANT_SOLDTOPARTY"] = chunk["PLANT"].astype(str).str.strip() + "_" + chunk["SOLDTOPARTY"].astype(str).str.strip()

        chunk["ERROR_PLANT"] = chunk["PLANT"].apply(
            lambda x: "Yes" if pd.isna(x) or x not in site_set else "")
        chunk["ERROR_BILLING_WEEK_START"] = chunk["BILLING_WEEK_START"].apply(
            lambda x: "Yes" if pd.isna(x) or not date_pattern.match(x) else "")
        chunk["ERROR_MATERIAL_PLANT"] = chunk["MATERIAL_PLANT"].apply(
            lambda x: "Yes" if pd.isna(x) or x not in part_set else "")
        chunk["ERROR_PLANT_SOLDTOPARTY"] = chunk["PLANT_SOLDTOPARTY"].apply(
            lambda x: "Yes" if pd.isna(x) or x not in customer_set else "")

        row_has_error = False
        for field, col, _ in rules:
            cnt = (chunk[col] == "Yes").sum()
            error_counts[field] += cnt
            row_has_error = row_has_error | (chunk[col] == "Yes")
        records_with_errors += row_has_error.sum()

        for error_col, (base_sheet, error_msg) in ERROR_SHEETS.items():
            error_rows = chunk[chunk[error_col] == "Yes"].copy()
            if error_rows.empty:
                continue

            error_rows["ERROR_COLUMNS"] = error_msg
            tracker = sheet_tracker[base_sheet]
            start = 0

            while start < len(error_rows):
                sheet_name = base_sheet if tracker["sheet_no"] == 1 else f"{base_sheet}_{tracker['sheet_no']}"
                remaining  = EXCEL_MAX_ROWS - tracker["current_row"]
                write_df   = error_rows.iloc[start:start + remaining]

                write_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=tracker["current_row"],
                    index=False,
                    header=(tracker["current_row"] == 0)
                )

                start += len(write_df)
                tracker["current_row"] += len(write_df)

                if tracker["current_row"] >= EXCEL_MAX_ROWS:
                    tracker["sheet_no"] += 1
                    tracker["current_row"] = 0

records_passing = total_records - records_with_errors

print("📝 Generating standardized reports...")

# ======================================================
# LOAD WORKBOOK FOR STYLING
# ======================================================
wb = load_workbook(OUTPUT_EXCEL)

# Remove placeholder sheet
if "_placeholder" in wb.sheetnames:
    del wb["_placeholder"]
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

bold            = Font(bold=True)
center          = Alignment(horizontal="center")
thin_side       = Side(style="thin")
border          = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
title_fill      = PatternFill("solid", fgColor="BDD7EE")
header_fill     = PatternFill("solid", fgColor="D9E1F2")
green_fill      = PatternFill("solid", fgColor="E2EFDA")
total_fill      = PatternFill("solid", fgColor="F2F2F2")
pale_yellow_fill = PatternFill("solid", fgColor="FFF2CC")
red_fill         = PatternFill("solid", fgColor="FF0000")
blue_header_fill = PatternFill("solid", fgColor="BDD7EE")

# ======================================================
# APPLY STYLING TO ERROR SHEETS
# Sort by base_sheet name length descending so "PLANT_SOLDTOPARTY"
# is always matched before "PLANT" — fixes wrong-column-red bug
# ======================================================
sorted_error_sheets = sorted(
    ERROR_SHEETS.items(),
    key=lambda x: len(x[1][0]),
    reverse=True
)

# Collect all error sheet names present in workbook
all_error_sheet_names = set()
for error_col, (base_sheet, _) in sorted_error_sheets:
    for sname in wb.sheetnames:
        if sname == base_sheet or sname.startswith(base_sheet + "_"):
            all_error_sheet_names.add(sname)

for sheet_name in all_error_sheet_names:
    ws = wb[sheet_name]

    # Match to base_sheet using longest-first sorted list
    matched_base = None
    for error_col, (base_sheet, _) in sorted_error_sheets:
        if sheet_name == base_sheet or sheet_name.startswith(base_sheet + "_"):
            matched_base = base_sheet
            break
    if matched_base is None:
        continue

    # Find the column index of the error-highlighted column from header row
    highlight_col_idx = None
    for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]:
        if cell.value == matched_base:
            highlight_col_idx = cell.column
            break

    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.fill      = blue_header_fill
                cell.font      = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.border    = border
            else:
                cell.border = border
                if highlight_col_idx is not None and col_idx == highlight_col_idx:
                    cell.fill = red_fill
                else:
                    cell.fill = pale_yellow_fill

# ======================================================
# SUMMARY SHEET
# ======================================================
ws = wb.create_sheet("Summary")

ws.merge_cells("A1:G1")
ws["A1"]           = "HDA Validation Summary"
ws["A1"].font      = Font(bold=True, size=14)
ws["A1"].fill      = title_fill
ws["A1"].alignment = center

ws.append(["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"])
for col in range(1, 8):
    c            = ws.cell(row=2, column=col)
    c.font       = bold
    c.fill       = header_fill
    c.border     = border
    c.alignment  = center

row = 3
for idx, (field, _, reason) in enumerate(rules, start=1):
    cnt        = error_counts[field]
    pct_error  = round((cnt / total_records) * 100, 2) if total_records else 0
    pct_health = round(100 - pct_error, 2)
    display_reason = reason if cnt > 0 else ""
    ws.append([idx, field, cnt, total_records, f"{pct_health}%", f"{pct_error}%", display_reason])
    for col in range(1, 8):
        ws.cell(row=row, column=col).border = border
    row += 1

total_errors       = sum(error_counts[field] for field, _, _ in rules)
total_record_count = total_records * len(rules)
total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
total_pct_health   = round(100 - total_pct_error, 2)

ws.append(["", "TOTAL", total_errors, total_record_count,
           f"{total_pct_health}%", f"{total_pct_error}%", ""])
for col in range(1, 8):
    c        = ws.cell(row=row, column=col)
    c.font   = bold
    c.fill   = total_fill
    c.border = border
row += 1

row += 1
for label, value in [
    ("Total Records",        total_records),
    ("Records with Errors",  records_with_errors),
    ("Records Passing",      records_passing),
]:
    ws.cell(row=row, column=1).value = label
    ws.cell(row=row, column=1).font  = bold
    ws.cell(row=row, column=2).value = value
    row += 1

# ======================================================
# RULESETS SHEET
# ======================================================
wsr                = wb.create_sheet("Rulesets")
wsr.merge_cells("A1:C1")
wsr["A1"]           = "HDA – Validation Rules"
wsr["A1"].font      = Font(bold=True, size=14)
wsr["A1"].fill      = title_fill
wsr["A1"].alignment = center

wsr.append(["#", "Field", "Rule Description"])
for col in range(1, 4):
    c           = wsr.cell(row=2, column=col)
    c.font      = bold
    c.fill      = header_fill
    c.border    = border
    c.alignment = center

row = 3
for idx, (field, _, reason) in enumerate(rules, start=1):
    wsr.append([idx, field, reason])
    wsr.cell(row=row, column=2).fill = green_fill
    for col in range(1, 4):
        wsr.cell(row=row, column=col).border = border
    row += 1

# ======================================================
# AUTOFIT ALL SHEETS
# ======================================================
for sheet in wb.sheetnames:
    wsx = wb[sheet]
    for col_idx, col_cells in enumerate(wsx.columns, start=1):
        wsx.column_dimensions[get_column_letter(col_idx)].width = (
            max(len(str(c.value)) if c.value else 0 for c in col_cells) + 3
        )

wb.save(OUTPUT_EXCEL)
print(f"✅ ALL FEATURES INCLUDED — script completed successfully → {OUTPUT_EXCEL}")
