import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====================================================
# File paths
# ====================================================
HDA_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA_Secondary\HDA(SecSales)2026-05-06-1606.tab"
PART_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-05-21-1510.tab"
CUSTOMER_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Cutomer_2026-05-20-1205.tab"
SITE_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-20-1153.tab"
OUTPUT_EXCEL  = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA_Secondary\Validated_HDA_Secondary_Technical2.xlsx"

# ====================================================
# Constants & Styling
# ====================================================
EXCEL_MAX_ROWS = 1_048_000

# Duplicate key columns
DUPLICATE_KEY_COLS = ["DISTRIBUTOR_CODE", "PLANT", "INVOICE_DATE", "CSKU"]

RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL  = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


# ====================================================
# Helper: Universal loader
# ====================================================
def read_input(path: str) -> pd.DataFrame:
    p = path.lower()
    if p.endswith(".csv"):
        return pd.read_csv(path, dtype=str)
    elif p.endswith(".tab"):
        return pd.read_csv(path, sep="\t", dtype=str)
    elif p.endswith((".xlsx", ".xls")):
        return pd.read_excel(path, dtype=str)
    else:
        raise ValueError(f"Unsupported file format: {path}")


# ====================================================
# Load master data
# ====================================================
print("📂 Loading reference files...")

part_df = read_input(PART_FILE)
part_df.columns = part_df.columns.str.strip().str.upper()
part_set = set(part_df["MATERIALNUMBER"].dropna().str.strip())

# Build (CUSTOMER, SUPPLYINGPLANT) combo set from Customer master
customer_df = read_input(CUSTOMER_FILE)
customer_df.columns = customer_df.columns.str.strip().str.upper()
customer_set = set(
    zip(
        customer_df["CUSTOMER"].fillna("").str.strip(),
        customer_df["SUPPLYINGPLANT"].fillna("").str.strip(),
    )
)

site_df = read_input(SITE_FILE)
site_df.columns = site_df.columns.str.strip().str.upper()
site_set = set(site_df["PLANT"].dropna().str.strip())


# ====================================================
# Regex & Mappings
# ====================================================
date_pattern = re.compile(r"^\d{8}$")

ERROR_REASON_MAP = {
    "ERROR_CSKU":             "CSKU: CSKU missing in Part master",
    "ERROR_DISTRIBUTOR_CODE": "DISTRIBUTOR_CODE: DISTRIBUTOR_CODE + PLANT combination does not exist in Customer master (CUSTOMER + SUPPLYINGPLANT)",
    "ERROR_INVOICE_DATE":     "INVOICE_DATE: Invoice week start is blank or not in YYYYMMDD format",
    "ERROR_PLANT":            "PLANT: Plant does not exist in Site master or is blank",
    "ERROR_DUPLICATE":        "DUPLICATE_CHECK: Duplicate record — DISTRIBUTOR_CODE, PLANT, INVOICE_DATE, CSKU combination already exists.",
}

ERROR_COLUMNS = list(ERROR_REASON_MAP.keys())

SUMMARY_RULES = [
    ("CSKU",             "ERROR_CSKU",             ERROR_REASON_MAP["ERROR_CSKU"]),
    ("DISTRIBUTOR_CODE", "ERROR_DISTRIBUTOR_CODE",  ERROR_REASON_MAP["ERROR_DISTRIBUTOR_CODE"]),
    ("INVOICE_DATE",     "ERROR_INVOICE_DATE",       ERROR_REASON_MAP["ERROR_INVOICE_DATE"]),
    ("PLANT",            "ERROR_PLANT",              ERROR_REASON_MAP["ERROR_PLANT"]),
    ("DUPLICATE_CHECK",  "ERROR_DUPLICATE",          ERROR_REASON_MAP["ERROR_DUPLICATE"]),
]

RULESET_DESCRIPTIONS = {
    "CSKU":             ["Must not be blank.", "Must exist as MATERIALNUMBER in Part master."],
    "DISTRIBUTOR_CODE": [
        "Must not be blank.",
        "DISTRIBUTOR_CODE + PLANT combination must exist as CUSTOMER + SUPPLYINGPLANT in Customer master.",
    ],
    "INVOICE_DATE":     ["Must not be blank.", "Must strictly be in YYYYMMDD format."],
    "PLANT":            ["Must not be blank.", "Must exist in Site master."],
    "DUPLICATE_CHECK":  [
        "Must not be blank.",
        "No duplicate combinations of DISTRIBUTOR_CODE + PLANT + INVOICE_DATE + CSKU allowed.",
        "All occurrences of a duplicate group are flagged as errors.",
    ],
}

ATTRIBUTE_SHEETS = {
    "ERROR_CSKU":             ("CSKU",             ERROR_REASON_MAP["ERROR_CSKU"]),
    "ERROR_DISTRIBUTOR_CODE": ("DISTRIBUTOR_CODE",  ERROR_REASON_MAP["ERROR_DISTRIBUTOR_CODE"]),
    "ERROR_INVOICE_DATE":     ("INVOICE_DATE",       ERROR_REASON_MAP["ERROR_INVOICE_DATE"]),
    "ERROR_PLANT":            ("PLANT",              ERROR_REASON_MAP["ERROR_PLANT"]),
    "ERROR_DUPLICATE":        ("DUPLICATE_CHECK",    ERROR_REASON_MAP["ERROR_DUPLICATE"]),
}

attribute_sheet_rows = {
    base: {"sheet_no": 1, "row": 0}
    for base in ["CSKU", "DISTRIBUTOR_CODE", "INVOICE_DATE", "PLANT", "DUPLICATE_CHECK"]
}

error_counts        = {field: 0 for field, _, _ in SUMMARY_RULES}
total_records       = 0
records_with_errors = 0


# ====================================================
# LOAD FULL FILE
# (chunking removed — full load required for duplicate detection)
# ====================================================
print("📂 Loading full HDA Secondary file for validation...")
hda_df = pd.read_csv(HDA_FILE, sep="\t", dtype=str)
hda_df = hda_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
hda_df.columns = hda_df.columns.str.strip().str.upper()

total_records = len(hda_df)
print(f"   → {total_records:,} records loaded.")


# ====================================================
# RUN VALIDATION CHECKS
# ====================================================
print("🔍 Running validation checks...")

hda_df["ERROR_CSKU"] = hda_df["CSKU"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or x not in part_set else ""
)

# DISTRIBUTOR_CODE + PLANT combo must exist in Customer master as CUSTOMER + SUPPLYINGPLANT
hda_df["ERROR_DISTRIBUTOR_CODE"] = hda_df.apply(
    lambda row: "Yes"
    if (
        pd.isna(row["DISTRIBUTOR_CODE"]) or str(row["DISTRIBUTOR_CODE"]).strip() == ""
        or pd.isna(row["PLANT"]) or str(row["PLANT"]).strip() == ""
        or (str(row["DISTRIBUTOR_CODE"]).strip(), str(row["PLANT"]).strip()) not in customer_set
    )
    else "",
    axis=1,
)

hda_df["ERROR_INVOICE_DATE"] = hda_df["INVOICE_DATE"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or not date_pattern.fullmatch(str(x)) else ""
)

hda_df["ERROR_PLANT"] = hda_df["PLANT"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or x not in site_set else ""
)

# Duplicate check: flag ALL occurrences of duplicate key combinations
# (DISTRIBUTOR_CODE + PLANT + INVOICE_DATE + CSKU)
available_dup_cols = [c for c in DUPLICATE_KEY_COLS if c in hda_df.columns]
if len(available_dup_cols) == len(DUPLICATE_KEY_COLS):
    hda_df["ERROR_DUPLICATE"] = hda_df.duplicated(
        subset=available_dup_cols, keep=False
    ).map({True: "Yes", False: ""})
else:
    missing = set(DUPLICATE_KEY_COLS) - set(available_dup_cols)
    print(f"⚠️  Warning: Duplicate check skipped — missing columns: {missing}")
    hda_df["ERROR_DUPLICATE"] = ""


# ====================================================
# COUNT ERRORS & RECORDS WITH ERRORS
# ====================================================
row_has_error = pd.Series(False, index=hda_df.index)
for field, col, _ in SUMMARY_RULES:
    cnt = (hda_df[col] == "Yes").sum()
    error_counts[field] += cnt
    row_has_error = row_has_error | (hda_df[col] == "Yes")

records_with_errors = int(row_has_error.sum())


# ====================================================
# WRITE ERROR SHEETS TO EXCEL
# ====================================================
print("📝 Writing error data to Excel...")

with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:

    for err_col, (base_name, reason_text) in ATTRIBUTE_SHEETS.items():
        error_rows = hda_df[hda_df[err_col] == "Yes"].copy()
        if error_rows.empty:
            continue

        # Drop all error flag columns, then append the human-readable reason
        error_rows = error_rows.drop(columns=ERROR_COLUMNS, errors="ignore")
        error_rows["ERROR_COLUMNS"] = reason_text

        state = attribute_sheet_rows[base_name]
        start = 0

        while start < len(error_rows):
            sheet_name = (
                base_name if state["sheet_no"] == 1
                else f"{base_name}_{state['sheet_no']}"
            )
            remaining = EXCEL_MAX_ROWS - state["row"]
            write_df  = error_rows.iloc[start:start + remaining]

            write_df.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                startrow=state["row"],
                index=False,
                header=(state["row"] == 0),
            )

            start        += len(write_df)
            state["row"] += len(write_df)

            if state["row"] >= EXCEL_MAX_ROWS:
                state["sheet_no"] += 1
                state["row"]       = 0

print("📝 Generating standardized reports...")


# ====================================================
# STYLING & RULESETS
# ====================================================
wb = load_workbook(OUTPUT_EXCEL)
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# ── Summary Sheet ──────────────────────────────────
ws_sum = wb.create_sheet("Summary", 0)

ws_sum.merge_cells("A1:G1")
title_cell           = ws_sum.cell(row=1, column=1, value="HDA Secondary Validation Summary")
title_cell.font      = Font(name="Arial", bold=True, size=14)
title_cell.fill      = TITLE_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 24

headers = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
for c_idx, h in enumerate(headers, start=1):
    cell           = ws_sum.cell(row=2, column=c_idx, value=h)
    cell.fill      = TITLE_FILL
    cell.font      = Font(name="Arial", bold=True)
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_num = 3
for idx, (field, _, reason) in enumerate(SUMMARY_RULES, start=1):
    cnt           = error_counts[field]
    pct_error     = round((cnt / total_records) * 100, 2) if total_records else 0
    pct_health    = round(100 - pct_error, 2)
    display_reason = reason if cnt > 0 else ""

    ws_sum.cell(row=row_num, column=1, value=idx)
    ws_sum.cell(row=row_num, column=2, value=field)
    ws_sum.cell(row=row_num, column=3, value=cnt)
    ws_sum.cell(row=row_num, column=4, value=total_records)
    ws_sum.cell(row=row_num, column=5, value=f"{pct_health}%")
    ws_sum.cell(row=row_num, column=6, value=f"{pct_error}%")
    ws_sum.cell(row=row_num, column=7, value=display_reason)

    for col in range(1, 8):
        c           = ws_sum.cell(row=row_num, column=col)
        c.font      = BODY_FONT
        c.border    = THIN_BORDER
        c.alignment = (
            Alignment(horizontal="left", vertical="center", wrap_text=True)
            if col == 7
            else Alignment(horizontal="center", vertical="center")
        )
    row_num += 1

total_errors       = sum(error_counts[f] for f, _, _ in SUMMARY_RULES)
total_record_count = total_records * len(SUMMARY_RULES)
total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
total_pct_health   = round(100 - total_pct_error, 2)

ws_sum.cell(row=row_num, column=1, value="")
ws_sum.cell(row=row_num, column=2, value="TOTAL")
ws_sum.cell(row=row_num, column=3, value=total_errors)
ws_sum.cell(row=row_num, column=4, value=total_record_count)
ws_sum.cell(row=row_num, column=5, value=f"{total_pct_health}%")
ws_sum.cell(row=row_num, column=6, value=f"{total_pct_error}%")
ws_sum.cell(row=row_num, column=7, value="")

for col in range(1, 8):
    c           = ws_sum.cell(row=row_num, column=col)
    c.font      = Font(name="Arial", bold=True)
    c.fill      = TOTAL_FILL
    c.border    = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

row_num += 2

for label, value in [
    ("Total Records:",        total_records),
    ("Records with Errors:",  records_with_errors),
    ("Records Passing:",      total_records - records_with_errors),
]:
    ws_sum.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
    lc           = ws_sum.cell(row=row_num, column=1, value=label)
    lc.font      = Font(name="Arial", bold=True, size=10)
    lc.fill      = STATS_FILL
    lc.border    = THIN_BORDER
    lc.alignment = Alignment(horizontal="left", vertical="center")

    vc           = ws_sum.cell(row=row_num, column=3, value=value)
    vc.font      = BODY_FONT
    vc.border    = THIN_BORDER
    vc.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1

for col in ws_sum.columns:
    length = max((len(str(c.value)) if c.value else 0) for c in col)
    ws_sum.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 3, 10), 60)


# ── Ruleset Sheet ──────────────────────────────────
wsr = wb.create_sheet("Rule_Set", 1)

wsr.merge_cells("A1:C1")
title_cell           = wsr.cell(row=1, column=1, value="HDA(Secondary) – Validation Rules")
title_cell.font      = Font(name="Arial", bold=True, size=13)
title_cell.fill      = TITLE_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
wsr.row_dimensions[1].height = 22

for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
    cell           = wsr.cell(row=3, column=c_idx, value=h)
    cell.fill      = HDR_FILL
    cell.font      = HDR_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

current_row = 4
rule_num    = 1

for field, rules_list in RULESET_DESCRIPTIONS.items():
    num_rules = len(rules_list)

    for r_idx, rule_text in enumerate(rules_list):
        num_cell           = wsr.cell(row=current_row, column=1, value=rule_num if r_idx == 0 else "")
        num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
        num_cell.fill      = RULE_FILL
        num_cell.border    = THIN_BORDER
        num_cell.alignment = Alignment(horizontal="center", vertical="center")

        field_cell           = wsr.cell(row=current_row, column=2, value=field if r_idx == 0 else "")
        field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
        field_cell.fill      = RULE_FILL
        field_cell.border    = THIN_BORDER
        field_cell.alignment = Alignment(horizontal="center", vertical="center")

        desc_cell           = wsr.cell(row=current_row, column=3, value=rule_text)
        desc_cell.font      = BODY_FONT
        desc_cell.border    = THIN_BORDER
        desc_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        current_row += 1

    if num_rules > 1:
        s = current_row - num_rules
        e = current_row - 1
        wsr.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
        wsr.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

    rule_num += 1

wsr.column_dimensions["A"].width = 6
wsr.column_dimensions["B"].width = 30
wsr.column_dimensions["C"].width = 65


# ── Restyle Data Sheets ────────────────────────────
attribute_base_names = set(attribute_sheet_rows.keys())

attr_sheets_in_wb = []
for sname in wb.sheetnames:
    if sname in ("Summary", "Rule_Set"):
        continue
    for base in attribute_base_names:
        if sname == base or sname.startswith(base + "_"):
            attr_sheets_in_wb.append((sname, base))
            break

for sheet_name, base_col in attr_sheets_in_wb:
    ws = wb[sheet_name]

    highlight_col_idx = None
    for cell in ws[1]:
        if cell.value == base_col:
            highlight_col_idx = cell.column

    max_row = ws.max_row
    max_col = ws.max_column

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.fill      = HDR_FILL
                cell.font      = HDR_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = THIN_BORDER
            else:
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if highlight_col_idx is not None and col_idx == highlight_col_idx:
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT
                else:
                    cell.fill = ROW_FILL

    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 3, 10), 60)

    ws.freeze_panes = "A2"

wb.save(OUTPUT_EXCEL)
print(f"✅ Final processing complete. Saved to: {OUTPUT_EXCEL}")
