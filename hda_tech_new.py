import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  (all inputs are .tab)
# ─────────────────────────────────────────────
HDA_INPUT_FILE       = r"C:\Users\SW526XH\Downloads\Go Live-2\HDA Primary Sales\HDA_Primary_Sales.tab"
PART_INPUT_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-2\HDA Primary Sales\Part.tab"
SITE_INPUT_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-2\HDA Primary Sales\Site.tab"
CUSTOMER_INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-2\HDA Primary Sales\Customer.tab"
OUTPUT_FILE          = r"C:\Users\SW526XH\Downloads\Go Live-2\HDA Primary Sales\Validated_HDA_Primary_Sales.xlsx"


# ─────────────────────────────────────────────
#  DATE FORMAT
# ─────────────────────────────────────────────
DATE_PATTERN = re.compile(r'^\d{4}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$')

# Composite key used for the duplicate check
DUP_KEY_COLS = ["MATERIAL", "PLANT", "SOLDTOPARTY", "BILLINGDOCUMENTDATE"]

# Above this many error rows, per-field sheets get lightweight styling
# (header + red-highlighted error cell only) instead of full cell-by-cell
# font/border/fill, since that per-cell styling is the slow part of
# writing a 600k-row sheet in openpyxl.
FULL_STYLE_ROW_LIMIT = 50_000


# ─────────────────────────────────────────────
#  Colours / Styles
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUB_FILL        = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_ALIGN  = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─────────────────────────────────────────────
#  FIELD ORDER & METADATA
# ─────────────────────────────────────────────
FIELD_ORDER = ["MATERIAL", "PLANT", "SOLDTOPARTY", "BILLINGDOCUMENTDATE", "DUPLICATE_CHECK"]

# Fields that use sub-rows in the summary (blank vs. invalid-value breakdown)
FIELDS_WITH_SUB_ROWS = {"MATERIAL", "PLANT", "SOLDTOPARTY", "BILLINGDOCUMENTDATE"}

# Per-field single-line reason shown in summary (blank for sub-row fields)
FIELD_REASON = {
    "MATERIAL":             "",
    "PLANT":                "",
    "SOLDTOPARTY":          "",
    "BILLINGDOCUMENTDATE":  "",
    "DUPLICATE_CHECK":      "DUPLICATE_CHECK: Duplicate MATERIAL-PLANT-SOLDTOPARTY-BILLINGDOCUMENTDATE combination",
}

# (label, subcount key, single-line reason) used to render the indented
# sub-rows under each multi-rule field in the Summary sheet
SUB_ROW_LABELS = {
    "MATERIAL": [
        ("  ↳ Blank Material",          "blank", "MATERIAL: Field is blank"),
        ("  ↳ Not in Part Master",      "other", "MATERIAL: Not present in Part master"),
    ],
    "PLANT": [
        ("  ↳ Blank Plant Code",        "blank", "PLANT: Field is blank"),
        ("  ↳ Not in Site Master",      "other", "PLANT: Plant code not found in the Site master"),
    ],
    "SOLDTOPARTY": [
        ("  ↳ Blank Sold-To Party",     "blank", "SOLDTOPARTY: Field is blank"),
        ("  ↳ Not in Customer Master",  "other", "SOLDTOPARTY: Customer not found in the Customer master"),
    ],
    "BILLINGDOCUMENTDATE": [
        ("  ↳ Blank Billing Document Date", "blank", "BILLINGDOCUMENTDATE: Field is blank"),
        ("  ↳ Invalid Format (not YYYYMMDD)", "other", "BILLINGDOCUMENTDATE: Does not follow required format YYYYMMDD"),
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine  (vectorized — no row-by-row loops)
# ══════════════════════════════════════════════
class HDARuleEngine:

    def __init__(self, part_materials: set, site_plants: set, customer_ids: set):
        self.part_materials = part_materials
        self.site_plants    = site_plants
        self.customer_ids   = customer_ids

    @staticmethod
    def _clean(series: pd.Series) -> pd.Series:
        return series.fillna("").astype(str).str.strip()

    # ── MATERIAL ──────────────────────────────
    def validate_material(self, df: pd.DataFrame):
        val    = self._clean(df["MATERIAL"])
        blank  = val == ""
        invalid = (~blank) & (~val.str.upper().isin(self.part_materials))
        reason = np.select(
            [blank, invalid],
            ["MATERIAL: Field is blank",
             "MATERIAL: '" + val + "' is not present in Part master"],
            default="",
        )
        return pd.Series(reason, index=df.index), blank, invalid

    # ── PLANT ─────────────────────────────────
    def validate_plant(self, df: pd.DataFrame):
        val    = self._clean(df["PLANT"])
        blank  = val == ""
        invalid = (~blank) & (~val.isin(self.site_plants))
        reason = np.select(
            [blank, invalid],
            ["PLANT: Field is blank",
             "PLANT: '" + val + "' is not present in the Site master"],
            default="",
        )
        return pd.Series(reason, index=df.index), blank, invalid

    # ── SOLDTOPARTY ───────────────────────────
    def validate_soldtoparty(self, df: pd.DataFrame):
        val    = self._clean(df["SOLDTOPARTY"])
        blank  = val == ""
        invalid = (~blank) & (~val.str.upper().isin(self.customer_ids))
        reason = np.select(
            [blank, invalid],
            ["SOLDTOPARTY: Field is blank",
             "SOLDTOPARTY: '" + val + "' is not present in Customer master"],
            default="",
        )
        return pd.Series(reason, index=df.index), blank, invalid

    # ── BILLINGDOCUMENTDATE ───────────────────
    def validate_billingdocumentdate(self, df: pd.DataFrame):
        val    = self._clean(df["BILLINGDOCUMENTDATE"])
        blank  = val == ""
        valid_format = val.str.match(DATE_PATTERN)
        invalid = (~blank) & (~valid_format)
        reason = np.select(
            [blank, invalid],
            ["BILLINGDOCUMENTDATE: Field is blank",
             "BILLINGDOCUMENTDATE: '" + val + "' does not follow the required format YYYYMMDD"],
            default="",
        )
        return pd.Series(reason, index=df.index), blank, invalid

    # ── DUPLICATE_CHECK ───────────────────────
    def validate_duplicate(self, df: pd.DataFrame):
        key_df = pd.DataFrame({c: self._clean(df[c]) for c in DUP_KEY_COLS}, index=df.index)
        dup_mask = key_df.duplicated(keep=False)
        reason = np.where(dup_mask, FIELD_REASON["DUPLICATE_CHECK"], "")
        return pd.Series(reason, index=df.index), dup_mask


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class HDATableValidator:

    def __init__(self, hda_path: str, part_path: str, site_path: str, customer_path: str):
        self.hda_path      = hda_path
        self.part_path     = part_path
        self.site_path     = site_path
        self.customer_path = customer_path

        self.df               = pd.DataFrame()
        self.part_materials   = set()
        self.site_plants      = set()
        self.customer_ids     = set()

        self.field_reasons        = {}   # field -> pd.Series of reason strings
        self.field_masks          = {}   # field -> boolean Series (True = error)
        self.sub_masks            = {}   # field -> {"blank": mask, "other": mask}
        self.combined_error_series = pd.Series(dtype=str)

    def load(self):
        read_kwargs = dict(sep="\t", dtype=str, na_filter=False, engine="c")

        self.df = pd.read_csv(self.hda_path, **read_kwargs)
        self.df.columns = [c.strip().upper() for c in self.df.columns]
        print(f"    HDA Primary Sales rows loaded : {len(self.df):,}")

        for col in DUP_KEY_COLS:
            if col not in self.df.columns:
                raise ValueError(f"{col} column not found in HDA Primary Sales extract.")

        part_df = pd.read_csv(self.part_path, **read_kwargs)
        part_df.columns = [c.strip().upper() for c in part_df.columns]
        if "MATERIALNUMBER" not in part_df.columns:
            raise ValueError("MATERIALNUMBER column not found in Part master.")
        self.part_materials = set(part_df["MATERIALNUMBER"].str.strip().str.upper())
        print(f"    Part master materials loaded  : {len(self.part_materials):,} unique values")

        site_df = pd.read_csv(self.site_path, **read_kwargs)
        site_df.columns = [c.strip().upper() for c in site_df.columns]
        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")
        self.site_plants = set(site_df["PLANT"].str.strip())
        print(f"    Site master plants loaded     : {len(self.site_plants):,} unique values")

        cust_df = pd.read_csv(self.customer_path, **read_kwargs)
        cust_df.columns = [c.strip().upper() for c in cust_df.columns]
        if "CUSTOMER" not in cust_df.columns:
            raise ValueError("CUSTOMER column not found in Customer master.")
        self.customer_ids = set(cust_df["CUSTOMER"].str.strip().str.upper())
        print(f"    Customer master ids loaded    : {len(self.customer_ids):,} unique values")

    def validate(self):
        engine = HDARuleEngine(self.part_materials, self.site_plants, self.customer_ids)

        mat_reason,  mat_blank,  mat_invalid  = engine.validate_material(self.df)
        plt_reason,  plt_blank,  plt_invalid  = engine.validate_plant(self.df)
        cust_reason, cust_blank, cust_invalid = engine.validate_soldtoparty(self.df)
        date_reason, date_blank, date_invalid = engine.validate_billingdocumentdate(self.df)
        dup_reason,  dup_mask                 = engine.validate_duplicate(self.df)

        self.field_reasons = {
            "MATERIAL":            mat_reason,
            "PLANT":               plt_reason,
            "SOLDTOPARTY":         cust_reason,
            "BILLINGDOCUMENTDATE": date_reason,
            "DUPLICATE_CHECK":     dup_reason,
        }
        self.sub_masks = {
            "MATERIAL":            {"blank": mat_blank,  "other": mat_invalid},
            "PLANT":               {"blank": plt_blank,  "other": plt_invalid},
            "SOLDTOPARTY":         {"blank": cust_blank, "other": cust_invalid},
            "BILLINGDOCUMENTDATE": {"blank": date_blank, "other": date_invalid},
        }
        self.field_masks = {f: (r != "") for f, r in self.field_reasons.items()}

        # Combine per-field reasons into one ERROR_COLUMNS string per row,
        # built with vectorized numpy string ops (no python-level row loop).
        combined = pd.Series("", index=self.df.index)
        for field in FIELD_ORDER:
            reason = self.field_reasons[field]
            sep = np.where((combined != "") & (reason != ""), " | ", "")
            combined = combined + sep + reason
        self.combined_error_series = combined

    def get_errors_by_field(self) -> dict:
        return {f: self.df.index[mask] for f, mask in self.field_masks.items() if mask.any()}

    def get_subcounts(self, field: str) -> dict:
        m = self.sub_masks[field]
        return {"blank": int(m["blank"].sum()), "other": int(m["other"].sum())}


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class HDAReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIAL": [
            "Must not be blank.",
            "Must be present in the Part master (MATERIALNUMBER column).",
        ],
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Site master (PLANT column).",
        ],
        "SOLDTOPARTY": [
            "Must not be blank.",
            "Must be present in the Customer master (CUSTOMER column).",
        ],
        "BILLINGDOCUMENTDATE": [
            "Must not be blank.",
            "Must follow the format: YYYYMMDD.",
        ],
        "DUPLICATE_CHECK": [
            "The combination of MATERIAL, PLANT, SOLDTOPARTY and BILLINGDOCUMENTDATE "
            "must be unique across the extract.",
        ],
    }

    def __init__(self, validator: HDATableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                 bold: bool = False, fill: PatternFill = None,
                                 italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    # ══════════════════════════════════════════
    #  Summary sheet
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, total_rows: int):
        v = self.validator

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="HDA Primary Sales Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = int(v.field_masks[col_name].sum())
            has_errors = count > 0

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            reason_text = "" if col_name in FIELDS_WITH_SUB_ROWS else (
                FIELD_REASON.get(col_name, "") if has_errors else ""
            )

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            self._style_summary_data_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1

            # ── generic sub-rows for multi-rule fields ──
            if col_name in FIELDS_WITH_SUB_ROWS and has_errors:
                subcounts = v.get_subcounts(col_name)
                for sub_label, sub_key, sub_reason in SUB_ROW_LABELS[col_name]:
                    sub_count      = subcounts[sub_key]
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1)
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True)
                    row_num += 1

            field_num += 1

        # ── TOTAL row ──
        total_errors       = sum(int(v.field_masks[f].sum()) for f in FIELD_ORDER)
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # ── Quick-glance stats block ──
        records_with_errors = int((v.combined_error_series != "").sum())
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        col_widths = [6, 42, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v = self.validator
        field_errors = v.get_errors_by_field()
        columns      = list(df.columns) + ["ERROR_COLUMNS"]  # business validator: show ALL source columns

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices]
            n_rows = len(subset)

            self._write_header(ws, columns)

            # Bulk value write — far faster than ws.cell(...) per cell for large sheets
            error_col_values = v.combined_error_series.loc[row_indices].to_numpy()
            data_matrix = subset.to_numpy()
            for row_vals, err_val in zip(data_matrix, error_col_values):
                ws.append(list(row_vals) + [err_val])

            # Column(s) to highlight red for this field's error rows
            if field_name == "DUPLICATE_CHECK":
                highlight_cols = [columns.index(c) + 1 for c in DUP_KEY_COLS if c in columns]
            else:
                highlight_cols = [columns.index(field_name) + 1] if field_name in columns else []

            full_style = n_rows <= FULL_STYLE_ROW_LIMIT

            if full_style:
                for r_idx in range(2, n_rows + 2):
                    for c_idx in range(1, len(columns) + 1):
                        cell           = ws.cell(row=r_idx, column=c_idx)
                        cell.font      = BODY_FONT
                        cell.fill      = WHITE_FILL
                        cell.border    = THIN_BORDER
                        cell.alignment = BODY_ALIGN
                    for hc in highlight_cols:
                        hcell      = ws.cell(row=r_idx, column=hc)
                        hcell.fill = RED_FILL
                        hcell.font = ERR_FONT

                for col in ws.columns:
                    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
            else:
                # Lightweight styling for very large sheets: only highlight the
                # offending cell(s); skip per-cell font/border/alignment to
                # keep write time reasonable on 100k+ row sheets.
                for r_idx in range(2, n_rows + 2):
                    for hc in highlight_cols:
                        hcell      = ws.cell(row=r_idx, column=hc)
                        hcell.fill = RED_FILL
                        hcell.font = ERR_FONT
                for c_idx, col_name in enumerate(columns, start=1):
                    ws.column_dimensions[get_column_letter(c_idx)].width = min(max(len(str(col_name)) + 4, 14), 40)

            ws.freeze_panes = "A2"

            note_row = n_rows + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {n_rows:,}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)
            if not full_style:
                ws.cell(
                    row=note_row + 1, column=1,
                    value="(Lightweight styling applied — row count exceeded the full-style threshold.)",
                ).font = Font(name="Arial", italic=True, size=8)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="HDA Primary Sales – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 75

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        records_with_errors = int((v.combined_error_series != "").sum())
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df):,}")
        print(f"   Error rows    : {records_with_errors:,}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class HDATableProcessor:

    def __init__(self, hda_path: str, part_path: str, site_path: str,
                 customer_path: str, output_path: str):
        self.validator = HDATableValidator(hda_path, part_path, site_path, customer_path)
        self.writer    = HDAReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    HDA Primary Sales columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules (vectorized) …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = HDATableProcessor(
        hda_path      = HDA_INPUT_FILE,
        part_path     = PART_INPUT_FILE,
        site_path     = SITE_INPUT_FILE,
        customer_path = CUSTOMER_INPUT_FILE,
        output_path   = OUTPUT_FILE,
    )
    processor.run()
