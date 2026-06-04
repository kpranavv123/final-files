"""
Independent Demand Business Validator
═════════════════════════════════════════════════════════════════════════════

Purpose:
  Validates Independent Demand file against business rules only.

Business Rules:
  - SCHEDULELINEORDERQUANTITY should not contain negative values
  - NETPRICE should not contain negative values
  - REQUESTEDDELIVERYDATE: year-month must not be earlier than
    (BASE_DATE year-month  minus 2 months).
    The day component of BASE_DATE is ignored — only year+month matter.
  - SDPROCESSSTATUS: if value is "C", then REQUESTEDQTYINBASEUNIT must
    equal DELIVEREDQTYINBASEUNIT.

Output:
  - Validated_IndependentDemand_Business.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as _date
import re


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
INDEPENDENT_DEMAND_INPUT = r"C:\Users\SW526XH\Downloads\Go Live-1\ID\Independent Demand_2026-05-20-1754.tab"
OUTPUT_FILE              = r"C:\Users\SW526XH\Downloads\Go Live-1\ID\Validated_IndependentDemand_Business2.xlsx"


# ─────────────────────────────────────────────
#  BASE DATE CONFIGURATION
# ─────────────────────────────────────────────
BASE_DATE = None


def _resolve_base_date() -> _date:
    return BASE_DATE if BASE_DATE is not None else _date.today()


def _cutoff_year_month() -> tuple[int, int]:
    base   = _resolve_base_date()
    year   = base.year
    month  = base.month - 2

    if month <= 0:
        month += 12
        year  -= 1

    return year, month


def _cutoff_label() -> str:
    y, m = _cutoff_year_month()
    month_name = _date(y, m, 1).strftime("%B %Y")
    return f"{y}-{m:02d} ({month_name})"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Business Ruleset Info
# ══════════════════════════════════════════════
BUSINESS_RULESET_INFO = {
    "SCHEDULELINEORDERQUANTITY": [
        "SCHEDULELINEORDERQUANTITY / REQUESTEDQTYINBASEUNIT contains negative value",
    ],
    "NETPRICE": [
        "NETPRICE contains negative value",
    ],
    "REQUESTEDDELIVERYDATE": [
        "REQUESTEDDELIVERYDATE year-month is more than 2 months before the base date",
    ],
    "SDPROCESSSTATUS": [
        "When SDPROCESSSTATUS is 'C', REQUESTEDQTYINBASEUNIT must equal DELIVEREDQTYINBASEUNIT",
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class IndependentDemandBusinessRuleEngine:

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_quantity_negative(self, row) -> tuple:
        val = row.get("SCHEDULELINEORDERQUANTITY")

        if pd.isna(val):
            val = row.get("REQUESTEDQTYINBASEUNIT")

        if self._is_blank(val):
            return True, ""

        try:
            if float(val) < 0:
                return (
                    False,
                    "SCHEDULELINEORDERQUANTITY / REQUESTEDQTYINBASEUNIT contains negative value",
                )
        except (ValueError, TypeError):
            pass

        return True, ""

    def validate_netprice_negative(self, row) -> tuple:
        val = row.get("NETPRICE")

        if self._is_blank(val):
            return True, ""

        try:
            if float(val) < 0:
                return False, "NETPRICE contains negative value"
        except (ValueError, TypeError):
            pass

        return True, ""

    def validate_requested_delivery_date_range(self, row) -> tuple:
        val = row.get("REQUESTEDDELIVERYDATE")

        if self._is_blank(val):
            return True, ""

        val_str = str(val).strip()

        if not re.match(r"^\d{8}$", val_str):
            return True, ""

        try:
            year  = int(val_str[:4])
            month = int(val_str[4:6])
        except ValueError:
            return True, ""

        if not (1 <= month <= 12):
            return True, ""

        cutoff_year, cutoff_month = _cutoff_year_month()

        if (year, month) < (cutoff_year, cutoff_month):
            base_str   = _resolve_base_date().strftime("%Y-%m-%d")
            cutoff_str = _cutoff_label()
            return (
                False,
                f"REQUESTEDDELIVERYDATE {val_str[:4]}-{val_str[4:6]} is before the "
                f"2-month cutoff {cutoff_str} (base date: {base_str})",
            )

        return True, ""

    def validate_sdprocessstatus_completed(self, row) -> tuple:
        """
        If SDPROCESSSTATUS is 'C', then REQUESTEDQTYINBASEUNIT must equal
        DELIVEREDQTYINBASEUNIT. Both values are compared as floats to avoid
        string formatting mismatches (e.g. '10' vs '10.0').

        Skips the check when:
          - SDPROCESSSTATUS is blank or not 'C'
          - Either quantity column is blank or non-numeric
        """
        status = row.get("SDPROCESSSTATUS")

        if self._is_blank(status) or str(status).strip().upper() != "C":
            return True, ""

        requested  = row.get("REQUESTEDQTYINBASEUNIT")
        delivered  = row.get("DELIVEREDQTYINBASEUNIT")

        if self._is_blank(requested) or self._is_blank(delivered):
            return (
                False,
                "SDPROCESSSTATUS is 'C' but REQUESTEDQTYINBASEUNIT or "
                "DELIVEREDQTYINBASEUNIT is blank",
            )

        try:
            req_val = float(requested)
            del_val = float(delivered)
        except (ValueError, TypeError):
            return (
                False,
                "SDPROCESSSTATUS is 'C' but quantity values are non-numeric "
                f"(REQUESTEDQTYINBASEUNIT={requested}, DELIVEREDQTYINBASEUNIT={delivered})",
            )

        if req_val != del_val:
            return (
                False,
                f"SDPROCESSSTATUS is 'C' but REQUESTEDQTYINBASEUNIT ({req_val}) "
                f"!= DELIVEREDQTYINBASEUNIT ({del_val})",
            )

        return True, ""

    def get_rules(self) -> dict:
        return {
            "SCHEDULELINEORDERQUANTITY": self.validate_quantity_negative,
            "NETPRICE":                  self.validate_netprice_negative,
            "REQUESTEDDELIVERYDATE":     self.validate_requested_delivery_date_range,
            "SDPROCESSSTATUS":           self.validate_sdprocessstatus_completed,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class IndependentDemandBusinessValidator:

    def __init__(self):
        self.df        = pd.DataFrame()
        self.error_map = {}

    def load(self):
        print("[LOAD] Loading files...")

        if INDEPENDENT_DEMAND_INPUT.lower().endswith(".xlsx"):
            self.df = pd.read_excel(INDEPENDENT_DEMAND_INPUT, dtype=str)
        else:
            self.df = pd.read_csv(INDEPENDENT_DEMAND_INPUT, sep="\t", dtype=str)

        print(f"    Independent Demand - rows loaded: {len(self.df)}")

        self.df.columns = [str(c).strip().upper() for c in self.df.columns]

        if (
            "SCHEDULELINEORDERQUANTITY" not in self.df.columns
            and "REQUESTEDQTYINBASEUNIT" in self.df.columns
        ):
            self.df["SCHEDULELINEORDERQUANTITY"] = self.df["REQUESTEDQTYINBASEUNIT"]

    def validate(self):
        print("[VALIDATE] Running business validation rules...")
        print(f"    Base date         : {_resolve_base_date().strftime('%Y-%m-%d')}")
        print(f"    2-month cutoff    : {_cutoff_label()}")

        engine = IndependentDemandBusinessRuleEngine()
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            errors = {}

            for col_name, rule_fn in rules.items():
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {str(e)}"

                if not passed:
                    errors[col_name] = reason

            if errors:
                self.error_map[idx] = errors

    def get_error_series(self) -> pd.Series:
        error_details = {}
        for idx, error_dict in self.error_map.items():
            error_details[idx] = "; ".join([f"{f}: {r}" for f, r in error_dict.items()])
        return pd.Series(error_details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class IndependentDemandBusinessReportWriter:
    SHEET_SUMMARY  = "Summary"
    SHEET_RULESETS = "Rulesets"

    def __init__(self, validator: IndependentDemandBusinessValidator, output_path: str):
        self.validator             = validator
        self.output_path           = output_path
        self._summary_fields_order = []
        self.base_date_str         = _resolve_base_date().strftime("%d-%b-%Y")
        self.cutoff_label_str      = _cutoff_label()

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)
        for ch in invalid_chars:
            name = name.replace(ch, "-")
        name = name.strip() or "Sheet"
        name = name[:31]
        if name not in wb.sheetnames:
            return name
        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _get_ruleset_columns(self):
        ruleset_fields = [
            "SALESORDERITEM",
            "SALESORDER",
            "PRODUCTIONPLANT",
            "SALESORDERTYPE",
            "SOLDTOPARTY",
            "REQUESTEDDELIVERYDATE",
            "MATERIAL",
            "SCHEDULELINEORDERQUANTITY",
            "REQUESTEDQTYINBASEUNIT",
            "DELIVEREDQTYINBASEUNIT",
            "NETPRICE",
            "SDPROCESSSTATUS",
        ]
        ruleset_columns = [col for col in ruleset_fields if col in self.validator.df.columns]
        ruleset_columns.append("ERROR_FIELDS")
        return ruleset_columns

    # ── Rulesets sheet ────────────────────────
    def _write_ruleset_sheet(self, wb, summary_fields=None):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        title_cell           = ws.cell(
            row=1, column=1,
            value="Independent Demand - Business Validation Rules",
        )
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:C1")

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        ruleset_info = {
            "SCHEDULELINEORDERQUANTITY": [
                "No negative values.",
            ],
            "NETPRICE": [
                "No negative values.",
            ],
            "REQUESTEDDELIVERYDATE": [
                "Transactions beyond current - 2 months should not be present",
            ],
            "SDPROCESSSTATUS": [
                "When SDPROCESSSTATUS is 'C' (Completed), REQUESTEDQTYINBASEUNIT "
                "must equal DELIVEREDQTYINBASEUNIT. A mismatch indicates the order "
                "was marked complete but quantities do not reconcile.",
            ],
        }

        ordered_fields = summary_fields if summary_fields else list(ruleset_info.keys())
        current_row    = 4
        rule_num       = 1

        for field in ordered_fields:
            if field not in ruleset_info:
                continue

            rules_list = ruleset_info[field]
            num_rules  = len(rules_list)

            for r_idx, rule_desc in enumerate(rules_list):
                num_cell           = ws.cell(
                    row=current_row, column=1,
                    value=rule_num if r_idx == 0 else "",
                )
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(
                    row=current_row, column=2,
                    value=field if r_idx == 0 else "",
                )
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_desc)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 85
        ws.row_dimensions[1].height     = 22

    # ── Summary sheet ─────────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        title_cell = ws.cell(
            row=1, column=1,
            value=(
                f"Independent Demand Business Rules Summary  "
                f"(Base date: {self.base_date_str}  |  "
                f"2-month cutoff: {self.cutoff_label_str})"
            ),
        )
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:G1")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count",
                      "% Health", "% of Error", "Reason"]
        col_widths = [6, 35, 16, 16, 16, 16, 80]

        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ruleset_field_order = [
            "SCHEDULELINEORDERQUANTITY",
            "NETPRICE",
            "REQUESTEDDELIVERYDATE",
            "SDPROCESSSTATUS",
        ]

        col_error_counts  = {field: 0 for field in ruleset_field_order}
        rule_error_counts = {}

        for bad_cols in error_map.values():
            for col, reason in bad_cols.items():
                if col in col_error_counts:
                    col_error_counts[col] += 1
                else:
                    col_error_counts[col] = 1
                rule_error_counts[(col, reason)] = rule_error_counts.get((col, reason), 0) + 1

        sorted_fields = sorted(
            [(field, col_error_counts.get(field, 0)) for field in ruleset_field_order],
            key=lambda x: x[1],
            reverse=True,
        )

        self._summary_fields_order = [f for f, _ in sorted_fields]

        row_num    = 4
        total_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

        for field_num, (col_name, count) in enumerate(sorted_fields, start=1):
            error_percent  = count / total_rows if total_rows > 0 else 0
            health_percent = 1 - error_percent

            canonical_reasons = {
                "SCHEDULELINEORDERQUANTITY": (
                    "SCHEDULELINEORDERQUANTITY / REQUESTEDQTYINBASEUNIT contains negative value"
                ),
                "NETPRICE": "NETPRICE contains negative value",
                "REQUESTEDDELIVERYDATE": (
                    f"REQUESTEDDELIVERYDATE year-month is before the 2-month cutoff "
                    f"{self.cutoff_label_str}"
                ),
                "SDPROCESSSTATUS": (
                    "SDPROCESSSTATUS is 'C' but REQUESTEDQTYINBASEUNIT != DELIVEREDQTYINBASEUNIT"
                ),
            }
            reason_text = canonical_reasons.get(col_name, "") if count > 0 else ""

            ws.cell(row=row_num, column=1, value=field_num).font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font  = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font     = BODY_FONT
            ws.cell(row=row_num, column=4, value=total_rows).font = BODY_FONT

            cell_health               = ws.cell(row=row_num, column=5, value=health_percent)
            cell_health.font          = BODY_FONT
            cell_health.number_format = "0.00%"

            cell_pct               = ws.cell(row=row_num, column=6, value=error_percent)
            cell_pct.font          = BODY_FONT
            cell_pct.number_format = "0.00%"

            ws.cell(row=row_num, column=7, value=reason_text).font = BODY_FONT

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(
                    horizontal="center" if c != 7 else "left",
                    wrap_text=(c == 7),
                )

            row_num += 1

        # ── TOTAL row ───────────────────────
        total_errors      = sum(col_error_counts.values())
        sum_record_counts = len(sorted_fields) * total_rows
        total_error_pct   = total_errors / sum_record_counts if sum_record_counts > 0 else 0
        total_health_pct  = 1 - total_error_pct

        ws.cell(row=row_num, column=2, value="TOTAL").font           = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font      = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=sum_record_counts).font = Font(name="Arial", bold=True)

        cell_th               = ws.cell(row=row_num, column=5, value=total_health_pct)
        cell_th.font          = Font(name="Arial", bold=True)
        cell_th.number_format = "0.00%"

        cell_tp               = ws.cell(row=row_num, column=6, value=total_error_pct)
        cell_tp.font          = Font(name="Arial", bold=True)
        cell_tp.number_format = "0.00%"

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).fill      = total_fill
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(
                horizontal="center" if c != 7 else "left"
            )

        row_num += 2

        # ── Stats block ──────────────────────
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors
        stats_label_fill    = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = stats_label_fill
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=2)

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v                = self.validator
        all_error_fields = set()

        for err_dict in v.error_map.values():
            all_error_fields.update(err_dict.keys())

        order             = self._summary_fields_order if self._summary_fields_order else list(all_error_fields)
        fields_to_process = [f for f in order if f in all_error_fields]

        for field in fields_to_process:
            row_indices = [idx for idx, err in v.error_map.items() if field in err]

            if not row_indices:
                continue

            sheet_name = self._safe_sheet_name(wb, field)
            ws         = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()
            subset["ERROR_FIELDS"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(field, "")
            )

            for c_idx, col_name in enumerate(subset.columns, start=1):
                cell           = ws.cell(row=1, column=c_idx, value=col_name)
                cell.fill      = HDR_FILL
                cell.font      = HDR_FONT
                cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 20

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell      = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font = BODY_FONT
                    cell.fill = ROW_FILL

                # For SDPROCESSSTATUS rule, highlight both quantity columns + status
                if field == "SDPROCESSSTATUS":
                    for highlight_col in ["SDPROCESSSTATUS", "REQUESTEDQTYINBASEUNIT", "DELIVEREDQTYINBASEUNIT"]:
                        if highlight_col in col_idx_map:
                            target_cell      = ws.cell(row=excel_row, column=col_idx_map[highlight_col])
                            target_cell.fill = RED_FILL
                            target_cell.font = ERR_FONT
                elif field in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT
                elif "ERROR_FIELDS" in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map["ERROR_FIELDS"])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    # ── Main write entry point ────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series       = v.get_error_series()
        df["ERROR_FIELDS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        ruleset_columns = self._get_ruleset_columns()
        filtered_cols   = [col for col in df.columns if col in ruleset_columns]
        df              = df[filtered_cols]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, len(df))
        self._write_ruleset_sheet(wb, summary_fields=self._summary_fields_order)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        print(f"\n[SAVE] Business output saved  : {self.output_path}")
        print(f"       Base date              : {self.base_date_str}")
        print(f"       2-month cutoff         : {self.cutoff_label_str}")
        print(f"       Error rows             : {len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class IndependentDemandBusinessProcessor:
    def __init__(self):
        self.validator = IndependentDemandBusinessValidator()
        self.writer    = IndependentDemandBusinessReportWriter(self.validator, OUTPUT_FILE)

    def run(self):
        print("=" * 70)
        print("Independent Demand Business Validation")
        print(f"Base date      : {_resolve_base_date().strftime('%d-%b-%Y')}")
        print(f"2-month cutoff : {_cutoff_label()}")
        print("=" * 70)

        self.validator.load()
        self.validator.validate()
        self.writer.write()

        print("\n" + "=" * 70)
        print("Business Validation Complete!")
        print("=" * 70)


if __name__ == "__main__":
    op = IndependentDemandBusinessProcessor()
    op.run()
