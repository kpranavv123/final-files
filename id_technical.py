"""
Independent Demand Technical Validator,updated
═════════════════════════════════════════════════════════════════════════════
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
INDEPENDENT_DEMAND_INPUT = r"C:\Users\SW526XH\Downloads\Go Live-1\ID\Independent Demand_2026-05-20-1754.tab"
PART_REFERENCE_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-06-04-1737.tab"
SITE_REFERENCE_FILE      = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-20-1153.tab"
CUSTOMER_REFERENCE_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\CustomerHierarchy_updated_2026-06-05-0958 1.tab"
OUTPUT_FILE              = r"C:\Users\SW526XH\Downloads\Go Live-1\ID\Validated_IndependentDemand_Technical2.xlsx"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Composite key columns for duplicate detection
DUPLICATE_KEY_COLS = ["SALESORDERITEM","SALESORDER", "PRODUCTIONPLANT", "SALESORDERTYPE", "SOLDTOPARTY", "MATERIAL"]


# ══════════════════════════════════════════════
#  Technical Ruleset Info
# ══════════════════════════════════════════════
TECHNICAL_RULESET_INFO = {
    "SALESORDERITEM": [
        "SALESORDERITEM is blank",
    ],
    "SALESORDER": [
        "SALESORDER is blank",
    ],
    "PRODUCTIONPLANT": [
        "PRODUCTIONPLANT is blank",
        "PRODUCTIONPLANT not found in Site Master",
    ],
    "SALESORDERTYPE": [
        "SALESORDERTYPE is blank",
    ],
    "SOLDTOPARTY": [
        "SOLDTOPARTY is blank",
        # "SOLDTOPARTY and PRODUCTIONPLANT combination not found in Customer master",  ← DISABLED
    ],
    "REQUESTEDDELIVERYDATE": [
        "REQUESTEDDELIVERYDATE is blank",
        "REQUESTEDDELIVERYDATE does not follow YYYYMMDD format",
        "REQUESTEDDELIVERYDATE is not a valid date",
    ],
    "MATERIAL": [
        "MATERIAL is blank",
        "MATERIAL and PRODUCTIONPLANT combination not found in Part Master",
    ],
    "SCHEDULELINEORDERQUANTITY": [
        "SCHEDULELINEORDERQUANTITY / REQUESTEDQTYINBASEUNIT is blank",
    ],
    "NETPRICE": [
        "NETPRICE is blank",
    ],
    "SDPROCESSSTATUS": [
        "SDPROCESSSTATUS is blank",
    ],
    "DUPLICATE_CHECK": [
        "Duplicate row: SALESORDER + PRODUCTIONPLANT + SALESORDERTYPE + SOLDTOPARTY + MATERIAL combination appears more than once in the extract",
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class IndependentDemandTechnicalRuleEngine:

    def __init__(self, site_codes: set, part_site_combos: set,
                 customer_site_combos: set, duplicate_keys: set):
        self.site_codes           = site_codes
        self.part_site_combos     = part_site_combos
        self.customer_site_combos = customer_site_combos
        self.duplicate_keys       = duplicate_keys

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_salesorderitem(self, row) -> tuple:
        if self._is_blank(row.get("SALESORDERITEM")):
            return False, "SALESORDERITEM is blank"
        return True, ""

    def validate_salesorder(self, row) -> tuple:
        if self._is_blank(row.get("SALESORDER")):
            return False, "SALESORDER is blank"
        return True, ""

    def validate_productionplant(self, row) -> tuple:
        val = row.get("PRODUCTIONPLANT")
        if self._is_blank(val):
            return False, "PRODUCTIONPLANT is blank"
        if str(val).strip() in self.site_codes:
            return True, ""
        return False, "PRODUCTIONPLANT not found in Site Master"

    def validate_salesordertype(self, row) -> tuple:
        if self._is_blank(row.get("SALESORDERTYPE")):
            return False, "SALESORDERTYPE is blank"
        return True, ""

    def validate_soldtoparty(self, row) -> tuple:
        cust_id = row.get("SOLDTOPARTY")
        if self._is_blank(cust_id):
            return False, "SOLDTOPARTY is blank"
        # ── Rule disabled: site/customer master cross-check ──────────────
        # cust_str  = str(cust_id).strip()
        # prod_plant = row.get("PRODUCTIONPLANT")
        # plant_str = str(prod_plant).strip() if not self._is_blank(prod_plant) else ""
        # if (cust_str, plant_str) in self.customer_site_combos:
        #     return True, ""
        # return False, "SOLDTOPARTY and PRODUCTIONPLANT combination not found in Customer master"
        # ─────────────────────────────────────────────────────────────────
        return True, ""

    def validate_req_del_date(self, row) -> tuple:
        import re
        val = row.get("REQUESTEDDELIVERYDATE")
        if self._is_blank(val):
            return False, "REQUESTEDDELIVERYDATE is blank"
        val_str = str(val).strip()
        if not re.match(r"^\d{8}$", val_str):
            return False, "REQUESTEDDELIVERYDATE does not follow YYYYMMDD format"
        try:
            datetime.strptime(val_str, "%Y%m%d")
        except ValueError:
            return False, "REQUESTEDDELIVERYDATE is not a valid date"
        return True, ""

    def validate_material(self, row) -> tuple:
        material   = row.get("MATERIAL")
        prod_plant = row.get("PRODUCTIONPLANT")
        if self._is_blank(material):
            return False, "MATERIAL is blank"
        mat_str   = str(material).strip()
        plant_str = str(prod_plant).strip() if not self._is_blank(prod_plant) else ""
        if (mat_str, plant_str) in self.part_site_combos:
            return True, ""
        return False, "MATERIAL and PRODUCTIONPLANT combination not found in Part Master"

    def validate_quantity_blank(self, row) -> tuple:
        val = row.get("SCHEDULELINEORDERQUANTITY")
        if pd.isna(val):
            val = row.get("REQUESTEDQTYINBASEUNIT")
        if self._is_blank(val):
            return False, "SCHEDULELINEORDERQUANTITY / REQUESTEDQTYINBASEUNIT is blank"
        return True, ""

    def validate_netprice_blank(self, row) -> tuple:
        if self._is_blank(row.get("NETPRICE")):
            return False, "NETPRICE is blank"
        return True, ""

    def validate_status(self, row) -> tuple:
        if self._is_blank(row.get("SDPROCESSSTATUS")):
            return False, "SDPROCESSSTATUS is blank"
        return True, ""

    def validate_duplicate_check(self, row) -> tuple:
        key = tuple(
            str(row.get(col, "")).strip() for col in DUPLICATE_KEY_COLS
        )
        if all(v and v != "nan" for v in key):
            if key in self.duplicate_keys:
                return False, (
                    "Duplicate row: SALESORDER + PRODUCTIONPLANT + SALESORDERTYPE "
                    "+ SOLDTOPARTY + MATERIAL combination appears more than once in the extract"
                )
        return True, ""

    def get_rules(self) -> dict:
        return {
            "SALESORDERITEM":            self.validate_salesorderitem,
            "SALESORDER":                self.validate_salesorder,
            "PRODUCTIONPLANT":           self.validate_productionplant,
            "SALESORDERTYPE":            self.validate_salesordertype,
            "SOLDTOPARTY":               self.validate_soldtoparty,
            "REQUESTEDDELIVERYDATE":     self.validate_req_del_date,
            "MATERIAL":                  self.validate_material,
            "SCHEDULELINEORDERQUANTITY": self.validate_quantity_blank,
            "NETPRICE":                  self.validate_netprice_blank,
            "SDPROCESSSTATUS":           self.validate_status,
            "DUPLICATE_CHECK":           self.validate_duplicate_check,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class IndependentDemandTechnicalValidator:

    def __init__(self):
        self.df                   = pd.DataFrame()
        self.site_codes           = set()
        self.part_site_combos     = set()
        self.customer_site_combos = set()
        self.error_map            = {}

    def _read_file(self, path: str) -> pd.DataFrame:
        p = path.lower()
        if p.endswith(".csv"):
            return pd.read_csv(path, dtype=str)
        elif p.endswith((".tab", ".tsv")):
            return pd.read_csv(path, sep="\t", dtype=str,
                               encoding="utf-8", encoding_errors="replace", engine="python")
        elif p.endswith((".xlsx", ".xls")):
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {path}")

    def load(self):
        print("[LOAD] Loading files...")

        # Independent Demand
        self.df = self._read_file(INDEPENDENT_DEMAND_INPUT)
        self.df.columns = [str(c).strip().upper() for c in self.df.columns]
        print(f"    Independent Demand - rows loaded: {len(self.df)}")

        if ("SCHEDULELINEORDERQUANTITY" not in self.df.columns
                and "REQUESTEDQTYINBASEUNIT" in self.df.columns):
            self.df["SCHEDULELINEORDERQUANTITY"] = self.df["REQUESTEDQTYINBASEUNIT"]

        # Site master
        site_df = self._read_file(SITE_REFERENCE_FILE)
        site_df.columns = [str(c).strip().upper() for c in site_df.columns]
        if "PLANT" not in site_df.columns:
            raise ValueError(f"PLANT column missing in Site master! Found: {list(site_df.columns)}")
        self.site_codes = set(site_df["PLANT"].dropna().str.strip().tolist())

        # Part master
        part_df = self._read_file(PART_REFERENCE_FILE)
        part_df.columns = [str(c).strip().upper() for c in part_df.columns]
        mat_col = (
            "MATERIALNUMBER" if "MATERIALNUMBER" in part_df.columns
            else ("PRODUCT" if "PRODUCT" in part_df.columns else "MATERIAL")
        )
        for _, row in part_df.iterrows():
            part_name = str(row[mat_col]).strip() if pd.notna(row.get(mat_col)) else ""
            part_site = str(row.get("PLANT", "")).strip() if pd.notna(row.get("PLANT")) else ""
            if part_name and part_site:
                self.part_site_combos.add((part_name, part_site))

        # Customer master (loaded but not used for validation — cross-check disabled)
        cust_df = self._read_file(CUSTOMER_REFERENCE_FILE)
        cust_df.columns = [str(c).strip().upper() for c in cust_df.columns]
        for _, row in cust_df.iterrows():
            cust_id = str(row.get("CUSTOMER", "")).strip() if pd.notna(row.get("CUSTOMER")) else ""
            site    = str(row.get("SUPPLYINGPLANT", "")).strip() if pd.notna(row.get("SUPPLYINGPLANT")) else ""
            if cust_id and site:
                self.customer_site_combos.add((cust_id, site))

    def validate(self):
        print("[VALIDATE] Running technical validation rules...")

        # Pre-compute composite duplicates
        present_key_cols = [c for c in DUPLICATE_KEY_COLS if c in self.df.columns]

        if len(present_key_cols) == len(DUPLICATE_KEY_COLS):
            combo = self.df[DUPLICATE_KEY_COLS].fillna("").apply(
                lambda row: "|||".join(row.str.strip()), axis=1
            )
            empty_sentinel = "|||".join([""] * len(DUPLICATE_KEY_COLS))
            dup_mask       = combo.duplicated(keep=False) & (combo != empty_sentinel)
            duplicate_keys = set(
                tuple(v.split("|||")) for v in combo[dup_mask].tolist()
            )
        else:
            duplicate_keys = set()

        engine = IndependentDemandTechnicalRuleEngine(
            self.site_codes,
            self.part_site_combos,
            self.customer_site_combos,
            duplicate_keys,
        )
        rules = engine.get_rules()

        for idx, row in self.df.iterrows():
            errors = {}
            for col_name, rule_fn in rules.items():
                if col_name != "DUPLICATE_CHECK" and col_name not in self.df.columns:
                    continue
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {str(e)}"
                if not passed:
                    errors[col_name] = reason
            if errors:
                self.error_map[idx] = errors

    def get_error_series(self) -> pd.Series:
        error_details = {}
        for idx, error_dict in self.error_map.items():
            error_details[idx] = "; ".join([f"{f}: {r}" for f, r in error_dict.items()])
        return pd.Series(error_details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class IndependentDemandTechnicalReportWriter:
    SHEET_SUMMARY  = "Summary"
    SHEET_RULESETS = "Rulesets"

    def __init__(self, validator: IndependentDemandTechnicalValidator, output_path: str):
        self.validator             = validator
        self.output_path           = output_path
        self._summary_fields_order = []

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)
        for ch in invalid_chars:
            name = name.replace(ch, "-")
        name = name.strip() or "Sheet"
        name = name[:31]
        if name not in wb.sheetnames:
            return name
        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _get_ruleset_columns(self):
        ruleset_fields = [
            "SALESORDERITEM", "SALESORDER", "PRODUCTIONPLANT", "SALESORDERTYPE",
            "SOLDTOPARTY", "REQUESTEDDELIVERYDATE", "MATERIAL",
            "SCHEDULELINEORDERQUANTITY", "NETPRICE", "SDPROCESSSTATUS",
        ]
        cols = [col for col in ruleset_fields if col in self.validator.df.columns]
        cols.append("ERROR_FIELDS")
        return cols

    def _write_ruleset_sheet(self, wb, summary_fields=None):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1,
                                       value="Independent Demand - Technical Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        ruleset_info = {
            "SALESORDERITEM":   ["Field should not be blank."],
            "SALESORDER":       ["Field should not be blank."],
            "PRODUCTIONPLANT":  [
                "Field should not be blank.",
                "To be present in the Site master.",
            ],
            "SALESORDERTYPE":   ["Field should not be blank."],
            "SOLDTOPARTY":      [
                "Field should not be blank.",
                # "Site customer combination to be present in customer master.",  ← DISABLED
            ],
            "REQUESTEDDELIVERYDATE": [
                "Field should not be blank.",
                "Field should follow the format: YYYYMMDD.",
                "Field should be a valid date.",
            ],
            "MATERIAL": [
                "Field should not be blank.",
                "Part site combination to be present in the Part master.",
            ],
            "SCHEDULELINEORDERQUANTITY": ["Field should not be blank."],
            "NETPRICE":         ["Field should not be blank."],
            "SDPROCESSSTATUS":  ["Field should not be blank."],
            "DUPLICATE_CHECK":  [
                "The combination of SALESORDER + PRODUCTIONPLANT + SALESORDERTYPE "
                "+ SOLDTOPARTY + MATERIAL must be unique across the entire extract. "
                "If the same values appear together in more than one row, all such rows are flagged.",
            ],
        }

        ordered_fields = summary_fields if summary_fields else list(ruleset_info.keys())
        current_row    = 4
        rule_num       = 1

        for field in ordered_fields:
            if field not in ruleset_info:
                continue

            rules_list = ruleset_info[field]
            num_rules  = len(rules_list)

            for r_idx, rule_desc in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_desc)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 85

    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        ws.merge_cells("A1:E1")
        title_cell           = ws.cell(row=1, column=1,
                                       value="Independent Demand Technical Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count",
                      "% Health", "% of Error", "Reason"]
        col_widths = [6, 35, 16, 16, 16, 16, 80]

        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ruleset_field_order = [
            "SALESORDERITEM", "SALESORDER", "PRODUCTIONPLANT", "SALESORDERTYPE",
            "SOLDTOPARTY", "REQUESTEDDELIVERYDATE", "MATERIAL",
            "SCHEDULELINEORDERQUANTITY", "NETPRICE", "SDPROCESSSTATUS",
            "DUPLICATE_CHECK",
        ]

        col_error_counts  = {field: 0 for field in ruleset_field_order}
        rule_error_counts = {}

        for bad_cols in error_map.values():
            for col, reason in bad_cols.items():
                if col in col_error_counts:
                    col_error_counts[col] += 1
                else:
                    col_error_counts[col] = 1
                rule_error_counts[(col, reason)] = rule_error_counts.get((col, reason), 0) + 1

        sorted_fields = sorted(
            [(field, col_error_counts.get(field, 0)) for field in ruleset_field_order],
            key=lambda x: x[1],
            reverse=True,
        )
        self._summary_fields_order = [f for f, _ in sorted_fields]

        row_num = 4

        for field_num, (col_name, count) in enumerate(sorted_fields, start=1):
            error_percent  = count / total_rows if total_rows > 0 else 0
            health_percent = 1 - error_percent

            if col_name == "DUPLICATE_CHECK":
                reason_text = (
                    "Duplicate row: SALESORDER + PRODUCTIONPLANT + SALESORDERTYPE "
                    "+ SOLDTOPARTY + MATERIAL combination appears more than once in the extract"
                ) if count > 0 else ""
            else:
                actual_reasons = [r for (f, r) in rule_error_counts.keys() if f == col_name]
                unique_reasons = sorted(set(actual_reasons),
                                        key=actual_reasons.count, reverse=True)
                reason_text = unique_reasons[0] if unique_reasons else ""

            ws.cell(row=row_num, column=1, value=field_num).font  = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font   = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font      = BODY_FONT
            ws.cell(row=row_num, column=4, value=total_rows).font = BODY_FONT

            cell_health              = ws.cell(row=row_num, column=5, value=health_percent)
            cell_health.font         = BODY_FONT
            cell_health.number_format = "0.00%"

            cell_pct              = ws.cell(row=row_num, column=6, value=error_percent)
            cell_pct.font         = BODY_FONT
            cell_pct.number_format = "0.00%"

            ws.cell(row=row_num, column=7, value=reason_text).font = BODY_FONT

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(
                    horizontal="center" if c != 7 else "left",
                    wrap_text=(c == 7),
                )

            row_num += 1

        # TOTAL row
        total_errors         = sum(col_error_counts.values())
        total_fill           = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        sum_record_counts    = len(sorted_fields) * total_rows
        total_error_percent  = total_errors / sum_record_counts if sum_record_counts > 0 else 0
        total_health_percent = 1 - total_error_percent

        ws.cell(row=row_num, column=2, value="TOTAL").font      = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=sum_record_counts).font = Font(name="Arial", bold=True)

        cell_th              = ws.cell(row=row_num, column=5, value=total_health_percent)
        cell_th.font         = Font(name="Arial", bold=True)
        cell_th.number_format = "0.00%"

        cell_tp              = ws.cell(row=row_num, column=6, value=total_error_percent)
        cell_tp.font         = Font(name="Arial", bold=True)
        cell_tp.number_format = "0.00%"

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).fill      = total_fill
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(
                horizontal="center" if c != 7 else "left"
            )

        row_num += 2

        # Stats block
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors
        stats_label_fill    = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = stats_label_fill
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v = self.validator
        all_error_fields = set()
        for err_dict in v.error_map.values():
            all_error_fields.update(err_dict.keys())

        order             = self._summary_fields_order if self._summary_fields_order else list(all_error_fields)
        fields_to_process = [f for f in order if f in all_error_fields]

        for field in fields_to_process:
            row_indices = [idx for idx, err in v.error_map.items() if field in err]
            if not row_indices:
                continue

            sheet_name = self._safe_sheet_name(wb, field)
            ws         = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()
            subset["ERROR_FIELDS"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(field, "")
            )

            for c_idx, col_name in enumerate(subset.columns, start=1):
                cell           = ws.cell(row=1, column=c_idx, value=col_name)
                cell.fill      = HDR_FILL
                cell.font      = HDR_FONT
                cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 20

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell      = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font = BODY_FONT
                    cell.fill = ROW_FILL

                if field == "DUPLICATE_CHECK":
                    for key_col in DUPLICATE_KEY_COLS:
                        if key_col in col_idx_map:
                            cell      = ws.cell(row=excel_row, column=col_idx_map[key_col])
                            cell.fill = RED_FILL
                            cell.font = ERR_FONT
                else:
                    if field in col_idx_map:
                        target_cell      = ws.cell(row=excel_row, column=col_idx_map[field])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 20

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series       = v.get_error_series()
        df["ERROR_FIELDS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        ruleset_columns = self._get_ruleset_columns()
        filtered_cols   = [col for col in df.columns if col in ruleset_columns]
        df              = df[filtered_cols]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, len(df))
        self._write_ruleset_sheet(wb, summary_fields=self._summary_fields_order)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        print(f"\n[SAVE] Technical output saved: {self.output_path}")
        print(f"       Error rows: {len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class IndependentDemandTechnicalProcessor:
    def __init__(self):
        self.validator = IndependentDemandTechnicalValidator()
        self.writer    = IndependentDemandTechnicalReportWriter(self.validator, OUTPUT_FILE)

    def run(self):
        print("=" * 70)
        print("Independent Demand Technical Validation")
        print("=" * 70)
        self.validator.load()
        self.validator.validate()
        self.writer.write()
        print("\n" + "=" * 70)
        print("Technical Validation Complete!")
        print("=" * 70)


if __name__ == "__main__":
    op = IndependentDemandTechnicalProcessor()
    op.run()
