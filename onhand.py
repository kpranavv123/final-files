import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
ONHAND_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\OnHand\Onhand_updated.tab"
SITE_INPUT_FILE   = r"C:\Users\SW526XH\Downloads\Go Live-2\OnHand\Site_2026-04-09-1058.xlsx"
OUTPUT_FILE       = r"C:\Users\SW526XH\Downloads\Go Live-2\Onhand\Validated_Onhand.xlsx"


# ─────────────────────────────────────────────
#  VALID TYPE VALUES
# ─────────────────────────────────────────────
VALID_TYPE_VALUES = {
    "VALUATEDUNRESTRICTEDUSESTOCK",
    "STOCKINTRANSFER",
    "QUALITYINSPECTION",
    "RESTRICTEDBATCHES",
    "BLOCKEDSTOCK",
    "BLOCKEDSTOCKRETURNS",
    "SCHEDULEDFORDELIVERY",
}

MATERIAL_RANGE_1_MIN = 14_000_000_000_000
MATERIAL_RANGE_1_MAX = 14_999_999_999_999
MATERIAL_RANGE_2_MIN = 15_000_000_000_000
MATERIAL_RANGE_2_MAX = 15_999_999_999_999

KEEP_COLS = [
    "MATERIALNUMBER", "MATERIALTYPE", "BATCHNUMBER", "PLANT",
    "STORAGELOCATION", "WAREHOUSENUMBER", "DATEOFLASTGOODSRECEIPT",
    "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE", "TYPE", "QTY", "STANDARDPRICE",
]

# ─────────────────────────────────────────────
#  Colours / Styles  — matched to Site template
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUB_FILL        = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Canonical field order
FIELD_ORDER = [
    "MATERIALNUMBER", "MATERIALTYPE", "BATCHNUMBER", "PLANT",
    "STORAGELOCATION", "WAREHOUSENUMBER", "DATEOFLASTGOODSRECEIPT",
    "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE", "TYPE", "QTY", "STANDARDPRICE",
]

# Fields that use sub-rows in the summary (parent reason cell left blank)
FIELDS_WITH_SUB_ROWS = {"MATERIALNUMBER", "PLANT"}

# ─────────────────────────────────────────────
#  Per-field single-line reason shown in summary
# ─────────────────────────────────────────────
FIELD_REASON = {
    "MATERIALNUMBER":                        "",   # sub-rows carry reasons
    # "MATERIALTYPE":                          "",   # no validation rules
    "BATCHNUMBER":                           "BATCHNUMBER: Field is blank ",
    "PLANT":                                 "",   # sub-rows carry reasons
    "STORAGELOCATION":                       "STORAGELOCATION: Field is blank",
    "WAREHOUSENUMBER":                       "WAREHOUSENUMBER: Field is blank ",
    "DATEOFLASTGOODSRECEIPT":               "DATEOFLASTGOODSRECEIPT: Field is blank",
    "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE": "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE: Field is blank",
    "TYPE":                                  "TYPE: Field is blank or not one of the 7 valid stock type values",
    "QTY":                                   "QTY: Field is blank",
    "STANDARDPRICE":                         "STANDARDPRICE: Field is blank",
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class OnhandRuleEngine:

    def __init__(self, site_plants: set):
        self.site_plants = set(str(p).strip() for p in site_plants)

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_materialnumber(self, row) -> str:
        val = row.get("MATERIALNUMBER", "")
        if self._is_blank(val):
            return "MATERIALNUMBER: Field is blank"
        try:
            num = int(str(val).strip())
        except ValueError:
            return f"MATERIALNUMBER: '{str(val).strip()}' is not a valid numeric material number"
        in_range1 = MATERIAL_RANGE_1_MIN <= num <= MATERIAL_RANGE_1_MAX
        in_range2 = MATERIAL_RANGE_2_MIN <= num <= MATERIAL_RANGE_2_MAX
        if in_range1 or in_range2:
            return ""
        return (
            f"MATERIALNUMBER: '{str(val).strip()}' is out of valid range "
            "(must be 14000000000000–14999999999999 OR 15000000000000–15999999999999)"
        )

    # def validate_materialtype(self, row) -> str:
    #     # No validation rules defined for this field
    #     return ""

    def validate_batchnumber(self, row) -> str:
        if self._is_blank(row.get("BATCHNUMBER")):
            return "BATCHNUMBER: Field is blank"
        return ""

    def validate_plant(self, row) -> str:
        val = str(row.get("PLANT", "")).strip()
        if not val or val == "nan":
            return "PLANT: Field is blank"
        if val not in self.site_plants:
            return f"PLANT: '{val}' is not present in the Site master"
        return ""

    def validate_storagelocation(self, row) -> str:
        if self._is_blank(row.get("STORAGELOCATION")):
            return "STORAGELOCATION: Field is blank"
        return ""

    def validate_warehousenumber(self, row) -> str:
        if self._is_blank(row.get("WAREHOUSENUMBER")):
            return "WAREHOUSENUMBER: Field is blank"
        return ""

    def validate_dateoflastgoodsreceipt(self, row) -> str:
        if self._is_blank(row.get("DATEOFLASTGOODSRECEIPT")):
            return "DATEOFLASTGOODSRECEIPT: Field is blank "
        return ""

    def validate_shelflife(self, row) -> str:
        if self._is_blank(row.get("SHELFLIFEEXPIRATION_DATEOFMANUFACTURE")):
            return "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE: Field is blank"
        return ""

    def validate_type(self, row) -> str:
        val = row.get("TYPE", "")
        if self._is_blank(val):
            return "TYPE: Field is blank"
        if str(val).strip().upper() not in VALID_TYPE_VALUES:
            return (
                f"TYPE: '{str(val).strip()}' is invalid — must be one of: "
                "VALUATEDUNRESTRICTEDUSESTOCK, STOCKINTRANSFER, QUALITYINSPECTION, "
                "RESTRICTEDBATCHES, BLOCKEDSTOCK, BLOCKEDSTOCKRETURNS, SCHEDULEDFORDELIVERY"
            )
        return ""

    def validate_qty(self, row) -> str:
        if self._is_blank(row.get("QTY")):
            return "QTY: Field is blank"
        return ""

    def validate_standardprice(self, row) -> str:
        if self._is_blank(row.get("STANDARDPRICE")):
            return "STANDARDPRICE: Field is blank"
        return ""

    def get_rules(self) -> dict:
        return {
            "MATERIALNUMBER":                        self.validate_materialnumber,
            # "MATERIALTYPE":                          self.validate_materialtype,
            "BATCHNUMBER":                           self.validate_batchnumber,
            "PLANT":                                 self.validate_plant,
            "STORAGELOCATION":                       self.validate_storagelocation,
            "WAREHOUSENUMBER":                       self.validate_warehousenumber,
            "DATEOFLASTGOODSRECEIPT":               self.validate_dateoflastgoodsreceipt,
            "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE": self.validate_shelflife,
            "TYPE":                                  self.validate_type,
            "QTY":                                   self.validate_qty,
            "STANDARDPRICE":                         self.validate_standardprice,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class OnhandTableValidator:

    def __init__(self, onhand_path: str, site_path: str):
        self.onhand_path = onhand_path
        self.site_path   = site_path
        self.df          = pd.DataFrame()
        self.site_plants = set()
        self.error_map   = {}
        self.reason_map  = {}

    def load(self):
        self.df = pd.read_csv(self.onhand_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        site_df = pd.read_excel(self.site_path, dtype=str,engine="openpyxl")
        site_df.columns = [c.strip().upper() for c in site_df.columns]

        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")

        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site master plants loaded : {len(self.site_plants)} unique values")

    def validate(self):
        engine = OnhandRuleEngine(self.site_plants)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols    = []
            col_reason_map = {}

            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)
                except Exception:
                    reason = f"{col}: Unexpected validation error"

                if reason:
                    failed_cols.append(col)
                    col_reason_map[col] = reason

            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = col_reason_map

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors

    def get_materialnumber_error_subcounts(self) -> dict:
        counts = {"blank": 0, "out_of_range": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("MATERIALNUMBER", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            else:
                counts["out_of_range"] += 1
        return counts

    def get_plant_error_subcounts(self) -> dict:
        counts = {"blank": 0, "not_in_site": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("PLANT", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            else:
                counts["not_in_site"] += 1
        return counts


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class OnhandReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIALNUMBER": [
            "Must not be blank.",
            "Must be in range 14000000000000-14999999999999 OR 15000000000000-15999999999999.",
        ],
        # "MATERIALTYPE": [
        #     "No specific validation rules defined for this field.",
        # ],
        "BATCHNUMBER": [
            "Must not be blank.",
        ],
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Site master (PLANT column).",
        ],
        "STORAGELOCATION": [
            "Must not be blank.",
        ],
        "WAREHOUSENUMBER": [
            "Must not be blank.",
        ],
        "DATEOFLASTGOODSRECEIPT": [
            "Must not be blank.",
        ],
        "SHELFLIFEEXPIRATION_DATEOFMANUFACTURE": [
            "Must not be blank.",
        ],
        "TYPE": [
            "Must not be blank.",
            "Value must be one of: VALUATEDUNRESTRICTEDUSESTOCK, STOCKINTRANSFER, "
            "QUALITYINSPECTION, RESTRICTEDBATCHES, BLOCKEDSTOCK, "
            "BLOCKEDSTOCKRETURNS, SCHEDULEDFORDELIVERY.",
        ],
        "QTY": [
            "Must not be blank.",
        ],
        "STANDARDPRICE": [
            "Must not be blank.",
        ],
    }

    def __init__(self, validator: OnhandTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            # ERROR_COLUMNS header → WHITE fill + black bold font so it's readable
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    # ══════════════════════════════════════════
    #  Summary sheet
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="Onhand Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        # ── Build per-field error counts ──
        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        mat_subcounts   = self.validator.get_materialnumber_error_subcounts()
        plant_subcounts = self.validator.get_plant_error_subcounts()

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            if col_name in FIELDS_WITH_SUB_ROWS:
                reason_text = ""            # sub-rows carry reasons
            elif has_errors:
                reason_text = FIELD_REASON.get(col_name, "")
            else:
                reason_text = ""

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            self._style_summary_data_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1

            # ── MATERIALNUMBER sub-rows ──
            if col_name == "MATERIALNUMBER" and has_errors:
                sub_definitions = [
                    (
                        "  ↳ Blank Material Number",
                        mat_subcounts["blank"],
                        "MATERIALNUMBER: Field is blank",
                    ),
                    (
                        "  ↳ Out of Valid Range (14x / 15x trillion)",
                        mat_subcounts["out_of_range"],
                        "MATERIALNUMBER: Not in range 14000000000000–14999999999999 "
                        "OR 15000000000000–15999999999999",
                    ),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)

                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            # ── PLANT sub-rows ──
            if col_name == "PLANT" and has_errors:
                sub_definitions = [
                    (
                        "  ↳ Blank Plant Code",
                        plant_subcounts["blank"],
                        "PLANT: Field is blank",
                    ),
                    (
                        "  ↳ Not in Site Master",
                        plant_subcounts["not_in_site"],
                        "PLANT: Plant code not found in the Site master",
                    ),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)

                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            field_num += 1

        # ── TOTAL row ──
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2   # blank spacer

        # ── Quick-glance stats block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 40, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        """
        Creates one sheet per field that has at least one error row.
        Sheets follow FIELD_ORDER for consistent tab ordering.
        Error rows use WHITE background; only the failing column cell is RED.
        The ERROR_COLUMNS header column uses WHITE fill so it remains readable.
        """
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            subset    = df.loc[row_indices].copy()
            keep_here = [c for c in KEEP_COLS if c in subset.columns] + ["ERROR_COLUMNS"]
            subset    = subset[keep_here]

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            # Header row (ERROR_COLUMNS column gets WHITE fill)
            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL   # ← white rows (no yellow)
                    cell.border    = THIN_BORDER

                # Only the failing field cell gets RED highlight
                if field_name in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Onhand Table – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 75

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        keep_cols = [c for c in KEEP_COLS if c in df.columns] + ["ERROR_COLUMNS"]
        df        = df[keep_cols]

        # ── Create workbook — first sheet is Summary ──
        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        # Rules sheet
        self._write_rules_sheet(wb)

        # Per-field error sheets (only for fields WITH errors, in FIELD_ORDER)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class OnhandTableProcessor:

    def __init__(self, onhand_path: str, site_path: str, output_path: str):
        self.validator = OnhandTableValidator(onhand_path, site_path)
        self.writer    = OnhandReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    Onhand columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = OnhandTableProcessor(
        onhand_path = ONHAND_INPUT_FILE,
        site_path   = SITE_INPUT_FILE,
        output_path = OUTPUT_FILE,
    )
    processor.run()
