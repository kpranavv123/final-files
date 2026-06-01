import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-05-28-1951.tab"
OUTPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Validated_Part_Business_Rulesets_with_IBPStatus.xlsx"


# ─────────────────────────────────────────────
#  STYLING CONSTANTS  (identical to technical script)
# ─────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FF0000")
ROW_FILL   = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL   = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL  = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
STATS_FILL = PatternFill("solid", fgColor="EDEDED")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  BUSINESS RULE IDENTIFIERS
#  Rule 1 : DUPLICATE_ROW
#  Rule 2 : MATERIALNUMBER  → mapped to single PRODUCTDESCRIPTION
#  Rule 3 : BASEUNIT        → consistent UOM per MATERIALNUMBER across sites
#  Rule 4 : MINREMSHELFLIFE → not blank for FERT/HAWA; must be > 0
#  Rule 5 : PROCUREMENTTYPE → must be E / F / X
# ─────────────────────────────────────────────
# RULE1_KEY = "DUPLICATE_ROW"
RULE2_KEY = "MATERIALNUMBER"
RULE3_KEY = "BASEUNIT"
RULE4_KEY = "MINREMSHELFLIFE"
RULE5_KEY = "PROCUREMENTTYPE"
RULE6_KEY = "IBPSTATUTS"

VALID_PROCUREMENT_TYPES = {"E", "F", "X"}
SHELF_LIFE_MATERIAL_TYPES = {"FERT", "HAWA"}

# Error-sheet creation order
ERROR_SHEET_PRIORITY = [RULE2_KEY, RULE3_KEY, RULE4_KEY, RULE5_KEY,RULE6_KEY]

# Summary / Rules sheet labels
RULES_FIELDS_ORDERED = [RULE2_KEY, RULE3_KEY, RULE4_KEY, RULE5_KEY,RULE6_KEY]

REASON_MAP = {
    # RULE1_KEY: (
    #     "DUPLICATE_ROW: The entire row is an exact duplicate of another row in the extract"
    # ),
    RULE2_KEY: (
        "MATERIALNUMBER: MATERIALNUMBER is mapped to more than one PRODUCTDESCRIPTION"
    ),
    RULE3_KEY: (
        "BASEUNIT: BASEUNIT (Base UOM) is inconsistent for the same MATERIALNUMBER — "
        "a single MATERIALNUMBER must have the same BASEUNIT across all sites"
    ),
    RULE4_KEY: (
        "MINREMSHELFLIFE: Field is blank or ≤ 0 for material type FERT / HAWA — "
        "must be present and greater than zero"
    ),
    RULE5_KEY: (
        "PROCUREMENTTYPE: Invalid or blank value — must be one of E / F / X"
    ),
    
RULE6_KEY: (
        "IBPSTATUS: Must be 'IBP' — blank or unexpected value found"
    ),

}

RULES_CONTENT = {
    # RULE1_KEY: [
    #     "No duplicate rows are allowed in the extract.",
    # ],
    RULE2_KEY: [
        "MATERIALNUMBER column should be mapped to single description.",
        
    ],
    RULE3_KEY: [
        "BASEUNIT (Base UOM) must be consistent for the same MATERIALNUMBER across all sites.",
    ],
    RULE4_KEY: [
        "MINREMSHELFLIFE field must not be blank for material types FERT and HAWA.",
        "MINREMSHELFLIFE must be greater than 0",
    ],
    RULE5_KEY: [
        "Field should be E/F/X",
    ],
    
RULE6_KEY: [
        "Field value must be 'IBP'.",
        "Blank or any other value is treated as an error.",
    ],

}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class BusinessRuleEngine:
    """
    Returns error_map:
        { row_index : { rule_key : { "reason": str, "highlight_cols": [col, ...] } } }
    """

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def run(self, df: pd.DataFrame) -> dict:
        error_map: dict = {}

        # ── Rule 1: Fully identical duplicate rows ──────────────────────────
        # dup_mask = df.duplicated(keep=False)          # marks ALL occurrences
        # for idx in df[dup_mask].index:
        #     error_map.setdefault(idx, {})[RULE1_KEY] = {
        #         "reason":         REASON_MAP[RULE1_KEY],
        #         "highlight_cols": [],                  # whole row is the issue
        #     }

        # ── Rule 2: MaterialNumber → multiple descriptions ──────────────────
        if "MATERIALNUMBER" in df.columns and "PRODUCTDESCRIPTION" in df.columns:
            multi_desc = (
                df.groupby("MATERIALNUMBER")["PRODUCTDESCRIPTION"]
                .nunique()
            )
            bad_materials = set(multi_desc[multi_desc > 1].index)
            for idx, row in df.iterrows():
                mat = str(row.get("MATERIALNUMBER", "")).strip()
                if mat in bad_materials:
                    error_map.setdefault(idx, {})[RULE2_KEY] = {
                        "reason":         REASON_MAP[RULE2_KEY],
                        "highlight_cols": ["MATERIALNUMBER", "PRODUCTDESCRIPTION"],
                    }

        # ── Rule 3: BASEUNIT consistency per MATERIALNUMBER ─────────────────
        if "MATERIALNUMBER" in df.columns and "BASEUNIT" in df.columns:
            multi_uom = (
                df.groupby("MATERIALNUMBER")["BASEUNIT"]
                .nunique()
            )
            bad_uom_materials = set(multi_uom[multi_uom > 1].index)
            for idx, row in df.iterrows():
                mat = str(row.get("MATERIALNUMBER", "")).strip()
                if mat in bad_uom_materials:
                    error_map.setdefault(idx, {})[RULE3_KEY] = {
                        "reason":         REASON_MAP[RULE3_KEY],
                        "highlight_cols": ["MATERIALNUMBER", "BASEUNIT"],
                    }

        # ── Rule 4: MINREMSHELFLIFE — blank / ≤ 0 for FERT and HAWA ────────
        if "MINREMSHELFLIFE" in df.columns and "MATERIALTYPE" in df.columns:
            for idx, row in df.iterrows():
                mat_type = str(row.get("MATERIALTYPE", "")).strip().upper()
                if mat_type not in SHELF_LIFE_MATERIAL_TYPES:
                    continue                           # rule only applies to FERT / HAWA

                val = row.get("MINREMSHELFLIFE")

                if self._is_blank(val):
                    error_map.setdefault(idx, {})[RULE4_KEY] = {
                        "reason":         REASON_MAP[RULE4_KEY],
                        "highlight_cols": ["MINREMSHELFLIFE"],
                    }
                    continue

                try:
                    numeric_val = float(str(val).strip())
                    if numeric_val <= 0:
                        error_map.setdefault(idx, {})[RULE4_KEY] = {
                            "reason":         REASON_MAP[RULE4_KEY],
                            "highlight_cols": ["MINREMSHELFLIFE"],
                        }
                except ValueError:
                    # Non-numeric value in the field
                    error_map.setdefault(idx, {})[RULE4_KEY] = {
                        "reason":         REASON_MAP[RULE4_KEY],
                        "highlight_cols": ["MINREMSHELFLIFE"],
                    }

        # ── Rule 5: PROCUREMENTTYPE — must be E / F / X ─────────────────────
        if "PROCUREMENTTYPE" in df.columns:
            for idx, row in df.iterrows():
                val = row.get("PROCUREMENTTYPE")
                if self._is_blank(val):
                    continue
                is_invalid = str(val).strip().upper() not in VALID_PROCUREMENT_TYPES

                if is_invalid:
                    error_map.setdefault(idx, {})[RULE5_KEY] = {
                        "reason":         REASON_MAP[RULE5_KEY],
                        "highlight_cols": ["PROCUREMENTTYPE"],
                    }
        if "IBPSTATUS" in df.columns:
           for idx, row in df.iterrows():
               val = row.get("IBPSTATUS")

               if self._is_blank(val) or str(val).strip().upper() != "IBP":
                error_map.setdefault(idx, {})[RULE6_KEY] = {
                "reason": REASON_MAP[RULE6_KEY],
                "highlight_cols": ["IBPSTATUS"],
             }
               
        return error_map



# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartBusinessValidator:

    def __init__(self, filepath: str):
        self.filepath  = filepath
        self.df        = pd.DataFrame()
        self.error_map = {}

    def load(self):
        path = self.filepath.lower()
        if path.endswith(".csv"):
            self.df = pd.read_csv(self.filepath, dtype=str)
        elif path.endswith(".tab") or path.endswith(".tsv"):
            self.df = pd.read_csv(
                self.filepath,
                sep="\t",
                dtype=str,
                encoding="utf-8",
                engine="python",
            )
        elif path.endswith(".xlsx") or path.endswith(".xls"):
            self.df = pd.read_excel(self.filepath, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {self.filepath}")

        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        engine         = BusinessRuleEngine()
        self.error_map = engine.run(self.df)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class BusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    def __init__(self, validator: PartBusinessValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── Helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _auto_width(self, ws, min_w=10, max_w=60):
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(length + 3, min_w), max_w
            )

    # ── Summary Sheet ─────────────────────────
    def _write_summary_sheet(self, wb):
        ws         = wb.create_sheet(self.SHEET_SUMMARY)
        v          = self.validator
        total_rows = len(v.df)

        # Count error rows per rule
        col_error_counts = {r: 0 for r in RULES_FIELDS_ORDERED}
        for rule_dict in v.error_map.values():
            for rule_key in rule_dict.keys():
                if rule_key in col_error_counts:
                    col_error_counts[rule_key] += 1

        # Title
        ws.merge_cells("A1:G1")
        tc           = ws.cell(row=1, column=1,
                               value="Part Master FG – Business Rules Validation Summary")
        tc.font      = Font(name="Arial", bold=True, size=14)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Header row
        headers = ["#", "Rule Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        row_num = 3
        for rule_num, rule_key in enumerate(RULES_FIELDS_ORDERED, start=1):
            count      = col_error_counts.get(rule_key, 0)
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason     = REASON_MAP.get(rule_key, "") if count > 0 else ""

            values = [rule_num, rule_key, count, total_rows,
                      f"{pct_health}%", f"{pct_error}%", reason]
            for c_idx, val in enumerate(values, start=1):
                cell           = ws.cell(row=row_num, column=c_idx, value=val)
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.fill      = WHITE_FILL
                cell.alignment = Alignment(
                    horizontal="left" if c_idx == 7 else "center",
                    vertical="center",
                    wrap_text=(c_idx == 7),
                )
            row_num += 1

        # TOTAL row
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(RULES_FIELDS_ORDERED)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c_idx, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""],
            start=1,
        ):
            cell           = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        records_with_errors = len(v.error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=2)
            lc           = ws.cell(row=row_num, column=1, value=label)
            lc.font      = Font(name="Arial", bold=True, size=10)
            lc.fill      = STATS_FILL
            lc.border    = THIN_BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")

            vc           = ws.cell(row=row_num, column=3, value=value)
            vc.font      = BODY_FONT
            vc.border    = THIN_BORDER
            vc.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        col_widths = [6, 24, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Rules Sheet ───────────────────────────
    def _write_ruleset_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        tc           = ws.cell(row=1, column=1,
                               value="Part Master FG – Business Validation Rules")
        tc.font      = Font(name="Arial", bold=True, size=13)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Rule Name", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = 4
        for rule_num, (rule_key, rules_list) in enumerate(RULES_CONTENT.items(), start=1):
            num_rules = len(rules_list)
            for r_idx, rule_text in enumerate(rules_list):
                nc           = ws.cell(row=current_row, column=1,
                                       value=rule_num if r_idx == 0 else "")
                nc.fill      = RULE_FILL
                nc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                nc.border    = THIN_BORDER
                nc.alignment = Alignment(horizontal="center", vertical="center")

                fc           = ws.cell(row=current_row, column=2,
                                       value=rule_key if r_idx == 0 else "")
                fc.fill      = RULE_FILL
                fc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                fc.border    = THIN_BORDER
                fc.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=True)

                dc           = ws.cell(row=current_row, column=3, value=rule_text)
                dc.font      = BODY_FONT
                dc.border    = THIN_BORDER
                dc.alignment = Alignment(wrap_text=True, vertical="center",
                                         horizontal="left")
                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 70

    # ── Error Sheets ──────────────────────────
    def _write_error_sheets(self, wb):
        """
        One sheet per business rule.
        ALL columns from the input file are shown.
        Highlighted columns per rule:
          DUPLICATE_ROW    → no specific column (whole row in yellow)
          MATERIALNUMBER   → MATERIALNUMBER + PRODUCTDESCRIPTION in red
          BASEUNIT         → MATERIALNUMBER + BASEUNIT in red
          MINREMSHELFLIFE  → MINREMSHELFLIFE in red
          PROCUREMENTTYPE  → PROCUREMENTTYPE in red
        """
        v              = self.validator
        df             = v.df
        all_input_cols = list(df.columns)

        for rule_key in ERROR_SHEET_PRIORITY:
            row_indices = [
                idx for idx, rule_dict in v.error_map.items()
                if rule_key in rule_dict
            ]
            if not row_indices:
                continue

            subset = df.loc[row_indices, all_input_cols].copy()
            subset["ERROR_REASON"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(rule_key, {}).get("reason", "")
            )

            # Safe sheet name (Excel max 31 chars)
            sheet_name = rule_key[:31]
            existing   = [s.title for s in wb.worksheets]
            counter    = 1
            base       = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base[:28]}_{counter}"
                counter   += 1

            ws = wb.create_sheet(sheet_name)
            self._write_header(ws, list(subset.columns))

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            # Determine highlight columns for this rule from first error row
            highlight_cols: list = v.error_map[row_indices[0]][rule_key].get(
                "highlight_cols", []
            )

            for r_idx, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for col, value in zip(subset.columns, row_data):
                    c_idx          = col_idx_map[col]
                    cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.fill = ROW_FILL

                # Red overlay on the specific failing column(s)
                for h_col in highlight_cols:
                    if h_col in col_idx_map:
                        tc      = ws.cell(row=r_idx, column=col_idx_map[h_col])
                        tc.fill = RED_FILL
                        tc.font = ERR_FONT

            self._auto_width(ws, min_w=10, max_w=60)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{rule_key}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Orchestrate ───────────────────────────
    def write(self):
        v  = self.validator
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb)
        self._write_ruleset_sheet(wb)
        self._write_error_sheets(wb)

        wb.save(self.output_path)

        # rule1_errors = sum(1 for rd in v.error_map.values() if RULE1_KEY in rd)
        rule2_errors = sum(1 for rd in v.error_map.values() if RULE2_KEY in rd)
        rule3_errors = sum(1 for rd in v.error_map.values() if RULE3_KEY in rd)
        rule4_errors = sum(1 for rd in v.error_map.values() if RULE4_KEY in rd)
        rule5_errors = sum(1 for rd in v.error_map.values() if RULE5_KEY in rd)
        rule6_errors = sum(1 for rd in v.error_map.values() if RULE6_KEY in rd)

        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows                        : {len(v.df)}")
        print(f"   Rows with any error               : {len(v.error_map)}")
        # print(f"   DUPLICATE_ROW errors              : {rule1_errors}")
        print(f"   MATERIALNUMBER errors             : {rule2_errors}")
        print(f"   BASEUNIT errors                   : {rule3_errors}")
        print(f"   MINREMSHELFLIFE errors            : {rule4_errors}")
        print(f"   PROCUREMENTTYPE errors            : {rule5_errors}")
        print(f"   IBPSTATUS errors                : {rule6_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartBusinessProcessor:

    def __init__(self, input_path: str, output_path: str):
        self.validator = PartBusinessValidator(input_path)
        self.writer    = BusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Columns detected : {list(self.validator.df.columns)}")
        print("🔍  Running business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PartBusinessProcessor(INPUT_FILE, OUTPUT_FILE)
    processor.run()
