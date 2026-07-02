import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
PARTSOURCE_MAKE_INPUT_FILE   = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource(Make)\PartSource_Make.tab"
BOM_INPUT_FILE                = r"C:\Users\SW526XH\Downloads\Go Live-2\BOM\BOM.tab"
SOURCE_CONSTRAINT_INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-2\SourceConstraint\SourceConstraint.tab"
OUTPUT_FILE                   = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource(Make)\Validated_PartSourceMake_Business.xlsx"


# ─────────────────────────────────────────────
#  PARTSOURCE(MAKE) COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
PS_MATERIAL_COL = "MATERIAL"
PS_PLANT_COL    = "PLANT"

# ─────────────────────────────────────────────
#  BOM MASTER COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
BOM_MATERIAL_COL = "ROOT_MATERIAL"
BOM_PLANT_COL    = "ROOT_PLANT"

# ─────────────────────────────────────────────
#  SOURCE CONSTRAINT MASTER COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
SC_MATERIAL_COL = "MATERIAL"
SC_PLANT_COL    = "MKALPLANT"


# ─────────────────────────────────────────────
#  Colours / Styles  — matched to existing Business Validator template
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUBROW_FILL     = PatternFill("solid", start_color="F7F7F7", end_color="F7F7F7")

HDR_FONT     = Font(bold=True, name="Arial")
BODY_FONT    = Font(name="Arial", size=10)
ERR_FONT     = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SUBROW_FONT  = Font(name="Arial", size=10, italic=True)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  FIELD ORDER  (only fields with business rules)
# ─────────────────────────────────────────────
FIELD_ORDER = ["MAXIMUMLOTSIZE", "MINIMUMLOTSIZE", "ROUNDINGVALUE", "MATERIAL"]

# Single-line reason shown in Summary sheet for the simple numeric fields
FIELD_REASON = {
    "MAXIMUMLOTSIZE": "MAXIMUMLOTSIZE: Value not greater than 0",
    "MINIMUMLOTSIZE": "MINIMUMLOTSIZE: Value not greater than 0",
    "ROUNDINGVALUE":  "ROUNDINGVALUE: Value not greater than 0",
}

# Sub-category reasons shown as indented rows under MATERIAL in Summary
MATERIAL_SUB_CATEGORIES = ["Blank", "Not in BOM", "Not in Source Constraint"]


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class PartSourceMakeBusinessRuleEngine:
    """
    Rules:
      MAXIMUMLOTSIZE  - must be > 0.
      MINIMUMLOTSIZE  - must be > 0.
      ROUNDINGVALUE   - must be > 0.
      MATERIAL
        - The MATERIAL / PLANT (Part-Site) combination must be present in the
          BOM extract as ROOT_MATERIAL / ROOT_PLANT, AND must also be present
          in the Source Constraint extract as MATERIAL / MKALPLANT.
    All numeric rules are independent of each other. MATERIAL is independent
    of the numeric fields but cross-references two separate master sets.
    """

    def __init__(self, valid_bom_combos: set, valid_sc_combos: set):
        self.valid_bom_combos = valid_bom_combos
        self.valid_sc_combos  = valid_sc_combos

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    @staticmethod
    def _parse_number(value):
        s = str(value).strip()
        if s == "" or s.lower() == "nan":
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _validate_positive_field(self, row, field_name: str) -> str:
        """Returns a reason string if field_name is present but not > 0, else ''."""
        raw = row.get(field_name, "")
        num = self._parse_number(raw)
        if num is not None and num <= 0:
            return f"{field_name}: '{str(raw).strip()}' is not greater than 0"
        return ""

    def _validate_material(self, row) -> str:
        """Returns a combined reason string for MATERIAL, or '' if it passes."""
        mat_raw   = row.get(PS_MATERIAL_COL, "")
        plant_raw = row.get(PS_PLANT_COL, "")
        mat_blank   = self._is_blank(mat_raw)
        plant_blank = self._is_blank(plant_raw)

        if mat_blank or plant_blank:
            missing = []
            if mat_blank:
                missing.append(PS_MATERIAL_COL)
            if plant_blank:
                missing.append(PS_PLANT_COL)
            return f"MATERIAL: Blank field(s) - {', '.join(missing)}"

        mat   = str(mat_raw).strip()
        plant = str(plant_raw).strip()
        key   = f"{mat}|{plant}"

        problems = []
        if key not in self.valid_bom_combos:
            problems.append("not found in BOM extract (ROOT_MATERIAL/ROOT_PLANT)")
        if key not in self.valid_sc_combos:
            problems.append("not found in Source Constraint extract (MATERIAL/MKALPLANT)")

        if problems:
            return f"MATERIAL: Part-Site combination '{mat}-{plant}' " + "; ".join(problems)

        return ""

    def validate_row(self, row) -> dict:
        """Returns {field_name: reason} for whichever fields failed on this row."""
        reasons = {}

        for field_name in ("MAXIMUMLOTSIZE", "MINIMUMLOTSIZE", "ROUNDINGVALUE"):
            reason = self._validate_positive_field(row, field_name)
            if reason:
                reasons[field_name] = reason

        material_reason = self._validate_material(row)
        if material_reason:
            reasons["MATERIAL"] = material_reason

        return reasons


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartSourceMakeBusinessTableValidator:

    def __init__(self, ps_make_path: str, bom_path: str, source_constraint_path: str):
        self.ps_make_path           = ps_make_path
        self.bom_path               = bom_path
        self.source_constraint_path = source_constraint_path

        self.df               = pd.DataFrame()
        self.valid_bom_combos = set()
        self.valid_sc_combos  = set()
        self.error_map        = {}   # row_idx -> [failed field names]
        self.reason_map       = {}   # row_idx -> {field: reason}

        # MATERIAL sub-category counters for Summary sheet breakdown
        self.material_blank_count      = 0
        self.material_not_in_bom_count = 0
        self.material_not_in_sc_count  = 0

    def load(self):
        self.df = pd.read_csv(self.ps_make_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        bom_df = pd.read_csv(self.bom_path, sep="\t", dtype=str)
        bom_df.columns = [c.strip().upper() for c in bom_df.columns]
        for _, r in bom_df.iterrows():
            mat   = str(r.get(BOM_MATERIAL_COL, "")).strip()
            plant = str(r.get(BOM_PLANT_COL, "")).strip()
            if mat == "" or plant == "":
                continue
            self.valid_bom_combos.add(f"{mat}|{plant}")

        sc_df = pd.read_csv(self.source_constraint_path, sep="\t", dtype=str)
        sc_df.columns = [c.strip().upper() for c in sc_df.columns]
        for _, r in sc_df.iterrows():
            mat   = str(r.get(SC_MATERIAL_COL, "")).strip()
            plant = str(r.get(SC_PLANT_COL, "")).strip()
            if mat == "" or plant == "":
                continue
            self.valid_sc_combos.add(f"{mat}|{plant}")

    def validate(self):
        engine = PartSourceMakeBusinessRuleEngine(self.valid_bom_combos, self.valid_sc_combos)

        for idx, row in self.df.iterrows():
            try:
                reasons = engine.validate_row(row)
            except Exception:
                reasons = {}

            if reasons:
                self.error_map[idx]  = list(reasons.keys())
                self.reason_map[idx] = reasons

                material_reason = reasons.get("MATERIAL", "")
                if material_reason.startswith("MATERIAL: Blank field(s)"):
                    self.material_blank_count += 1
                else:
                    if "not found in BOM extract" in material_reason:
                        self.material_not_in_bom_count += 1
                    if "not found in Source Constraint extract" in material_reason:
                        self.material_not_in_sc_count += 1

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class PartSourceMakeBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MAXIMUMLOTSIZE": [
            "MAXIMUMLOTSIZE must be greater than 0.",
        ],
        "MINIMUMLOTSIZE": [
            "MINIMUMLOTSIZE must be greater than 0.",
        ],
        "ROUNDINGVALUE": [
            "ROUNDINGVALUE must be greater than 0.",
        ],
        "MATERIAL": [
            "The MATERIAL / PLANT (Part-Site) combination must be present in the "
            "BOM extract as ROOT_MATERIAL / ROOT_PLANT.",
            "The MATERIAL / PLANT (Part-Site) combination must also be present in "
            "the Source Constraint extract as MATERIAL / MKALPLANT.",
        ],
    }

    def __init__(self, validator: PartSourceMakeBusinessTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False, indent_col2: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill
        if indent_col2:
            ws.cell(row=row_num, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=2
            )

    # ══════════════════════════════════════════
    #  Summary sheet
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):
        v = self.validator

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="PartSource(Make) Business Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        # ── Per-field error counts (simple numeric fields) ──
        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            if col_name == "MATERIAL":
                reason_text = ("Blank / not found in BOM (ROOT_MATERIAL-ROOT_PLANT) / "
                               "not found in Source Constraint (MATERIAL-MKALPLANT)") if has_errors else ""
            else:
                reason_text = FIELD_REASON.get(col_name, "") if has_errors else ""

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            self._style_summary_data_row(ws, row_num, bold=(col_name == "MATERIAL"), fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1
            field_num += 1

            # ── MATERIAL sub-rows: Blank / Not in BOM / Not in Source Constraint ──
            if col_name == "MATERIAL":
                sub_counts = {
                    "Blank":                    v.material_blank_count,
                    "Not in BOM":               v.material_not_in_bom_count,
                    "Not in Source Constraint": v.material_not_in_sc_count,
                }
                for sub_label in MATERIAL_SUB_CATEGORIES:
                    sub_count      = sub_counts.get(sub_label, 0)
                    sub_pct_error  = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_error, 2)

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=f"└─ {sub_label}")
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_error}%")
                    ws.cell(row=row_num, column=7, value="")
                    self._style_summary_data_row(ws, row_num, italic=True, fill=SUBROW_FILL,
                                                  indent_col2=True)
                    row_num += 1

        # ── TOTAL row ──
        total_errors        = sum(col_error_counts.values())
        total_record_count  = total_rows * len(FIELD_ORDER)
        total_pct_error     = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health    = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2   # blank spacer

        # ── Quick-glance stats block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 42, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets  (ALL source columns) ──
    def _write_field_error_sheets(self, wb, df: pd.DataFrame, all_cols: list):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            keep_here = all_cols + ["ERROR_COLUMNS"]
            subset    = df.loc[row_indices, keep_here].copy()

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            # Fields to highlight per sheet: MATERIAL sheet highlights MATERIAL + PLANT
            # (the Part-Site combo); numeric sheets highlight just themselves.
            if field_name == "MATERIAL":
                highlight_cols = (PS_MATERIAL_COL, PS_PLANT_COL)
            else:
                highlight_cols = (field_name,)

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER

                for involved_col in highlight_cols:
                    if involved_col in col_idx_map:
                        target_cell      = ws.cell(row=excel_row, column=col_idx_map[involved_col])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="PartSource(Make) Table – Business Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 90

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        # All source columns, in original order — used for per-field error sheets
        all_cols = list(df.columns)

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df, all_cols)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows                       : {len(df)}")
        print(f"   Error rows                       : {len(v.error_map)}")
        print(f"     MATERIAL - Blank                : {v.material_blank_count}")
        print(f"     MATERIAL - Not in BOM            : {v.material_not_in_bom_count}")
        print(f"     MATERIAL - Not in Source Constr. : {v.material_not_in_sc_count}")
        print(f"   Field sheets                      : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartSourceMakeBusinessTableProcessor:

    def __init__(self, ps_make_path: str, bom_path: str, source_constraint_path: str, output_path: str):
        self.validator = PartSourceMakeBusinessTableValidator(ps_make_path, bom_path, source_constraint_path)
        self.writer    = PartSourceMakeBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    PartSource(Make) columns detected : {list(self.validator.df.columns)}")
        print(f"    Valid BOM Part-Site combinations   : {len(self.validator.valid_bom_combos)}")
        print(f"    Valid SourceConstraint combinations: {len(self.validator.valid_sc_combos)}")
        print("🔍  Validating business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PartSourceMakeBusinessTableProcessor(
        ps_make_path           = PARTSOURCE_MAKE_INPUT_FILE,
        bom_path               = BOM_INPUT_FILE,
        source_constraint_path = SOURCE_CONSTRAINT_INPUT_FILE,
        output_path            = OUTPUT_FILE,
    )
    processor.run()
