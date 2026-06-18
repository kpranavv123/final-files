import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  (fill these in before running)
# ─────────────────────────────────────────────
PARTSOURCE_MAKE_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource_Make\PartSource(Make)_2026-06-08-1646.tab"
PART_MASTER_INPUT_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\PartSite_2026-06-16-1757.tab"
SITE_INPUT_FILE            = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-06-17-1243.tab"
BOM_INPUT_FILE             = r"C:\Users\SW526XH\Downloads\Go Live-2\BOM\BOM_2026-06-08-1629.tab"
OUTPUT_FILE                = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource_Make\Validated_Part_Source_Make_Technical.xlsx"


# ─────────────────────────────────────────────
#  DATE FORMAT
# ─────────────────────────────────────────────
DATE_PATTERN = re.compile(r'^\d{4}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$')  # YYYYMMDD


# ─────────────────────────────────────────────
#  Colours / Styles
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUB_FILL        = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─────────────────────────────────────────────
#  FIELD ORDER & METADATA
# ─────────────────────────────────────────────
KEEP_COLS = [
    "MATERIAL",
    "PLANT",
    "PRODUCTIONVERSION",
    "VALIDFROM",
    "VALIDTO",
    "MATERIAL_PLANT",
    "MAXIMUMLOTSIZE",
    "MINIMUMLOTSIZE",
    "ROUNDINGVALUE",
]

FIELD_ORDER = [
    "MATERIAL",
    "PLANT",
    "PRODUCTIONVERSION",
    "VALIDFROM",
    "VALIDTO",
    "MATERIAL_PLANT",
    "MAXIMUMLOTSIZE",
    "MINIMUMLOTSIZE",
    "ROUNDINGVALUE",
]

FIELDS_WITH_SUB_ROWS = {
    "MATERIAL",
    "PLANT",
    "VALIDFROM",
    "VALIDTO",
}

FIELD_REASON = {
    "MATERIAL":           "",
    "PLANT":              "",
    "PRODUCTIONVERSION":  "PRODUCTIONVERSION: Field is blank",
    "VALIDFROM":          "",
    "VALIDTO":            "",
    "MATERIAL_PLANT":     "MATERIAL_PLANT: Field is blank",
    "MAXIMUMLOTSIZE":     "MAXIMUMLOTSIZE: Field is blank",
    "MINIMUMLOTSIZE":     "MINIMUMLOTSIZE: Field is blank",
    "ROUNDINGVALUE":      "ROUNDINGVALUE: Field is blank",
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class PartSourceMakeRuleEngine:

    def __init__(self,
                 part_master_materials: set,
                 site_plants: set,
                 # Set of MATERIAL_PLANT values (from PartSource) that passed the
                 # Part-Master filter (found in PM AND XPLANTMATSTATUS is blank or 02).
                 # Only these are candidates for the BOM cross-check.
                 eligible_for_bom_check: set,
                 # ROOT_MATERIAL_ROOT_PLANT values from BOM extract
                 bom_material_plant_pairs: set):
        self.part_master_materials  = set(str(m).strip().upper() for m in part_master_materials)
        self.site_plants            = set(str(p).strip() for p in site_plants)
        self.eligible_for_bom_check = eligible_for_bom_check   # already normalised by loader
        self.bom_material_plant_pairs = bom_material_plant_pairs  # already normalised by loader

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    @staticmethod
    def _valid_date(value) -> bool:
        return bool(DATE_PATTERN.match(str(value).strip()))

    # ── MATERIAL ──────────────────────────────
    def validate_material(self, row) -> str:
        val = row.get("MATERIAL", "")
        if self._is_blank(val):
            return "MATERIAL: Field is blank"

        val_str = str(val).strip().upper()

        if val_str not in self.part_master_materials:
            return f"MATERIAL: '{val_str}' is not present in Part master"

        # Rule 1c  ────────────────────────────────────────────────────────────
        # Use the pre-built MATERIAL_PLANT key from the PartSource row.
        # 1. If the key was not found in Part Master's MATERIALNUMBER_PLANT_XPLANTMATSTATUS
        #    column → skip (no error).
        # 2. If found but XPLANTMATSTATUS is not blank/02 → skip (no error).
        # 3. If eligible (found AND status blank/02) → must exist in BOM.
        mp_key = str(row.get("MATERIAL_PLANT", "")).strip()

        if mp_key in self.eligible_for_bom_check:
            if mp_key not in self.bom_material_plant_pairs:
                return (
                    f"MATERIAL: '{val_str}' + PLANT combo "
                    f"(XPLANTMATSTATUS blank/02) not found in BOM extract"
                )
        # Not eligible → no BOM check needed
        return ""

    # ── PLANT ─────────────────────────────────
    def validate_plant(self, row) -> str:
        val = str(row.get("PLANT", "")).strip()
        if not val or val == "nan":
            return "PLANT: Field is blank"
        if val not in self.site_plants:
            return f"PLANT: '{val}' is not present in the Site master"
        return ""

    # ── PRODUCTIONVERSION ─────────────────────
    def validate_productionversion(self, row) -> str:
        if self._is_blank(row.get("PRODUCTIONVERSION")):
            return "PRODUCTIONVERSION: Field is blank"
        return ""

    # ── VALIDFROM ─────────────────────────────
    def validate_validfrom(self, row) -> str:
        val = row.get("VALIDFROM", "")
        if self._is_blank(val):
            return "VALIDFROM: Field is blank"
        if not self._valid_date(val):
            return f"VALIDFROM: '{str(val).strip()}' does not follow the required format YYYYMMDD"
        return ""

    # ── VALIDTO ───────────────────────────────
    def validate_validto(self, row) -> str:
        val = row.get("VALIDTO", "")
        if self._is_blank(val):
            return "VALIDTO: Field is blank"
        if not self._valid_date(val):
            return f"VALIDTO: '{str(val).strip()}' does not follow the required format YYYYMMDD"
        return ""

    # ── MATERIAL_PLANT (derived) ───────────────
    def validate_material_plant(self, row) -> str:
        if self._is_blank(row.get("MATERIAL_PLANT", "")):
            return "MATERIAL_PLANT: Field is blank"
        return ""

    # ── MAXIMUMLOTSIZE ────────────────────────
    def validate_maximumlotsize(self, row) -> str:
        if self._is_blank(row.get("MAXIMUMLOTSIZE")):
            return "MAXIMUMLOTSIZE: Field is blank"
        return ""

    # ── MINIMUMLOTSIZE ────────────────────────
    def validate_minimumlotsize(self, row) -> str:
        if self._is_blank(row.get("MINIMUMLOTSIZE")):
            return "MINIMUMLOTSIZE: Field is blank"
        return ""

    # ── ROUNDINGVALUE ─────────────────────────
    def validate_roundingvalue(self, row) -> str:
        if self._is_blank(row.get("ROUNDINGVALUE")):
            return "ROUNDINGVALUE: Field is blank"
        return ""

    def get_rules(self) -> dict:
        return {
            "MATERIAL":          self.validate_material,
            "PLANT":             self.validate_plant,
            "PRODUCTIONVERSION": self.validate_productionversion,
            "VALIDFROM":         self.validate_validfrom,
            "VALIDTO":           self.validate_validto,
            "MATERIAL_PLANT":    self.validate_material_plant,
            "MAXIMUMLOTSIZE":    self.validate_maximumlotsize,
            "MINIMUMLOTSIZE":    self.validate_minimumlotsize,
            "ROUNDINGVALUE":     self.validate_roundingvalue,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartSourceMakeTableValidator:

    def __init__(self, partsource_path: str, part_master_path: str,
                 site_path: str, bom_path: str):
        self.partsource_path  = partsource_path
        self.part_master_path = part_master_path
        self.site_path        = site_path
        self.bom_path         = bom_path

        self.df                     = pd.DataFrame()
        self.part_master_materials  = set()
        self.site_plants            = set()
        self.eligible_for_bom_check = set()   # MATERIAL_PLANT keys passing PM filter
        self.bom_material_plant_pairs = set() # ROOT_MATERIAL_ROOT_PLANT keys from BOM

        self.error_map  = {}
        self.reason_map = {}

    def load(self):
        # ── PartSource(Make) ──────────────────────────────────────────────────
        self.df = pd.read_csv(self.partsource_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        # Use the pre-built MATERIAL_PLANT column from the file if present,
        # otherwise derive it on the fly (same as before).
        if "MATERIAL_PLANT" not in self.df.columns:
            mat   = self.df["MATERIAL"].fillna("").str.strip() if "MATERIAL" in self.df.columns else ""
            plant = self.df["PLANT"].fillna("").str.strip()    if "PLANT"    in self.df.columns else ""
            self.df["MATERIAL_PLANT"] = (mat + plant).where((mat != "") | (plant != ""), other="")

        # Normalise MATERIAL_PLANT to strip whitespace
        self.df["MATERIAL_PLANT"] = self.df["MATERIAL_PLANT"].fillna("").str.strip()

        # ── Part Master ───────────────────────────────────────────────────────
        pm_df = pd.read_csv(self.part_master_path, sep="\t", dtype=str)
        pm_df.columns = [c.strip().upper() for c in pm_df.columns]

        required_pm_cols = {"MATERIALNUMBER", "MATERIALNUMBER_PLANT_XPLANTMATSTATUS"}
        missing = required_pm_cols - set(pm_df.columns)
        if missing:
            raise ValueError(f"Part Master is missing columns: {missing}")

        self.part_master_materials = set(
            pm_df["MATERIALNUMBER"].dropna().str.strip().str.upper().tolist()
        )
        print(f"    Part Master materials loaded          : {len(self.part_master_materials)} unique values")

        # Build eligible_for_bom_check:
        #   Parse MATERIALNUMBER_PLANT_XPLANTMATSTATUS  →  "MATNUM_PLANT_STATUS"
        #   Extract the MATERIAL_PLANT portion (first two segments joined without the
        #   trailing underscore-STATUS) and keep only rows where STATUS is blank or '02'.
        #
        #   Column value examples:
        #     blank status  → "160000000000006834_1801_"          (ends with _)
        #     status=Z2     → "160000000000006834_1801_Z2"
        #
        #   The MATERIAL_PLANT in PartSource is a plain concatenation: "160000000000006834" + "1801"
        #   → "1600000000000068341801"
        #   So we reconstruct the MATERIAL_PLANT key the same way the PartSource file does it:
        #   strip the STATUS suffix (everything after the last underscore), then remove
        #   the remaining underscores to get the raw concatenation.

        eligible = set()
        col_vals = pm_df["MATERIALNUMBER_PLANT_XPLANTMATSTATUS"].dropna().str.strip()

        for raw in col_vals:
            # Split on the LAST underscore to isolate the STATUS part
            last_sep = raw.rfind("_")
            if last_sep == -1:
                # No underscore at all – malformed; skip
                continue
            status   = raw[last_sep + 1:].strip()   # "" if blank, "Z2" etc. otherwise
            mp_part  = raw[:last_sep]                # e.g. "160000000000006834_1801"

            # Only proceed if status is blank or "02"
            if status not in ("", "02"):
                continue

            # Convert "MAT_PLANT" → plain concatenation "MATPLANT"
            mp_key = mp_part.replace("_", "")
            eligible.add(mp_key)

        self.eligible_for_bom_check = eligible
        print(f"    Eligible MATERIAL_PLANT for BOM check : {len(self.eligible_for_bom_check)} (XPLANTMATSTATUS blank/02)")

        # ── Site Master ───────────────────────────────────────────────────────
        site_df = pd.read_csv(self.site_path, dtype=str, sep="\t")
        site_df.columns = [c.strip().upper() for c in site_df.columns]
        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")
        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site master plants loaded             : {len(self.site_plants)} unique values")

        # ── BOM Extract ───────────────────────────────────────────────────────
        bom_df = pd.read_csv(self.bom_path, sep="\t", dtype=str)
        bom_df.columns = [c.strip().upper() for c in bom_df.columns]

        required_bom_cols = {"ROOT_MATERIAL_ROOT_PLANT"}
        missing_bom = required_bom_cols - set(bom_df.columns)
        if missing_bom:
            raise ValueError(f"BOM extract is missing columns: {missing_bom}")

        # BOM also uses a plain concatenation column – normalise the same way
        self.bom_material_plant_pairs = set(
            bom_df["ROOT_MATERIAL_ROOT_PLANT"].dropna().str.strip().tolist()
        )
        print(f"    BOM MATERIAL_PLANT pairs loaded       : {len(self.bom_material_plant_pairs)} unique values")

    def validate(self):
        engine = PartSourceMakeRuleEngine(
            part_master_materials  = self.part_master_materials,
            site_plants            = self.site_plants,
            eligible_for_bom_check = self.eligible_for_bom_check,
            bom_material_plant_pairs = self.bom_material_plant_pairs,
        )
        rules = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols    = []
            col_reason_map = {}

            for col, rule_fn in rules.items():
                if col != "MATERIAL_PLANT" and col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)
                except Exception:
                    reason = f"{col}: Unexpected validation error"

                if reason:
                    failed_cols.append(col)
                    col_reason_map[col] = reason

            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = col_reason_map

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors

    # ── Sub-count helpers ─────────────────────
    def get_material_error_subcounts(self) -> dict:
        counts = {"blank": 0, "not_in_master": 0, "not_in_bom": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("MATERIAL", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            elif "not found in bom" in reason.lower():
                counts["not_in_bom"] += 1
            else:
                counts["not_in_master"] += 1
        return counts

    def get_plant_error_subcounts(self) -> dict:
        counts = {"blank": 0, "not_in_site": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("PLANT", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            else:
                counts["not_in_site"] += 1
        return counts

    def get_validfrom_error_subcounts(self) -> dict:
        counts = {"blank": 0, "bad_format": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("VALIDFROM", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            else:
                counts["bad_format"] += 1
        return counts

    def get_validto_error_subcounts(self) -> dict:
        counts = {"blank": 0, "bad_format": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("VALIDTO", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            else:
                counts["bad_format"] += 1
        return counts


# ══════════════════════════════════════════════
#  Report Writer  (unchanged from original)
# ══════════════════════════════════════════════
class PartSourceMakeReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIAL": [
            "Must not be blank.",
            "Must be present in the Part master (MATERIALNUMBER column).",
            "For rows where MATERIAL_PLANT exists in Part Master AND XPLANTMATSTATUS is blank or '02': "
            "the MATERIAL_PLANT must also be present in the BOM extract (ROOT_MATERIAL_ROOT_PLANT column).",
        ],
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Site master (PLANT column).",
        ],
        "PRODUCTIONVERSION": [
            "Must not be blank.",
        ],
        "VALIDFROM": [
            "Must not be blank.",
            "Must follow the format: YYYYMMDD.",
        ],
        "VALIDTO": [
            "Must not be blank.",
            "Must follow the format: YYYYMMDD.",
        ],
        "MATERIAL_PLANT": [
            "Derived column (concatenation of MATERIAL and PLANT). Must not be blank.",
        ],
        "MAXIMUMLOTSIZE": [
            "Must not be blank.",
        ],
        "MINIMUMLOTSIZE": [
            "Must not be blank.",
        ],
        "ROUNDINGVALUE": [
            "Must not be blank.",
        ],
    }

    def __init__(self, validator: PartSourceMakeTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):

        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="PartSource (Make) Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        material_sub  = self.validator.get_material_error_subcounts()
        plant_sub     = self.validator.get_plant_error_subcounts()
        validfrom_sub = self.validator.get_validfrom_error_subcounts()
        validto_sub   = self.validator.get_validto_error_subcounts()

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            if col_name in FIELDS_WITH_SUB_ROWS:
                reason_text = ""
            elif has_errors:
                reason_text = FIELD_REASON.get(col_name, "")
            else:
                reason_text = ""

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            self._style_summary_data_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1

            # ── MATERIAL sub-rows ──
            if col_name == "MATERIAL" and has_errors:
                sub_definitions = [
                    (
                        "  ↳ Blank Material",
                        material_sub["blank"],
                        "MATERIAL: Field is blank",
                    ),
                    (
                        "  ↳ Not in Part Master",
                        material_sub["not_in_master"],
                        "MATERIAL: Not present in Part master",
                    ),
                    (
                        "  ↳ MATERIAL_PLANT (XPLANTMATSTATUS blank/02) not in BOM",
                        material_sub["not_in_bom"],
                        "MATERIAL: MATERIAL_PLANT combo (XPLANTMATSTATUS blank/02) not found in BOM extract",
                    ),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            # ── PLANT sub-rows ──
            if col_name == "PLANT" and has_errors:
                sub_definitions = [
                    ("  ↳ Blank Plant",        plant_sub["blank"],       "PLANT: Field is blank"),
                    ("  ↳ Not in Site Master", plant_sub["not_in_site"], "PLANT: Site not found in the Site master"),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            # ── VALIDFROM sub-rows ──
            if col_name == "VALIDFROM" and has_errors:
                sub_definitions = [
                    ("  ↳ Blank Valid From Date",        validfrom_sub["blank"],      "VALIDFROM: Field is blank"),
                    ("  ↳ Invalid Format (not YYYYMMDD)", validfrom_sub["bad_format"], "VALIDFROM: Does not follow required format YYYYMMDD"),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            # ── VALIDTO sub-rows ──
            if col_name == "VALIDTO" and has_errors:
                sub_definitions = [
                    ("  ↳ Blank Valid To Date",           validto_sub["blank"],      "VALIDTO: Field is blank"),
                    ("  ↳ Invalid Format (not YYYYMMDD)", validto_sub["bad_format"], "VALIDTO: Does not follow required format YYYYMMDD"),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)
                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)
                    self._style_summary_data_row(ws, row_num, fill=SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            field_num += 1

        # ── TOTAL row ──
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        col_widths = [6, 52, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            subset    = df.loc[row_indices].copy()
            keep_here = [c for c in KEEP_COLS if c in subset.columns] + ["ERROR_COLUMNS"]
            subset    = subset[keep_here]

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER

                if field_name in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="PartSource (Make) Table – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 90

    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        keep_cols = [c for c in KEEP_COLS if c in df.columns] + ["ERROR_COLUMNS"]
        df        = df[keep_cols]

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartSourceMakeTableProcessor:

    def __init__(self, partsource_path: str, part_master_path: str,
                 site_path: str, bom_path: str, output_path: str):
        self.validator = PartSourceMakeTableValidator(
            partsource_path, part_master_path, site_path, bom_path
        )
        self.writer = PartSourceMakeReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    PartSource(Make) columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PartSourceMakeTableProcessor(
        partsource_path  = PARTSOURCE_MAKE_INPUT_FILE,
        part_master_path = PART_MASTER_INPUT_FILE,
        site_path        = SITE_INPUT_FILE,
        bom_path         = BOM_INPUT_FILE,
        output_path      = OUTPUT_FILE,
    )
    processor.run()
