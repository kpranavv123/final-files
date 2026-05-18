"""
Part UOM Technical Validator
═════════════════════════════════════════════════════════════════════════════

Purpose:
  Validates Part UOM Conversion file against technical/data-quality rules only.

Technical Rules:
  PRODUCT:
    1. PRODUCT should not be blank
    2. PRODUCT + PLANT combination should exist in Part Master

  PLANT:
    1. PLANT should not be blank
    2. PLANT should exist in Site Master

  ALTERNATIVEUNIT:
    1. ALTERNATIVEUNIT should not be blank
    2. Each PRODUCT + PLANT combination must have all three units: KG, CV, PAC
       (if any one is missing, all rows for that combination are flagged)

  CONVERSIONFACTOR:
    1. CONVERSIONFACTOR should not be blank
    2. CONVERSIONFACTOR should be numeric
    3. CONVERSIONFACTOR should be non-zero

Output:
  Validated_PartUOM_Technical.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
PART_UOM_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\PartUOM\PUOM_2026-05-14-1818.tab"
PART_REFERENCE_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-05-14-1817.tab"
SITE_REFERENCE_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-11-1205.tab"
OUTPUT_FILE         = r"C:\Users\SW526XH\Downloads\Go Live-1\PartUOM\Validated_PartUOM_Technical.xlsx"

# Required units that every PRODUCT+PLANT combination must have
REQUIRED_UNITS = {"KG", "CV", "PAC"}


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=9)
ERR_FONT    = Font(name="Arial", size=9, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Technical Ruleset Info
# ══════════════════════════════════════════════
SUMMARY_RULESET_INFO = {
    "PRODUCT": [
        "PRODUCT is blank",
        "PRODUCT and PLANT combination not found in part master",
    ],
    "PLANT": [
        "PLANT is blank",
        "PLANT not found in site master",
    ],
    "ALTERNATIVEUNIT": [
        "ALTERNATIVEUNIT is blank",
        "PRODUCT+PLANT combination is missing required unit(s): KG, CV or PAC",
    ],
    "CONVERSIONFACTOR": [
        "CONVERSIONFACTOR is blank",
        "CONVERSIONFACTOR is not numeric",
        "CONVERSIONFACTOR must be non-zero",
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class PartUOMTechnicalRuleEngine:
    """Technical validation rules for Part UOM Conversion."""

    def __init__(self, part_plant_combos: set, site_codes: set):
        self.part_plant_combos = part_plant_combos
        self.site_codes = set(str(s).strip() for s in site_codes)
        # Will be populated before row-level validation
        self.combo_missing_units: dict = {}   # (product, plant) -> set of missing units

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _is_numeric(self, value) -> bool:
        if self._is_blank(value):
            return False
        try:
            float(str(value).strip())
            return True
        except (ValueError, TypeError):
            return False

    def precompute_missing_units(self, df: pd.DataFrame):
        """
        For every PRODUCT+PLANT combination present in the dataframe, determine
        which of the required units (KG, CV, PAC) are absent.
        Combos that have all three units are NOT stored (they pass).
        """
        self.combo_missing_units = {}

        if "ALTERNATIVEUNIT" not in df.columns:
            return

        # Build a mapping: (product, plant) -> set of units present
        combo_units: dict = {}

        for _, row in df.iterrows():
            product = row.get("PRODUCT", "")
            plant   = row.get("PLANT", "")
            unit    = row.get("ALTERNATIVEUNIT", "")

            if self._is_blank(product) or self._is_blank(plant):
                continue

            product_str = str(product).strip()
            plant_str   = str(plant).strip()
            unit_str    = str(unit).strip().upper() if not self._is_blank(unit) else ""

            key = (product_str, plant_str)
            combo_units.setdefault(key, set())

            if unit_str:
                combo_units[key].add(unit_str)

        # Determine missing required units per combo
        for key, present_units in combo_units.items():
            missing = REQUIRED_UNITS - present_units
            if missing:
                self.combo_missing_units[key] = missing

    def validate_product(self, row) -> tuple:
        product = row.get("PRODUCT")
        plant   = row.get("PLANT")

        if self._is_blank(product):
            return False, "PRODUCT is blank"

        product_str = str(product).strip()
        plant_str   = str(plant).strip() if not self._is_blank(plant) else ""
        combo       = (product_str, plant_str)

        if combo in self.part_plant_combos:
            return True, ""

        return False, "PRODUCT and PLANT combination not found in part master"

    def validate_plant(self, row) -> tuple:
        plant = row.get("PLANT")

        if self._is_blank(plant):
            return False, "PLANT is blank"

        plant_str = str(plant).strip()

        if plant_str in self.site_codes:
            return True, ""

        return False, "PLANT not found in site master"

    def validate_alternativeunit(self, row) -> tuple:
        """
        Rule 1 – ALTERNATIVEUNIT must not be blank.
        Rule 2 – The PRODUCT+PLANT combination must contain all of KG, CV, PAC.
                 If one or more are missing, every row for that combination is
                 flagged with a message listing exactly which unit(s) are absent.
        """
        alt_unit = row.get("ALTERNATIVEUNIT")

        # Rule 1: blank check
        if self._is_blank(alt_unit):
            return False, "ALTERNATIVEUNIT is blank"

        # Rule 2: required-units check (group level)
        product = row.get("PRODUCT", "")
        plant   = row.get("PLANT", "")

        if not self._is_blank(product) and not self._is_blank(plant):
            key = (str(product).strip(), str(plant).strip())
            missing = self.combo_missing_units.get(key)

            if missing:
                missing_sorted = ", ".join(sorted(missing))
                return (
                    False,
                    f"For this PRODUCT+PLANT combination, required unit(s) missing: {missing_sorted}",
                )

        return True, ""

    def validate_conversionfactor(self, row) -> tuple:
        cf = row.get("CONVERSIONFACTOR")

        if self._is_blank(cf):
            return False, "CONVERSIONFACTOR is blank"

        if not self._is_numeric(cf):
            return False, "CONVERSIONFACTOR is not numeric"

        try:
            if float(str(cf).strip()) == 0:
                return False, "CONVERSIONFACTOR must be non-zero"
        except (ValueError, TypeError):
            return False, "CONVERSIONFACTOR is not numeric"

        return True, ""

    def get_rules(self) -> dict:
        return {
            "PRODUCT":          self.validate_product,
            "PLANT":            self.validate_plant,
            "ALTERNATIVEUNIT":  self.validate_alternativeunit,
            "CONVERSIONFACTOR": self.validate_conversionfactor,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartUOMTechnicalValidator:
    """Loads data, validates technical rules, builds error map."""

    def __init__(self):
        self.df = pd.DataFrame()
        self.part_plant_combos = set()
        self.site_codes = set()
        self.error_map = {}

    def _add_error(self, idx, field_name: str, reason: str):
        self.error_map.setdefault(idx, {})

        if field_name in self.error_map[idx] and self.error_map[idx][field_name] != reason:
            existing = self.error_map[idx][field_name]
            if reason not in existing.split(" | "):
                self.error_map[idx][field_name] = existing + " | " + reason
        else:
            self.error_map[idx][field_name] = reason

    # ── helpers ──────────────────────────────
    @staticmethod
    def _read_file(path: str) -> pd.DataFrame:
        lower = path.lower()
        if lower.endswith(".csv"):
            return pd.read_csv(path, dtype=str)
        elif lower.endswith((".tab", ".tsv")):
            return pd.read_csv(
                path, sep="\t", dtype=str,
                encoding="utf-8", encoding_errors="replace", engine="python",
            )
        elif lower.endswith((".xlsx", ".xls")):
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {path}")

    def load(self):
        print("[LOAD] Loading files...")

        # ── PartUOM input ────────────────────
        self.df = self._read_file(PART_UOM_INPUT_FILE)
        self.df.columns = [str(c).strip().upper() for c in self.df.columns]
        print(f"    PartUOM rows loaded: {len(self.df)}")

        # ── Part master reference ────────────
        part_df = self._read_file(PART_REFERENCE_FILE)
        part_df.columns = [str(c).strip().upper() for c in part_df.columns]

        mat_col = "MATERIALNUMBER" if "MATERIALNUMBER" in part_df.columns else "MATERIAL"
        if mat_col not in part_df.columns and "PRODUCT" in part_df.columns:
            mat_col = "PRODUCT"

        missing = [c for c in [mat_col, "PLANT"] if c not in part_df.columns]
        if missing:
            raise ValueError(
                f"Missing columns in Part master: {missing}. Available: {list(part_df.columns)}"
            )

        for _, row in part_df.iterrows():
            part_name = str(row[mat_col]).strip() if pd.notna(row[mat_col]) else ""
            plant     = str(row["PLANT"]).strip()  if pd.notna(row["PLANT"])  else ""
            if part_name and plant:
                self.part_plant_combos.add((part_name, plant))

        print(f"    Part master - PRODUCT/PLANT combos: {len(self.part_plant_combos)}")

        # ── Site master reference ────────────
        site_df = self._read_file(SITE_REFERENCE_FILE)
        site_df.columns = [str(c).strip().upper() for c in site_df.columns]

        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")

        self.site_codes = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site master - PLANT codes: {len(self.site_codes)}")

    def validate(self):
        print("[VALIDATE] Running technical validation rules...")

        engine = PartUOMTechnicalRuleEngine(self.part_plant_combos, self.site_codes)

        # Pre-compute which PRODUCT+PLANT combos are missing required units
        engine.precompute_missing_units(self.df)
        print(
            f"    PRODUCT+PLANT combos missing required units (KG/CV/PAC): "
            f"{len(engine.combo_missing_units)}"
        )

        rules = engine.get_rules()

        for idx, row in self.df.iterrows():
            for field, rule_fn in rules.items():
                if field not in self.df.columns:
                    continue
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {e}"

                if not passed:
                    self._add_error(idx, field, reason)

    def get_error_series(self) -> pd.Series:
        details = {}
        for idx, errdict in self.error_map.items():
            messages = [f"{fld}: {msg}" for fld, msg in errdict.items()]
            details[idx] = "; ".join(messages)
        return pd.Series(details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class PartUOMTechnicalReportWriter:
    """Builds technical validation workbook."""

    SHEET_SUMMARY  = "Summary"
    SHEET_RULESETS = "Rulesets"

    def __init__(self, validator: PartUOMTechnicalValidator, output_path: str):
        self.validator = validator
        self.output_path = output_path
        self._summary_fields_order = []

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)
        for ch in invalid_chars:
            name = name.replace(ch, "-")
        name = name.strip() or "Sheet"
        name = name[:31]
        if name not in wb.sheetnames:
            return name
        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = 0
            for c in col:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 4, 80))

    def _get_ruleset_columns(self):
        ruleset_fields = ["PRODUCT", "PLANT", "ALTERNATIVEUNIT", "CONVERSIONFACTOR"]
        ruleset_columns = [col for col in ruleset_fields if col in self.validator.df.columns]
        ruleset_columns.append("ERROR_FIELDS")
        return ruleset_columns

    def _summary_order(self):
        return ["PRODUCT", "PLANT", "ALTERNATIVEUNIT", "CONVERSIONFACTOR"]

    # ── Rulesets sheet ────────────────────────
    def _write_ruleset_sheet(self, wb, summary_fields=None):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        title_cell = ws.cell(row=1, column=1, value="Part UOM Table – Technical Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:C1")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell            = ws.cell(row=3, column=c_idx, value=h)
            cell.fill       = HDR_FILL
            cell.font       = HDR_FONT
            cell.border     = THIN_BORDER
            cell.alignment  = Alignment(horizontal="center")

        ruleset_info = {
            "PRODUCT": [
                "Must not be blank.",
                "PRODUCT and PLANT combination must exist in part master.",
            ],
            "PLANT": [
                "Must not be blank.",
                "Must exist in site master.",
            ],
            "ALTERNATIVEUNIT": [
                "Must not be blank.",
                (
                    "Each PRODUCT + PLANT combination must contain all three required units: "
                    "KG, CV, and PAC. If any one of these units is missing for a combination, "
                    "ALL rows belonging to that combination are flagged and the error message "
                    "specifies exactly which unit(s) are absent "
                    "(e.g. 'For this PRODUCT+PLANT combination, required unit(s) missing: KG'). "
                    "Combinations that already have all three units are NOT flagged, even if "
                    "they also contain additional units."
                ),
            ],
            "CONVERSIONFACTOR": [
                "Must not be blank.",
                "Must be numeric.",
                "Must be non-zero.",
            ],
        }

        ordered_fields = summary_fields or list(ruleset_info.keys())
        current_row = 4
        rule_num    = 1

        for field in ordered_fields:
            if field not in ruleset_info:
                continue

            rules_list = ruleset_info[field]
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell = ws.cell(
                    row=current_row, column=1,
                    value=rule_num if r_idx == 0 else "",
                )
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell = ws.cell(
                    row=current_row, column=2,
                    value=field if r_idx == 0 else "",
                )
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 100

    # ── Summary sheet ─────────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        title_cell = ws.cell(row=1, column=1, value="Part UOM Technical Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        col_widths = [6, 28, 16, 16, 16, 16, 90]

        for c_idx, h in enumerate(headers, start=1):
            cell            = ws.cell(row=3, column=c_idx, value=h)
            cell.fill       = TITLE_FILL
            cell.font       = Font(name="Arial", bold=True)
            cell.border     = THIN_BORDER
            cell.alignment  = Alignment(horizontal="center", vertical="center")

        field_order     = self._summary_order()
        col_error_counts = {field: 0 for field in field_order}
        rule_error_counts: dict = {}

        for bad_cols in error_map.values():
            for col, reason in bad_cols.items():
                if col not in col_error_counts:
                    continue
                col_error_counts[col] += 1

                # Bucket similar "missing unit" messages under a generic key
                display_reason = reason
                if "required unit(s) missing:" in reason:
                    display_reason = "PRODUCT+PLANT combination is missing required unit(s): KG, CV or PAC"

                rule_error_counts[(col, display_reason)] = (
                    rule_error_counts.get((col, display_reason), 0) + 1
                )

        sorted_fields = [(field, col_error_counts.get(field, 0)) for field in field_order]
        self._summary_fields_order = [field for field, _ in sorted_fields]

        row_num      = 4
        item_counter = 1

        for field_name, count in sorted_fields:
            reasons      = SUMMARY_RULESET_INFO.get(field_name, [])
            actual_reasons = set(r for (f, r) in rule_error_counts.keys() if f == field_name)
            all_reasons  = list(reasons) + list(actual_reasons - set(reasons))
            is_multi     = len(all_reasons) > 1

            if is_multi:
                ws.cell(row=row_num, column=1, value=item_counter)
                ws.cell(row=row_num, column=2, value=field_name)
                ws.cell(row=row_num, column=3, value=count)
                ws.cell(row=row_num, column=4, value=total_rows)
                err_pct = count / total_rows if total_rows else 0
                ws.cell(row=row_num, column=5, value=1 - err_pct).number_format = "0.00%"
                ws.cell(row=row_num, column=6, value=err_pct).number_format       = "0.00%"
                ws.cell(row=row_num, column=7, value="")

                for c in range(1, 8):
                    cell            = ws.cell(row=row_num, column=c)
                    cell.border     = THIN_BORDER
                    cell.alignment  = Alignment(horizontal="center" if c != 7 else "left")
                    cell.font       = BODY_FONT

                row_num += 1

                for reason in all_reasons:
                    sub_count = rule_error_counts.get((field_name, reason), 0)
                    sub_pct   = sub_count / total_rows if total_rows else 0

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=f"↳ {reason}")
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=1 - sub_pct).number_format = "0.00%"
                    ws.cell(row=row_num, column=6, value=sub_pct).number_format       = "0.00%"
                    ws.cell(row=row_num, column=7, value=reason if sub_count > 0 else "")

                    for c in range(1, 8):
                        cell = ws.cell(row=row_num, column=c)
                        cell.border    = THIN_BORDER
                        cell.alignment = Alignment(
                            horizontal="center" if c != 7 else "left",
                            indent=(1 if c == 2 else 0),
                            wrap_text=(c == 7),
                        )
                        cell.font = BODY_FONT

                    row_num += 1

            else:
                reason  = all_reasons[0] if all_reasons else ""
                err_pct = count / total_rows if total_rows else 0
                values  = [item_counter, field_name, count, total_rows, 1 - err_pct, err_pct,
                           reason if count > 0 else ""]

                for c_idx, value in enumerate(values, start=1):
                    cell            = ws.cell(row=row_num, column=c_idx, value=value)
                    cell.font       = BODY_FONT
                    cell.border     = THIN_BORDER
                    cell.alignment  = Alignment(
                        horizontal="center" if c_idx != 7 else "left",
                        wrap_text=(c_idx == 7),
                    )
                    if c_idx in [5, 6]:
                        cell.number_format = "0.00%"

                row_num += 1

            item_counter += 1

        # ── Totals row ───────────────────────
        total_errors        = sum(col_error_counts.values())
        sum_record_counts   = len(sorted_fields) * total_rows
        total_error_percent = total_errors / sum_record_counts if sum_record_counts else 0
        total_fill          = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

        ws.cell(row=row_num, column=2, value="TOTAL").font             = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font        = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=sum_record_counts).font   = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=5, value=1 - total_error_percent).number_format = "0.00%"
        ws.cell(row=row_num, column=6, value=total_error_percent).number_format      = "0.00%"

        for c in range(1, 8):
            cell            = ws.cell(row=row_num, column=c)
            cell.fill       = total_fill
            cell.border     = THIN_BORDER
            cell.font       = Font(name="Arial", bold=True)
            cell.alignment  = Alignment(horizontal="center" if c != 7 else "left")

        row_num += 2

        # ── Stats block ──────────────────────
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors
        stats = [
            ("Total Records:",        total_rows),
            ("Records with Errors:",  records_with_errors),
            ("Records Passing:",      records_passing),
        ]
        stats_fill = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

        for label, value in stats:
            label_cell            = ws.cell(row=row_num, column=1, value=label)
            label_cell.font       = Font(name="Arial", bold=True, size=10)
            label_cell.fill       = stats_fill
            label_cell.border     = THIN_BORDER
            label_cell.alignment  = Alignment(horizontal="left")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)

            value_cell            = ws.cell(row=row_num, column=3, value=value)
            value_cell.font       = Font(name="Arial", size=10)
            value_cell.border     = THIN_BORDER
            value_cell.alignment  = Alignment(horizontal="center")
            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v          = self.validator
        all_fields = set()

        for errdict in v.error_map.values():
            all_fields.update(errdict.keys())

        fields_to_process = [f for f in self._summary_fields_order if f in all_fields]

        for field_name in fields_to_process:
            row_indices = [idx for idx, errdict in v.error_map.items() if field_name in errdict]

            if not row_indices:
                continue

            sheet_name = self._safe_sheet_name(wb, field_name)
            ws         = wb.create_sheet(sheet_name)

            subset                = df.loc[row_indices].copy()
            subset["ERROR_FIELDS"] = subset.index.map(
                lambda i, fn=field_name: v.error_map.get(i, {}).get(fn, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell            = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font       = BODY_FONT
                    cell.fill       = ROW_FILL
                    cell.alignment  = Alignment(vertical="center", wrap_text=True)

                # Highlight the erroneous field cell in red
                if field_name in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._set_widths(ws)

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Main write entry point ────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series        = v.get_error_series()
        df["ERROR_FIELDS"]  = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        ruleset_columns = self._get_ruleset_columns()
        filtered_cols   = [col for col in df.columns if col in ruleset_columns]
        df              = df[filtered_cols]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, len(df))
        self._write_ruleset_sheet(wb, self._summary_fields_order)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        print(f"\n[SAVE] Technical output saved: {self.output_path}")
        print(f"   Total records    : {len(df)}")
        print(f"   Error records    : {len(v.error_map)}")
        print(f"   Passing records  : {len(df) - len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartUOMTechnicalProcessor:
    def __init__(self):
        self.validator = PartUOMTechnicalValidator()
        self.writer    = PartUOMTechnicalReportWriter(self.validator, OUTPUT_FILE)

    def run(self):
        print("=" * 70)
        print("Part UOM Technical Validation Pipeline")
        print("=" * 70)

        self.validator.load()
        self.validator.validate()
        self.writer.write()

        print("\n" + "=" * 70)
        print("Technical Validation Complete!")
        print("=" * 70)


if __name__ == "__main__":
    processor = PartUOMTechnicalProcessor()
    processor.run()
