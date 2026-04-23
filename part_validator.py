import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_FG2026-04-21-1448.tab"
OUTPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part.xlsx"

# ─────────────────────────────────────────────
#  CONSOLIDATED PL LIST  (hardcoded)
# ─────────────────────────────────────────────
VALID_PLANTS = {
    "1555", "1452", "1204", "1292", "1253", "1234", "1731", "1508",
    "5011637", "1100", "2091", "2081", "1436", "1448", "1438", "1501",
    "1248", "1257", "1623", "1601", "1451", "2092", "1125", "1649",
    "1646", "2088", "1494", "1137", "1275", "2084", "1437", "5018849",
    "1554", "1213", "1499", "1575", "1104", "1104A", "1145", "1430",
    "1249", "1509", "1774", "1295", "1127", "1146", "1180", "1432",
    "1500", "1742", "1642", "1214", "1426", "1241", "1463", "1105",
    "1423", "1648", "1733", "1780", "1428", "1495", "1462", "1738",
    "5123531", "1233", "1158", "1157", "1482", "1483", "1106", "1485",
    "1107", "1225", "5123742", "1481", "1203", "2082", "1416", "2087",
    "1569", "1647", "1440", "1645", "1724", "1503", "1651", "1421",
    "1205", "1563", "1640", "1109", "1471", "1155", "1110", "1948",
    "1473", "1477", "2089", "1449", "1211", "1788", "1726", "1578",
    "1135", "1491", "1226", "1650", "1505", "1433", "1739", "2085",
    "1659", "1256", "1653", "1652", "1111", "1757", "1559", "1558",
    "1489", "1657", "4007430", "1654", "1112", "1784", "1579", "1571",
    "1754", "2090", "1522", "1771", "1113", "1229", "1235", "1156",
    "1258", "1176", "1197", "1265", "1445", "1475", "1740", "2083",
    "1725", "1484", "1114", "1480", "1506", "5011407", "1520",
    "5015073", "1627", "1758", "1240", "1478", "1429", "1656",
    "1476", "1298", "1801", "1802", "1661", "1785", "1507",
    "5011308", "1643", "1638", "1521", "1118", "5013796", "1511",
    "1487", "1488", "1512", "1281", "1166", "1184", "1186",
    "1218", "1223", "1279", "1442", "1732", "1472", "2086",
    "1439", "1425", "1208", "1722", "1734", "4011702", "1525",
    "1568", "1296", "1589", "1658", "5123296", "1518", "1455",
    "1196", "1441", "1190"
}


# ─────────────────────────────────────────────
#  STYLING CONSTANTS
# ─────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FF0000")
ROW_FILL   = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL   = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL  = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
STATS_FILL = PatternFill("solid", fgColor="EDEDED")

HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  FEATURE 2: Only these columns shown in error sheets
#  (mirrors the Rules sheet fields exactly)
#  ABCINDICATOR removed entirely
# ─────────────────────────────────────────────
RULES_FIELDS_ORDERED = [
    "MATERIALNUMBER",
    "PLANT",
    "PRODUCTDESCRIPTION",
    "PRODUCTTYPE",
    "PRODUCTHIERARCHY",
    "BASEUNIT",
    "MRPTYPE",
    "PROCUREMENTTYPE",
    "IBPSTATUS",
    "XPLANTMATSTATUS",
]

# ─────────────────────────────────────────────
#  FEATURE 3: Error sub-sheet creation order
# ─────────────────────────────────────────────
ERROR_SHEET_PRIORITY = [
    "PLANT",
    "PROCUREMENTTYPE",
    "PRODUCTHIERARCHY",
    "MRPTYPE",
]


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class RuleEngine:

    def __init__(self, valid_plants: set):
        self.valid_plants = valid_plants

    @staticmethod
    def _is_blank(value) -> bool:
        if pd.isna(value):
            return True
        return str(value).strip() == ""

    M_NUMBER_REASON = "MATERIALNUMBER: Must be 14xxxxxxxxxxxxxx (FERT) or 15xxxxxxxxxxxxxx (HAWA) — invalid range or blank"
    PLANT_REASON    = "PLANT: Value is not in the Consolidated PL list"
    PRODDESC_REASON = "PRODUCTDESCRIPTION: Field is blank"
    PRODTYPE_REASON = "PRODUCTTYPE: Must be FERT or HAWA — invalid or blank value found"
    PRODHIER_REASON = "PRODUCTHIERARCHY: Field is blank "
    BASEUNIT_REASON = "BASEUNIT: Field is blank"
    MRPTYPE_REASON  = "MRPTYPE: Field is blank"
    PROC_REASON     = "PROCUREMENTTYPE: Field is blank"
    IBP_REASON      = "IBPSTATUS: Must be 'IBP' or blank — unexpected value found"
    XPLANT_REASON   = "XPLANTMATSTATUS: Must be '2' or blank — unexpected value found"

    def validate_material_number(self, row) -> tuple[bool, str]:
        raw = row.get("MATERIALNUMBER")
        if self._is_blank(raw):
            return False, self.M_NUMBER_REASON
        valstr = str(raw).strip()
        if not valstr.isdigit():
            return False, self.M_NUMBER_REASON
        val      = int(valstr)
        mat_type = str(row.get("PRODUCTTYPE", "")).strip().upper()
        if mat_type == "FERT" and 14000000000000 <= val <= 14999999999999:
            return True, ""
        if mat_type == "HAWA" and 15000000000000 <= val <= 15999999999999:
            return True, ""
        return False, self.M_NUMBER_REASON

    def validate_plant(self, row) -> tuple[bool, str]:
        val = str(row.get("PLANT", "")).strip().upper()
        if self._is_blank(val) or val not in self.valid_plants:
            return False, self.PLANT_REASON
        return True, ""

    def validate_product_description(self, row) -> tuple[bool, str]:
        if self._is_blank(row.get("PRODUCTDESCRIPTION")):
            return False, self.PRODDESC_REASON
        return True, ""

    def validate_product_type(self, row) -> tuple[bool, str]:
        val = str(row.get("PRODUCTTYPE", "")).strip().upper()
        if val in {"FERT", "HAWA"}:
            return True, ""
        return False, self.PRODTYPE_REASON

    def validate_product_hierarchy(self, row) -> tuple[bool, str]:
        if self._is_blank(row.get("PRODUCTHIERARCHY")):
            return False, self.PRODHIER_REASON
        return True, ""

    def validate_base_unit(self, row) -> tuple[bool, str]:
        if self._is_blank(row.get("BASEUNIT")):
            return False, self.BASEUNIT_REASON
        return True, ""

    def validate_mrp_type(self, row) -> tuple[bool, str]:
        val = str(row.get("MRPTYPE", "")).strip().upper()
        if val in {"ND", "PD"}:
            return True, ""
        return False, self.MRPTYPE_REASON

    def validate_procurement_type(self, row) -> tuple[bool, str]:
        if self._is_blank(row.get("PROCUREMENTTYPE")):
            return False, self.PROC_REASON
        return True, ""

    def validate_ibp_status(self, row) -> tuple[bool, str]:
        raw = row.get("IBPSTATUS")
        if self._is_blank(raw):
            return True, ""
        if str(raw).strip().upper() == "IBP":
            return True, ""
        return False, self.IBP_REASON

    def validate_xplant_mat_status(self, row) -> tuple[bool, str]:
        raw = row.get("XPLANTMATSTATUS")
        if self._is_blank(raw):
            return True, ""
        if str(raw).strip() in {"2", "02"}:
            return True, ""
        return False, self.XPLANT_REASON

    def get_rules(self) -> dict:
        return {
            "MATERIALNUMBER":     self.validate_material_number,
            "PLANT":              self.validate_plant,
            "PRODUCTDESCRIPTION": self.validate_product_description,
            "PRODUCTTYPE":        self.validate_product_type,
            "PRODUCTHIERARCHY":   self.validate_product_hierarchy,
            "BASEUNIT":           self.validate_base_unit,
            "MRPTYPE":            self.validate_mrp_type,
            "PROCUREMENTTYPE":    self.validate_procurement_type,
            "IBPSTATUS":          self.validate_ibp_status,
            "XPLANTMATSTATUS":    self.validate_xplant_mat_status,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartTableValidator:

    def __init__(self, filepath: str, valid_plants: set):
        self.filepath     = filepath
        self.valid_plants = valid_plants
        self.df           = pd.DataFrame()
        self.error_map    = {}

    def load(self):
        path = self.filepath.lower()

        if path.endswith(".csv"):
            self.df = pd.read_csv(self.filepath, dtype=str)

        elif path.endswith(".tab") or path.endswith(".tsv"):
            self.df = pd.read_csv(
                self.filepath,
                sep="\t",
                dtype=str,
                encoding="utf-8",
                engine="python"
            )

        elif path.endswith(".xlsx") or path.endswith(".xls"):
            self.df = pd.read_excel(
                self.filepath,
                dtype=str,
                engine="openpyxl"
            )

        else:
            raise ValueError(f"Unsupported file format: {self.filepath}")

        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        engine = RuleEngine(self.valid_plants)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            errors = {}
            for col, rule_fn in rules.items():
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {str(e)}"
                if not passed:
                    errors[col] = reason
            if errors:
                self.error_map[idx] = errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class ReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIALNUMBER": [
            "Must not be blank.",
            "For FERT type: Material number must be in range 14000000000000 – 14999999999999.",
            "For HAWA type: Material number must be in range 15000000000000 – 15999999999999.",
        ],
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Consolidated PL list.",
        ],
        "PRODUCTDESCRIPTION": ["Must not be blank."],
        "PRODUCTTYPE": [
            "Must not be blank.",
            "Value must be either FERT or HAWA.",
        ],
        "PRODUCTHIERARCHY": ["Must not be blank."],
        "BASEUNIT": [
            "Must not be blank.",
        ],
        "MRPTYPE": [
            "Must not be blank.",
            "Value must be either ND or PD.",
        ],
        "PROCUREMENTTYPE": ["Must not be blank."],
        "IBPSTATUS": [
            "Allowed values: IBP or blank.",
            "Any other value is treated as an error.",
        ],
        "XPLANTMATSTATUS": [
            "Allowed values: 2 or blank.",
            "Any other value is treated as an error.",
        ],
    }

    def __init__(self, validator: PartTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── Helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _auto_width(self, ws, min_w=10, max_w=60):
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(length + 3, min_w), max_w
            )

    # ── Summary sheet ────────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        col_error_counts = {col: 0 for col in RULES_FIELDS_ORDERED}
        for bad_cols in error_map.values():
            for col in bad_cols.keys():
                if col in col_error_counts:
                    col_error_counts[col] += 1

        REASON_MAP = {
            "MATERIALNUMBER":     "MATERIALNUMBER: Must be 14xxxxxxxxxxxxxx (FERT) or 15xxxxxxxxxxxxxx (HAWA) — invalid range or blank",
            "PLANT":              "PLANT: Value is not in the Consolidated PL list",
            "PRODUCTDESCRIPTION": "PRODUCTDESCRIPTION: Field is blank",
            "PRODUCTTYPE":        "PRODUCTTYPE: Must be FERT or HAWA — invalid or blank value found",
            "PRODUCTHIERARCHY":   "PRODUCTHIERARCHY: Field is blank",
            "BASEUNIT":           "BASEUNIT: Field is blank",
            "MRPTYPE":            "MRPTYPE: Field is blank",
            "PROCUREMENTTYPE":    "PROCUREMENTTYPE: Field is blank",
            "IBPSTATUS":          "IBPSTATUS: Must be 'IBP' or blank — unexpected value found",
            "XPLANTMATSTATUS":    "XPLANTMATSTATUS: Must be '2' or blank — unexpected value found",
        }

        # Title
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="Part Master FG Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        headers = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 3
        for field_num, col_name in enumerate(RULES_FIELDS_ORDERED, start=1):
            count       = col_error_counts.get(col_name, 0)
            pct_error   = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health  = round(100 - pct_error, 2)
            reason_text = REASON_MAP.get(col_name, "") if count > 0 else ""

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).font      = BODY_FONT
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                align = Alignment(
                    horizontal="left" if c == 7 else "center", vertical="center"
                )
                ws.cell(row=row_num, column=c).alignment = align
            row_num += 1

        # TOTAL row
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(RULES_FIELDS_ORDERED)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c_idx, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""],
            start=1,
        ):
            cell           = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font      = Font(name="Arial", bold=True)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = BODY_FONT
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        self._auto_width(ws, min_w=8, max_w=70)

    # ── Rule_Set sheet ────────────────────────
    def _write_ruleset_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Part Master FG – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in self.RULES_CONTENT.items():
            num_rules = len(rules_list)
            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.fill      = RULE_FILL
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                fld_cell           = ws.cell(row=current_row, column=2,
                                             value=field if r_idx == 0 else "")
                fld_cell.fill      = RULE_FILL
                fld_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                fld_cell.border    = THIN_BORDER
                fld_cell.alignment = Alignment(horizontal="center", vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center",
                                                horizontal="center")
                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 65

    # ── Field Error Sheets ───────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        """
        FEATURE 2: Only columns present in RULES_FIELDS_ORDERED are shown per sheet.
        FEATURE 3: Sheet order = ERROR_SHEET_PRIORITY first, then remaining fields.
        ABCINDICATOR has been fully removed from all processing.
        """
        v = self.validator

        # Only keep columns that exist in both the source data AND the rules list
        rules_cols_in_data = [c for c in RULES_FIELDS_ORDERED if c in df.columns]

        # Collect all fields that have at least one error row
        all_error_fields: set = set()
        for bad_cols in v.error_map.values():
            all_error_fields.update(bad_cols.keys())

        # Build ordered list: priority fields first, then the rest alphabetically
        ordered_fields = []
        for f in ERROR_SHEET_PRIORITY:
            if f in all_error_fields:
                ordered_fields.append(f)
        for f in sorted(all_error_fields):
            if f not in ordered_fields:
                ordered_fields.append(f)

        for field_name in ordered_fields:
            row_indices = [
                idx for idx, errdict in v.error_map.items() if field_name in errdict
            ]
            if not row_indices:
                continue

            # Subset: only rules-sheet columns + appended error reason column
            subset = df.loc[row_indices, rules_cols_in_data].copy()
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(field_name, "")
            )

            # Safe sheet name (Excel max 31 chars)
            sheet_name = field_name[:31]
            existing   = [s.title for s in wb.worksheets]
            counter    = 1
            base_name  = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter   += 1

            ws = wb.create_sheet(sheet_name)
            self._write_header(ws, subset.columns)

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for r_idx, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(
                    zip(subset.columns, row_data), start=1
                ):
                    cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.fill = ROW_FILL

                # Highlight the failing column in red
                if field_name in col_idx_map:
                    target_cell      = ws.cell(row=r_idx, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._auto_width(ws, min_w=10, max_w=60)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Orchestrate Writing ──────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, total_rows=len(df))
        self._write_ruleset_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartTableProcessor:

    def __init__(self, input_path: str, output_path: str):
        self.validator = PartTableValidator(input_path, VALID_PLANTS)
        self.writer    = ReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PartTableProcessor(INPUT_FILE, OUTPUT_FILE)
    processor.run()
