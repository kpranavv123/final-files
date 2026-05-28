"""
Part UOM Technical Validator
═════════════════════════════════════════════════════════════════════════════

Purpose:
  Validates Part UOM Conversion file against technical/data-quality rules only.

Technical Rules:
  PRODUCT:
    1. PRODUCT should not be blank
    2. PRODUCT + PLANT combination should exist in Part Master

  PLANT:
    1. PLANT should not be blank
    2. PLANT should exist in Site Master

  CONVERSIONFACTOR:
    1. CONVERSIONFACTOR should not be blank
    2. CONVERSIONFACTOR should be numeric
    3. CONVERSIONFACTOR should be non-zero

  ALTERNATIVEUNIT:
    1. ALTERNATIVEUNIT should not be blank
    2. Each PRODUCT must have all five units: KG, CV, PAC, TO, NO
       (if any one is missing, all rows for that PRODUCT are flagged)

  DUPLICATE_CHECK:
    1. No duplicate PRODUCT + PLANT + ALTERNATIVEUNIT combinations allowed

Output:
  Validated_PartUOM_Technical.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
PART_UOM_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\PartUOM\PUOM_2026-05-21-1328.tab"
PART_REFERENCE_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-05-21-1510.tab"
SITE_REFERENCE_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-20-1153.tab"
OUTPUT_FILE         = r"C:\Users\SW526XH\Downloads\Go Live-1\PartUOM\Validated_PartUOM_Technical2.xlsx"

# ── CHANGE 1: Required units expanded to KG, CV, PAC, TO, NO ──────────────
REQUIRED_UNITS = {"KG", "CV", "PAC", "TO", "NO"}

# Canonical label for missing-units sub-row in Summary
MISSING_UNITS_SUMMARY_LABEL = "PRODUCT is missing required unit(s): KG, CV, PAC, TO or NO"

# Canonical label for duplicate check in Summary
DUPLICATE_SUMMARY_LABEL = "Duplicate PRODUCT + PLANT + ALTERNATIVEUNIT combination"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL  = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=9)
ERR_FONT    = Font(name="Arial", size=9, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── CHANGE 3: Canonical field order — drives Summary, Rulesets, error sheets
FIELD_ORDER = [
    "PRODUCT",
    "PLANT",
    "CONVERSIONFACTOR",
    "ALTERNATIVEUNIT",
    "DUPLICATE_CHECK",
]

# ══════════════════════════════════════════════
#  Technical Ruleset Info
#  (Drives BOTH the Rulesets sheet and the
#   Summary sub-row labels — keep in sync.)
# ══════════════════════════════════════════════
SUMMARY_RULESET_INFO = {
    "PRODUCT": [
        "PRODUCT is blank",
        "PRODUCT and PLANT combination not found in part master",
    ],
    "PLANT": [
        "PLANT is blank",
        "PLANT not found in site master",
    ],
    # CHANGE 3: CONVERSIONFACTOR moved before ALTERNATIVEUNIT
    "CONVERSIONFACTOR": [
        "CONVERSIONFACTOR is blank",
        "CONVERSIONFACTOR is not numeric",
        "CONVERSIONFACTOR must be non-zero",
    ],
    "ALTERNATIVEUNIT": [
        "ALTERNATIVEUNIT is blank",
        MISSING_UNITS_SUMMARY_LABEL,   # one canonical label for all missing-unit errors
    ],
    "DUPLICATE_CHECK": [
        DUPLICATE_SUMMARY_LABEL,
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class PartUOMTechnicalRuleEngine:
    """Technical validation rules for Part UOM Conversion."""

    def __init__(self, part_plant_combos: set, site_codes: set):
        self.part_plant_combos    = part_plant_combos
        self.site_codes           = set(str(s).strip() for s in site_codes)
        # product_str -> set of missing unit strings  (populated by precompute_missing_units)
        self.combo_missing_units: dict = {}
        # set of row indices that are duplicates    (populated by precompute_duplicates)
        self.duplicate_combos: set     = set()

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _is_numeric(self, value) -> bool:
        if self._is_blank(value):
            return False
        try:
            float(str(value).strip())
            return True
        except (ValueError, TypeError):
            return False

    def precompute_missing_units(self, df: pd.DataFrame):
        """
        For every PRODUCT, determine which of the required units
        (KG, CV, PAC, TO, NO) are absent across the whole file.
        Stores the exact missing set so error messages can be specific.
        """
        self.combo_missing_units = {}

        if "ALTERNATIVEUNIT" not in df.columns:
            return

        product_units: dict = {}

        for _, row in df.iterrows():
            product = row.get("PRODUCT", "")
            unit    = row.get("ALTERNATIVEUNIT", "")

            if self._is_blank(product):
                continue

            product_str = str(product).strip()
            unit_str    = str(unit).strip().upper() if not self._is_blank(unit) else ""

            product_units.setdefault(product_str, set())
            if unit_str:
                product_units[product_str].add(unit_str)

        for product, present_units in product_units.items():
            missing = REQUIRED_UNITS - present_units
            if missing:
                self.combo_missing_units[product] = missing

    def precompute_duplicates(self, df: pd.DataFrame):
        """
        Find every (PRODUCT, PLANT, ALTERNATIVEUNIT) combination that appears
        more than once. All occurrences (including the first) are flagged.
        """
        self.duplicate_combos = set()

        required_cols = {"PRODUCT", "PLANT", "ALTERNATIVEUNIT"}
        if not required_cols.issubset(df.columns):
            return

        seen: dict = {}

        for idx, row in df.iterrows():
            product  = row.get("PRODUCT", "")
            plant    = row.get("PLANT", "")
            alt_unit = row.get("ALTERNATIVEUNIT", "")

            if self._is_blank(product) or self._is_blank(plant) or self._is_blank(alt_unit):
                continue

            combo = (
                str(product).strip(),
                str(plant).strip(),
                str(alt_unit).strip().upper(),
            )

            if combo in seen:
                self.duplicate_combos.add(seen[combo])
                self.duplicate_combos.add(idx)
            else:
                seen[combo] = idx

    # ── Individual validators ──────────────────

    def validate_product(self, row) -> tuple:
        product = row.get("PRODUCT")
        plant   = row.get("PLANT")

        if self._is_blank(product):
            return False, "PRODUCT is blank"

        product_str = str(product).strip()
        plant_str   = str(plant).strip() if not self._is_blank(plant) else ""
        combo       = (product_str, plant_str)

        if combo in self.part_plant_combos:
            return True, ""

        return False, "PRODUCT and PLANT combination not found in part master"

    def validate_plant(self, row) -> tuple:
        plant = row.get("PLANT")

        if self._is_blank(plant):
            return False, "PLANT is blank"

        if str(plant).strip() in self.site_codes:
            return True, ""

        return False, "PLANT not found in site master"

    def validate_conversionfactor(self, row) -> tuple:
        cf = row.get("CONVERSIONFACTOR")

        if self._is_blank(cf):
            return False, "CONVERSIONFACTOR is blank"

        if not self._is_numeric(cf):
            return False, "CONVERSIONFACTOR is not numeric"

        try:
            if float(str(cf).strip()) == 0:
                return False, "CONVERSIONFACTOR must be non-zero"
        except (ValueError, TypeError):
            return False, "CONVERSIONFACTOR is not numeric"

        return True, ""

    def validate_alternativeunit(self, row) -> tuple:
        """
        Rule 1 – ALTERNATIVEUNIT must not be blank.
        Rule 2 – Every PRODUCT must have all required UOMs (KG, CV, PAC, TO, NO).

        CHANGE 2: The error message stored in error_map now names the specific
        missing units (e.g. "required unit(s) missing for this PRODUCT: TO, NO")
        so that the ERROR_FIELDS column in the error sheet is actionable.
        The summary sheet still collapses all such messages under the one
        canonical MISSING_UNITS_SUMMARY_LABEL via bucketing logic in _write_summary_sheet.
        """
        alt_unit = row.get("ALTERNATIVEUNIT")

        # Rule 1: blank check
        if self._is_blank(alt_unit):
            return False, "ALTERNATIVEUNIT is blank"

        # Rule 2: required-units check (product level)
        product = row.get("PRODUCT", "")
        if not self._is_blank(product):
            key     = str(product).strip()
            missing = self.combo_missing_units.get(key)

            if missing:
                # CHANGE 2: specific message listing the actual missing units
                sorted_missing = ", ".join(sorted(missing))
                return (
                    False,
                    f"Required unit(s) missing for this PRODUCT: {sorted_missing}",
                )

        return True, ""

    def validate_duplicate_check(self, row, idx) -> tuple:
        if idx in self.duplicate_combos:
            return False, DUPLICATE_SUMMARY_LABEL
        return True, ""

    def get_rules(self) -> dict:
        """Return field -> validator mapping in FIELD_ORDER sequence."""
        return {
            "PRODUCT":          self.validate_product,
            "PLANT":            self.validate_plant,
            "CONVERSIONFACTOR": self.validate_conversionfactor,
            "ALTERNATIVEUNIT":  self.validate_alternativeunit,
            # DUPLICATE_CHECK handled separately (needs idx) in the main loop
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartUOMTechnicalValidator:
    """Loads data, validates technical rules, builds error map."""

    def __init__(self):
        self.df                = pd.DataFrame()
        self.part_plant_combos = set()
        self.site_codes        = set()
        self.error_map         = {}

    def _add_error(self, idx, field_name: str, reason: str):
        self.error_map.setdefault(idx, {})

        if field_name in self.error_map[idx] and self.error_map[idx][field_name] != reason:
            existing = self.error_map[idx][field_name]
            if reason not in existing.split(" | "):
                self.error_map[idx][field_name] = existing + " | " + reason
        else:
            self.error_map[idx][field_name] = reason

    @staticmethod
    def _read_file(path: str) -> pd.DataFrame:
        lower = path.lower()
        if lower.endswith(".csv"):
            return pd.read_csv(path, dtype=str)
        elif lower.endswith((".tab", ".tsv")):
            return pd.read_csv(
                path, sep="\t", dtype=str,
                encoding="utf-8", encoding_errors="replace", engine="python",
            )
        elif lower.endswith((".xlsx", ".xls")):
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {path}")

    def load(self):
        print("[LOAD] Loading files...")

        # ── PartUOM input ────────────────────
        self.df = self._read_file(PART_UOM_INPUT_FILE)
        self.df.columns = [str(c).strip().upper() for c in self.df.columns]
        print(f"    PartUOM rows loaded: {len(self.df)}")

        # ── Part master reference ────────────
        part_df = self._read_file(PART_REFERENCE_FILE)
        part_df.columns = [str(c).strip().upper() for c in part_df.columns]

        mat_col = "MATERIALNUMBER" if "MATERIALNUMBER" in part_df.columns else "MATERIAL"
        if mat_col not in part_df.columns and "PRODUCT" in part_df.columns:
            mat_col = "PRODUCT"

        missing = [c for c in [mat_col, "PLANT"] if c not in part_df.columns]
        if missing:
            raise ValueError(
                f"Missing columns in Part master: {missing}. Available: {list(part_df.columns)}"
            )

        for _, row in part_df.iterrows():
            part_name = str(row[mat_col]).strip() if pd.notna(row[mat_col]) else ""
            plant     = str(row["PLANT"]).strip()  if pd.notna(row["PLANT"])  else ""
            if part_name and plant:
                self.part_plant_combos.add((part_name, plant))

        print(f"    Part master - PRODUCT/PLANT combos: {len(self.part_plant_combos)}")

        # ── Site master reference ────────────
        site_df = self._read_file(SITE_REFERENCE_FILE)
        site_df.columns = [str(c).strip().upper() for c in site_df.columns]

        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")

        self.site_codes = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site master - PLANT codes: {len(self.site_codes)}")

    def validate(self):
        print("[VALIDATE] Running technical validation rules...")

        engine = PartUOMTechnicalRuleEngine(self.part_plant_combos, self.site_codes)

        engine.precompute_missing_units(self.df)
        print(
            f"    PRODUCTs missing required units (KG/CV/PAC/TO/NO): "
            f"{len(engine.combo_missing_units)}"
        )

        engine.precompute_duplicates(self.df)
        print(f"    Duplicate rows (PRODUCT+PLANT+ALTERNATIVEUNIT): {len(engine.duplicate_combos)}")

        rules = engine.get_rules()

        for idx, row in self.df.iterrows():
            # Standard field-level rules — iterate in FIELD_ORDER for consistency
            for field in FIELD_ORDER:
                if field == "DUPLICATE_CHECK":
                    continue   # handled below
                rule_fn = rules.get(field)
                if rule_fn is None or field not in self.df.columns:
                    continue
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"Exception: {e}"

                if not passed:
                    self._add_error(idx, field, reason)

            # Duplicate check (needs idx)
            try:
                passed, reason = engine.validate_duplicate_check(row, idx)
            except Exception as e:
                passed, reason = False, f"Exception: {e}"

            if not passed:
                self._add_error(idx, "DUPLICATE_CHECK", reason)

    def get_error_series(self) -> pd.Series:
        details = {}
        for idx, errdict in self.error_map.items():
            messages = [f"{fld}: {msg}" for fld, msg in errdict.items()]
            details[idx] = "; ".join(messages)
        return pd.Series(details, dtype=str)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class PartUOMTechnicalReportWriter:
    """Builds technical validation workbook."""

    SHEET_SUMMARY  = "Summary"
    SHEET_RULESETS = "Rulesets"

    def __init__(self, validator: PartUOMTechnicalValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    def _safe_sheet_name(self, wb, base_name: str) -> str:
        invalid_chars = ["/", "\\", "*", "?", ":", "[", "]"]
        name = str(base_name)
        for ch in invalid_chars:
            name = name.replace(ch, "-")
        name = name.strip() or "Sheet"
        name = name[:31]
        if name not in wb.sheetnames:
            return name
        counter = 1
        while True:
            suffix    = f"_{counter}"
            candidate = f"{name[:31 - len(suffix)]}{suffix}"
            if candidate not in wb.sheetnames:
                return candidate
            counter += 1

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 4, 80))

    # ── Rulesets sheet ────────────────────────
    def _write_ruleset_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULESETS, 1)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Part UOM Table – Technical Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # CHANGE 3: ruleset content ordered to match FIELD_ORDER
        ruleset_info = {
            "PRODUCT": [
                "Must not be blank.",
                "PRODUCT and PLANT combination must exist in part master.",
            ],
            "PLANT": [
                "Must not be blank.",
                "Must exist in site master.",
            ],
            "CONVERSIONFACTOR": [
                "Must not be blank.",
                "Must be numeric.",
                "Must be non-zero.",
            ],
            # CHANGE 1: updated to list all five required units
            "ALTERNATIVEUNIT": [
                "Must not be blank.",
                "Each PRODUCT must contain all five required units: KG, CV, PAC, TO, and NO.",
            ],
            "DUPLICATE_CHECK": [
                "No duplicate combinations of PRODUCT + PLANT + ALTERNATIVEUNIT are allowed in the extract.",
            ],
        }

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            if field not in ruleset_info:
                continue

            rules_list = ruleset_info[field]
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 100

    # ── Summary sheet ─────────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="Part UOM Technical Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 24

        headers    = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        col_widths = [6, 28, 16, 16, 16, 16, 90]

        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Count errors per field ────────────
        col_error_counts: dict = {field: 0 for field in FIELD_ORDER}
        # key: (field, canonical_reason) -> count
        rule_error_counts: dict = {}

        for bad_cols in error_map.values():
            for col, reason in bad_cols.items():
                if col not in col_error_counts:
                    continue
                col_error_counts[col] += 1

                # CHANGE 2: all specific "Required unit(s) missing…" messages are bucketed
                # under the single canonical label for the Summary sub-row
                if col == "ALTERNATIVEUNIT" and reason.startswith("Required unit(s) missing"):
                    canonical_reason = MISSING_UNITS_SUMMARY_LABEL
                else:
                    canonical_reason = reason

                key = (col, canonical_reason)
                rule_error_counts[key] = rule_error_counts.get(key, 0) + 1

        row_num      = 4
        item_counter = 1

        for field_name in FIELD_ORDER:
            count         = col_error_counts.get(field_name, 0)
            static_reasons = SUMMARY_RULESET_INFO.get(field_name, [])
            is_multi       = len(static_reasons) > 1

            pct_err    = count / total_rows if total_rows else 0
            pct_health = 1 - pct_err

            if is_multi:
                # Parent row
                ws.cell(row=row_num, column=1, value=item_counter)
                ws.cell(row=row_num, column=2, value=field_name)
                ws.cell(row=row_num, column=3, value=count)
                ws.cell(row=row_num, column=4, value=total_rows)
                ws.cell(row=row_num, column=5, value=pct_health).number_format = "0.00%"
                ws.cell(row=row_num, column=6, value=pct_err).number_format    = "0.00%"
                ws.cell(row=row_num, column=7, value="")

                for c in range(1, 8):
                    cell           = ws.cell(row=row_num, column=c)
                    cell.border    = THIN_BORDER
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(horizontal="center" if c != 7 else "left")

                row_num += 1

                # Sub-rows — one per canonical reason
                for reason in static_reasons:
                    sub_count  = rule_error_counts.get((field_name, reason), 0)
                    sub_pct    = sub_count / total_rows if total_rows else 0

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=f"↳ {reason}")
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=1 - sub_pct).number_format = "0.00%"
                    ws.cell(row=row_num, column=6, value=sub_pct).number_format     = "0.00%"
                    ws.cell(row=row_num, column=7, value=reason if sub_count > 0 else "")

                    for c in range(1, 8):
                        cell           = ws.cell(row=row_num, column=c)
                        cell.border    = THIN_BORDER
                        cell.font      = BODY_FONT
                        cell.alignment = Alignment(
                            horizontal="center" if c != 7 else "left",
                            indent=(1 if c == 2 else 0),
                            wrap_text=(c == 7),
                        )

                    row_num += 1

            else:
                # Single-sub-row field
                reason  = static_reasons[0] if static_reasons else ""
                values  = [
                    item_counter, field_name, count, total_rows,
                    pct_health, pct_err,
                    reason if count > 0 else "",
                ]

                for c_idx, value in enumerate(values, start=1):
                    cell           = ws.cell(row=row_num, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center" if c_idx != 7 else "left",
                        wrap_text=(c_idx == 7),
                    )
                    if c_idx in [5, 6]:
                        cell.number_format = "0.00%"

                row_num += 1

            item_counter += 1

        # ── TOTAL row ────────────────────────
        total_errors      = sum(col_error_counts.values())
        sum_record_counts = len(FIELD_ORDER) * total_rows
        total_err_pct     = total_errors / sum_record_counts if sum_record_counts else 0

        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=sum_record_counts)
        ws.cell(row=row_num, column=5, value=1 - total_err_pct).number_format = "0.00%"
        ws.cell(row=row_num, column=6, value=total_err_pct).number_format     = "0.00%"

        for c in range(1, 8):
            cell           = ws.cell(row=row_num, column=c)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.font      = Font(name="Arial", bold=True)
            cell.alignment = Alignment(horizontal="center" if c != 7 else "left")

        row_num += 2

        # ── Stats block ──────────────────────
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center")
            row_num += 1

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        v = self.validator

        all_error_fields: set = set()
        for errdict in v.error_map.values():
            all_error_fields.update(errdict.keys())

        # CHANGE 3: error sheets created in FIELD_ORDER sequence
        for field_name in FIELD_ORDER:
            if field_name not in all_error_fields:
                continue

            row_indices = [idx for idx, errdict in v.error_map.items() if field_name in errdict]
            if not row_indices:
                continue

            sheet_name = self._safe_sheet_name(wb, field_name)
            ws         = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()

            # CHANGE 2: ERROR_FIELDS column carries the specific per-row reason,
            # which for ALTERNATIVEUNIT now names the exact missing units
            subset["ERROR_FIELDS"] = subset.index.map(
                lambda i, fn=field_name: v.error_map.get(i, {}).get(fn, "")
            )

            # CHANGE 3: display columns follow FIELD_ORDER (data cols only, no DUPLICATE_CHECK)
            data_cols    = [c for c in FIELD_ORDER if c != "DUPLICATE_CHECK" and c in df.columns]
            display_cols = data_cols + ["ERROR_FIELDS"]
            subset       = subset[[c for c in display_cols if c in subset.columns]]

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.fill      = ROW_FILL
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border    = THIN_BORDER

                # Highlight the offending column(s) in red
                if field_name == "DUPLICATE_CHECK":
                    for dup_col in ("PRODUCT", "PLANT", "ALTERNATIVEUNIT"):
                        if dup_col in col_idx_map:
                            target_cell      = ws.cell(row=excel_row, column=col_idx_map[dup_col])
                            target_cell.fill = RED_FILL
                            target_cell.font = ERR_FONT
                elif field_name in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Main write entry point ────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series       = v.get_error_series()
        df["ERROR_FIELDS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        # CHANGE 3: keep only columns in FIELD_ORDER sequence (data cols) + ERROR_FIELDS
        data_cols     = [c for c in FIELD_ORDER if c != "DUPLICATE_CHECK" and c in df.columns]
        keep_cols     = data_cols + ["ERROR_FIELDS"]
        df            = df[[c for c in keep_cols if c in df.columns]]

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb, v.error_map, len(df))
        self._write_ruleset_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        print(f"\n[SAVE] Technical output saved: {self.output_path}")
        print(f"   Total records    : {len(df)}")
        print(f"   Error records    : {len(v.error_map)}")
        print(f"   Passing records  : {len(df) - len(v.error_map)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartUOMTechnicalProcessor:
    def __init__(self):
        self.validator = PartUOMTechnicalValidator()
        self.writer    = PartUOMTechnicalReportWriter(self.validator, OUTPUT_FILE)

    def run(self):
        print("=" * 70)
        print("Part UOM Technical Validation Pipeline")
        print("=" * 70)

        self.validator.load()
        self.validator.validate()
        self.writer.write()

        print("\n" + "=" * 70)
        print("Technical Validation Complete!")
        print("=" * 70)


if __name__ == "__main__":
    processor = PartUOMTechnicalProcessor()
    processor.run()
