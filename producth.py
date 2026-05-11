import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\ProductH\ProductHierarchy_2026-05-06-1313.tab"
OUTPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\ProductH\ProductHierarchy_Business_Rules.xlsx"


# ─────────────────────────────────────────────
#  PARENT-CHILD PAIR DEFINITIONS
#  Rule: one child value must map to exactly one unique parent value.
#  Tuple order: (parent_col, child_col)
# ─────────────────────────────────────────────
PARENT_CHILD_PAIRS = [
    ("CATEGORYDESCRIPTION",    "PRODUCTDESCRIPTION"),
    ("PRODUCTDESCRIPTION",     "VARIANTDESCRIPTION"),
    ("VARIANTDESCRIPTION",     "BRANDDESCRIPTION"),
    ("BRANDDESCRIPTION",       "SUBBRANDDESCRIPTION"),
    ("SUBBRANDDESCRIPTION",    "BRANDVARIANTDESCRIPTION"),
    ("BRANDVARIANTDESCRIPTION","PACKSIZEDESCRIPTION"),
    ("PACKSIZEDESCRIPTION",    "MARKETSKUDESCRIPTION"),
    ("MARKETSKUDESCRIPTION",   "MATERIALNUMBER"),
]

# Sheet name for each pair  (Excel max 31 chars)
def _sheet_name(parent_col: str, child_col: str) -> str:
    raw = f"{child_col}__{parent_col}"
    return raw[:31]

# Reason text per pair
def _reason(parent_col: str, child_col: str) -> str:
    return (
        f"{child_col}: mapped to multiple {parent_col} values — "
        f"one {child_col} must map to exactly one {parent_col}"
    )

# Rules sheet description lines per pair
def _rule_lines(parent_col: str, child_col: str) -> list:
    return [
        f"Each unique {child_col} value must map to exactly one {parent_col} value.",
        f"All rows where {child_col} appears with more than one distinct {parent_col} are flagged.",
        f"Both the {parent_col} and {child_col} cells are highlighted in red.",
    ]


# ─────────────────────────────────────────────
#  STYLING CONSTANTS  (identical across all scripts)
# ─────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FF0000")
ROW_FILL   = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL   = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL  = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
STATS_FILL = PatternFill("solid", fgColor="EDEDED")

HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class PHBusinessRuleEngine:
    """
    For each parent-child pair, builds child → set(distinct parents) map.
    Any child with more than one distinct parent is a violation.
    ALL rows in the violating group are flagged (keep=False logic).

    Returns:
        pc_error_map : dict
            {
              (parent_col, child_col) : {
                  row_index : reason_string,
                  ...
              }
            }
    """

    @staticmethod
    def _is_blank(value) -> bool:
        if value is None:
            return True
        try:
            import math
            if math.isnan(float(value)):
                return True
        except (TypeError, ValueError):
            pass
        return str(value).strip() == ""

    def run(self, df: pd.DataFrame) -> dict:
        pc_error_map: dict = {}

        for parent_col, child_col in PARENT_CHILD_PAIRS:
            if parent_col not in df.columns or child_col not in df.columns:
                continue

            # Build child → set of unique non-blank parent values
            child_to_parents: dict[str, set] = defaultdict(set)
            for _, row in df.iterrows():
                p_val = str(row[parent_col]).strip() if not self._is_blank(row[parent_col]) else ""
                c_val = str(row[child_col]).strip()  if not self._is_blank(row[child_col])  else ""
                if c_val and p_val:
                    child_to_parents[c_val].add(p_val)

            # Children that map to more than one parent
            violating_children = {
                c for c, parents in child_to_parents.items() if len(parents) > 1
            }

            if not violating_children:
                continue

            pair_errors: dict = {}
            reason_text = _reason(parent_col, child_col)

            for idx, row in df.iterrows():
                c_val = str(row[child_col]).strip() if not self._is_blank(row[child_col]) else ""
                if c_val in violating_children:
                    pair_errors[idx] = reason_text

            if pair_errors:
                pc_error_map[(parent_col, child_col)] = pair_errors

        return pc_error_map


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PHBusinessValidator:

    def __init__(self, filepath: str):
        self.filepath     = filepath
        self.df           = pd.DataFrame()
        self.pc_error_map = {}   # { (parent_col, child_col) : { row_idx : reason } }

    def load(self):
        path = self.filepath.lower()
        if path.endswith(".csv"):
            self.df = pd.read_csv(self.filepath, dtype=str)
        elif path.endswith(".tab") or path.endswith(".tsv"):
            self.df = pd.read_csv(
                self.filepath,
                sep="\t",
                dtype=str,
                encoding="latin-1",
                engine="python",
            )
        elif path.endswith(".xlsx") or path.endswith(".xls"):
            self.df = pd.read_excel(self.filepath, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {self.filepath}")

        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        engine            = PHBusinessRuleEngine()
        self.pc_error_map = engine.run(self.df)

    # Convenience: all row indices that have at least one PC error
    def all_error_indices(self) -> set:
        indices = set()
        for pair_errors in self.pc_error_map.values():
            indices.update(pair_errors.keys())
        return indices


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class PHBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    def __init__(self, validator: PHBusinessValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── Helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _auto_width(self, ws, min_w=10, max_w=60):
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(length + 3, min_w), max_w
            )

    def _safe_sheet_name(self, wb, name: str) -> str:
        name     = name[:31]
        existing = [s.title for s in wb.worksheets]
        base     = name
        counter  = 1
        while name in existing:
            name = f"{base[:28]}_{counter}"
            counter += 1
        return name

    # ── Summary Sheet ─────────────────────────
    def _write_summary_sheet(self, wb):
        ws         = wb.create_sheet(self.SHEET_SUMMARY)
        v          = self.validator
        total_rows = len(v.df)

        # ── Title ──
        ws.merge_cells("A1:G1")
        tc           = ws.cell(row=1, column=1,
                               value="Product Hierarchy – Business Rules Validation Summary")
        tc.font      = Font(name="Arial", bold=True, size=14)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Column headers ──
        headers = ["#", "Parent → Child Pair", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        row_num    = 3
        rule_num   = 1
        all_counts = []

        # One data row per PC pair (only pairs defined in PARENT_CHILD_PAIRS, in order)
        for parent_col, child_col in PARENT_CHILD_PAIRS:
            pair_errors = v.pc_error_map.get((parent_col, child_col), {})
            count       = len(pair_errors)
            all_counts.append(count)

            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason     = _reason(parent_col, child_col) if count > 0 else ""

            label = f"{parent_col}  →  {child_col}"

            ws.cell(row=row_num, column=1, value=rule_num)
            ws.cell(row=row_num, column=2, value=label)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason)

            for c in range(1, 8):
                cell           = ws.cell(row=row_num, column=c)
                cell.font      = BODY_FONT
                cell.fill      = WHITE_FILL
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="left" if c in (2, 7) else "center",
                    vertical="center",
                    wrap_text=(c == 7),
                )

            row_num  += 1
            rule_num += 1

        # ── TOTAL row ──
        total_errors       = sum(all_counts)
        total_record_count = total_rows * len(PARENT_CHILD_PAIRS)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c_idx, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""],
            start=1,
        ):
            cell           = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # ── Stats block ──
        all_error_indices   = v.all_error_indices()
        records_with_errors = len(all_error_indices)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=2)
            lc           = ws.cell(row=row_num, column=1, value=label)
            lc.font      = Font(name="Arial", bold=True, size=10)
            lc.fill      = STATS_FILL
            lc.border    = THIN_BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")

            vc           = ws.cell(row=row_num, column=3, value=value)
            vc.font      = BODY_FONT
            vc.border    = THIN_BORDER
            vc.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        # ── Column widths ──
        col_widths = [6, 55, 14, 16, 12, 12, 75]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Rules Sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        tc           = ws.cell(row=1, column=1,
                               value="Product Hierarchy – Business Validation Rules")
        tc.font      = Font(name="Arial", bold=True, size=13)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Sub-title explaining the main rule
        ws.merge_cells("A2:C2")
        sc           = ws.cell(row=2, column=1,
                               value="Main Rule: Each child value must map to exactly one parent value. "
                                     "All occurrences of a violating child are flagged.")
        sc.font      = Font(name="Arial", italic=True, size=10)
        sc.fill      = RULE_FILL
        sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 20

        for c_idx, h in enumerate(["#", "Parent  →  Child Pair", "Rule Description"], start=1):
            cell           = ws.cell(row=4, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = 5
        for rule_num, (parent_col, child_col) in enumerate(PARENT_CHILD_PAIRS, start=1):
            rules_list = _rule_lines(parent_col, child_col)
            num_rules  = len(rules_list)
            pair_label = f"{parent_col}  →  {child_col}"

            for r_idx, rule_text in enumerate(rules_list):
                nc           = ws.cell(row=current_row, column=1,
                                       value=rule_num if r_idx == 0 else "")
                nc.fill      = RULE_FILL
                nc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                nc.border    = THIN_BORDER
                nc.alignment = Alignment(horizontal="center", vertical="center")

                fc           = ws.cell(row=current_row, column=2,
                                       value=pair_label if r_idx == 0 else "")
                fc.fill      = RULE_FILL
                fc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                fc.border    = THIN_BORDER
                fc.alignment = Alignment(horizontal="left", vertical="center",
                                         wrap_text=True)

                dc           = ws.cell(row=current_row, column=3, value=rule_text)
                dc.font      = BODY_FONT
                dc.border    = THIN_BORDER
                dc.alignment = Alignment(wrap_text=True, vertical="center",
                                         horizontal="left")
                current_row += 1

            # Vertically merge # and pair columns across rule lines
            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 70

    # ── Error Sheets — one per PC pair ────────
    def _write_error_sheets(self, wb):
        """
        One sheet per violating PC pair (only created when violations exist).
        Sheet order mirrors PARENT_CHILD_PAIRS definition order.
        ALL columns from the input file are shown.
        Both parent and child cells are highlighted red on each error row.
        """
        v              = self.validator
        all_input_cols = list(v.df.columns)

        for parent_col, child_col in PARENT_CHILD_PAIRS:
            pair_errors = v.pc_error_map.get((parent_col, child_col))
            if not pair_errors:
                continue

            row_indices = list(pair_errors.keys())

            # Subset of the full dataframe — all input columns
            subset = v.df.loc[row_indices, all_input_cols].copy()

            # Append reason column
            subset["ERROR_REASON"] = subset.index.map(
                lambda i: pair_errors.get(i, "")
            )

            sheet_name = self._safe_sheet_name(
                wb, _sheet_name(parent_col, child_col)
            )
            ws = wb.create_sheet(sheet_name)

            all_cols    = list(subset.columns)
            col_idx_map = {col: i for i, col in enumerate(all_cols, start=1)}

            self._write_header(ws, all_cols)

            for r_idx, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for col, value in zip(all_cols, row_data):
                    c_idx          = col_idx_map[col]
                    cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.fill = ROW_FILL

                # Red highlight on both parent and child columns
                for highlight_col in (parent_col, child_col):
                    if highlight_col in col_idx_map:
                        tc      = ws.cell(row=r_idx, column=col_idx_map[highlight_col])
                        tc.fill = RED_FILL
                        tc.font = ERR_FONT

            self._auto_width(ws, min_w=10, max_w=60)
            ws.freeze_panes = "A2"

            note_row = len(row_indices) + 3
            ws.cell(
                row=note_row, column=1,
                value=(
                    f"Total error rows for '{child_col} → {parent_col}' "
                    f"mapping violation: {len(row_indices)}"
                ),
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Orchestrate ───────────────────────────
    def write(self):
        v  = self.validator
        wb = Workbook()

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb)
        self._write_rules_sheet(wb)
        self._write_error_sheets(wb)

        wb.save(self.output_path)

        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows            : {len(v.df)}")
        print(f"   Rows with any error   : {len(v.all_error_indices())}")
        print()
        print("   Per-pair violation counts:")
        for parent_col, child_col in PARENT_CHILD_PAIRS:
            count = len(v.pc_error_map.get((parent_col, child_col), {}))
            if count:
                print(f"     {parent_col:30s} → {child_col:30s} : {count} rows")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PHBusinessProcessor:

    def __init__(self, input_path: str, output_path: str):
        self.validator = PHBusinessValidator(input_path)
        self.writer    = PHBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Columns detected : {list(self.validator.df.columns)}")
        print("🔍  Running parent-child mapping rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PHBusinessProcessor(INPUT_FILE, OUTPUT_FILE)
    processor.run()
