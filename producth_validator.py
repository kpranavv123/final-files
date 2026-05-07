"""
PRODUCTHIERARCHY (FG) - Excel Data Validation Tool
Validates fields based on material type rules AND parent-child mapping integrity.
Generates a formatted Excel report.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import Optional
import os


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
INPUT_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-1\ProductH\ProductHierarchy_2026-05-06-1313.tab"
OUTPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\ProductH\Validated_Product Hierarchy2.xlsx"

# ─────────────────────────────────────────────
#  CONSTANTS & CONFIGURATION
# ─────────────────────────────────────────────

VALID_MATERIAL_TYPES = {"FERT", "HAWA"}
VALID_IBP_STATUSES   = {"IBP", ""}

NOT_BLANK_FIELDS = [
    "MATERIALNUMBER", "MATERIALDESCRIPTION", "PRODUCTGROUP", "MATLGRPDESC",
    "DIVISION", "DIVISIONDESCRIPTION", "PRODUCTTYPE", "PRODUCT_HIERARCHY_KEY",
    "CATEGORY", "CATEGORYDESCRIPTION", "PRODUCT", "PRODUCTDESCRIPTION",
    "VARIANT", "VARIANTDESCRIPTION", "BRAND", "BRANDDESCRIPTION",
    "SUBBRAND", "SUBBRANDDESCRIPTION", "BRANDVARIANT", "BRANDVARIANTDESCRIPTION",
    "PACKSIZE", "PACKSIZEDESCRIPTION", "MARKETSKU", "MARKETSKUDESCRIPTION",
    "SUPPLY_FAMILY",
]

# ── Parent-child pair definitions ────────────────────────────────────────────
# Rule: one child value must map to exactly one unique parent value.
# Tuple order: (parent_col, child_col)
PARENT_CHILD_PAIRS = [
    ("CATEGORYDESCRIPTION",  "PRODUCTDESCRIPTION"),
    ("VARIANTDESCRIPTION",   "BRANDDESCRIPTION"),
    ("SUBBRANDDESCRIPTION",  "BRANDVARIANTDESCRIPTION"),
    ("PACKSIZEDESCRIPTION",  "MARKETSKUDESCRIPTION"),
]

# Human-readable reason messages for PC errors (keyed by child column)
PC_REASON_MAP = {
    child: (
        f"{child}: mapped to multiple {parent} values. "
        f"One {child} must map to exactly one {parent}."
    )
    for parent, child in PARENT_CHILD_PAIRS
}

# Rules to display in the Rule_Set sheet for PC pairs
PC_RULES_CONTENT = {
    f"{child} → {parent}": [
        f"Each unique {child} value must map to exactly one {parent} value.",
        f"Rows where {child} appears with more than one distinct {parent} are flagged as errors.",
    ]
    for parent, child in PARENT_CHILD_PAIRS
}

RULES_CONTENT = {
    **{f: ["Must not be blank for FERT/HAWA material types."] for f in NOT_BLANK_FIELDS},
    "MATERIALTYPE": [
        "Must not be blank.",
        "Value must be either FERT or HAWA.",
    ],
    "IBPSTATUS": [
        "Allowed values: IBP or blank.",
        "Any other value is treated as an error.",
    ],
    **PC_RULES_CONTENT,
}

# Centralised reason messages for field-level errors
FIELD_REASON_MAP = {
    **{f: f"{f}: is blank for FERT/HAWA material types." for f in NOT_BLANK_FIELDS},
    "MATERIALTYPE": "MATERIALTYPE: is blank for FERT/HAWA material types",
    "IBPSTATUS":    "IBPSTATUS: Invalid value – must be 'IBP' or blank.",
}

# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL       = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_ERROR_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL       = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL      = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL     = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
WHITE_FILL     = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
TOTAL_FILL     = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL     = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─────────────────────────────────────────────
#  HELPER UTILITIES
# ─────────────────────────────────────────────

def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float):
        import math
        return math.isnan(value)
    return str(value).strip() == ""


def style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell           = ws.cell(row=row, column=col)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER


def auto_width(ws, min_w=10, max_w=60):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 3, min_w), max_w)


# ─────────────────────────────────────────────
#  VALIDATION ENGINE  (field-level)
# ─────────────────────────────────────────────

class FieldValidator:
    def __init__(self, field: str, rule_description: str):
        self.field            = field
        self.rule_description = rule_description

    def validate(self, row: pd.Series) -> Optional[str]:
        raise NotImplementedError


class NotBlankValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        mat_type = str(row.get("MATERIALTYPE", "")).strip().upper()
        if mat_type not in VALID_MATERIAL_TYPES:
            return None
        value = row.get(self.field)
        if is_blank(value):
            return FIELD_REASON_MAP.get(self.field, f"{self.field}: is blank for FERT/HAWA material types.")
        return None


class MaterialTypeValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        value = str(row.get("MATERIALTYPE", "")).strip().upper()
        if is_blank(value) or value not in VALID_MATERIAL_TYPES:
            return FIELD_REASON_MAP["MATERIALTYPE"]
        return None


class IBPStatusValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        raw   = row.get("IBPSTATUS", "")
        value = "" if is_blank(raw) else str(raw).strip()
        if value not in VALID_IBP_STATUSES and value.upper() not in VALID_IBP_STATUSES:
            return FIELD_REASON_MAP["IBPSTATUS"]
        return None


# ─────────────────────────────────────────────
#  VALIDATION ORCHESTRATOR
# ─────────────────────────────────────────────

class ProductHierarchyValidator:

    def __init__(self):
        self.validators: list[FieldValidator] = self._build_validators()

    def _build_validators(self) -> list[FieldValidator]:
        validators = [
            NotBlankValidator(f, "For FERT/HAWA materials – Field must not be blank")
            for f in NOT_BLANK_FIELDS
        ]
        validators.append(MaterialTypeValidator("MATERIALTYPE", "Field must be FERT or HAWA and must not be blank"))
        validators.append(IBPStatusValidator("IBPSTATUS",       "Field values must be either 'IBP' or blank"))
        return validators

    def validate_row(self, row: pd.Series) -> list[str]:
        errors = []
        for v in self.validators:
            if v.field in row.index or v.field in ("MATERIALTYPE", "IBPSTATUS"):
                result = v.validate(row)
                if result:
                    errors.append(result)
        return errors

    # ── Parent-child mapping validation (dataframe-level) ────────────────────
    def validate_parent_child_mappings(self, df: pd.DataFrame) -> pd.Series:
        """
        For each PC pair, build child→set(parents) map.
        Returns a Series (same index as df) where each element is a set of
        (parent_col, child_col) tuples for pairs that are violated on that row.
        Initialise with empty sets so non-error rows carry no overhead.
        """
        pc_errors: dict[int, set] = defaultdict(set)

        for parent_col, child_col in PARENT_CHILD_PAIRS:
            if parent_col not in df.columns or child_col not in df.columns:
                continue

            # Build child → set of unique parents (ignoring blank values)
            child_to_parents: dict[str, set] = defaultdict(set)
            for idx, row in df.iterrows():
                p_val = str(row[parent_col]).strip() if not is_blank(row[parent_col]) else ""
                c_val = str(row[child_col]).strip()  if not is_blank(row[child_col])  else ""
                if c_val and p_val:
                    child_to_parents[c_val].add(p_val)

            # Flag rows where child has >1 unique parent
            violating_children = {c for c, parents in child_to_parents.items() if len(parents) > 1}

            for idx, row in df.iterrows():
                c_val = str(row[child_col]).strip() if not is_blank(row[child_col]) else ""
                if c_val in violating_children:
                    pc_errors[idx].add((parent_col, child_col))

        # Return as a Series aligned to df.index, defaulting to empty set
        return pd.Series(
            [pc_errors.get(idx, set()) for idx in df.index],
            index=df.index,
        )

    def validate_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = df.columns.str.strip().str.upper()

        error_list        = []
        error_fields_list = []
        error_detail_list = []

        for _, row in df.iterrows():
            row_errors = self.validate_row(row)
            error_list.append(" | ".join(row_errors) if row_errors else "")

            fields_in_error = set()
            field_reason    = {}
            for e in row_errors:
                field = e.split(":")[0].strip()
                fields_in_error.add(field)
                field_reason[field] = e

            error_fields_list.append(fields_in_error)
            error_detail_list.append(field_reason)

        df["ERROR_COLUMNS"] = error_list
        df["_ERROR_FIELDS"] = error_fields_list
        df["_ERROR_DETAIL"] = error_detail_list

        # Parent-child errors stored separately – set of (parent_col, child_col) per row
        df["_PC_ERROR_PAIRS"] = self.validate_parent_child_mappings(df).values

        return df


# ─────────────────────────────────────────────
#  EXCEL REPORT BUILDER
# ─────────────────────────────────────────────

class ExcelReportBuilder:

    def __init__(self, df_validated: pd.DataFrame, output_path: str):
        self.df           = df_validated
        self.output       = output_path
        self.wb           = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

        self.error_df     = df_validated[df_validated["ERROR_COLUMNS"] != ""].copy()
        self.error_fields = df_validated["_ERROR_FIELDS"]
        self.error_detail = df_validated["_ERROR_DETAIL"]
        self.pc_errors    = df_validated["_PC_ERROR_PAIRS"]   # Series of sets of tuples

    def build(self):
        self._write_main_sheet()
        self._write_summary_sheet()
        self._write_ruleset_sheet()
        self._write_per_field_error_sheets()
        self._write_parent_child_error_sheets()
        self.wb.save(self.output)
        print(f"\n✅  Report saved → {self.output}")

    # ── Sheet 1: PRODUCTHIERARCHY_FG ────────────────────

    def _write_main_sheet(self):
        ws         = self.wb.create_sheet("PRODUCTHIERARCHY_FG")
        display_df = self.df.drop(columns=["_ERROR_FIELDS", "_ERROR_DETAIL", "_PC_ERROR_PAIRS"])
        headers    = list(display_df.columns)

        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"

        for r_idx, (orig_idx, row) in enumerate(display_df.iterrows(), start=2):
            df_row_pos     = r_idx - 2
            has_error      = bool(self.error_fields.iloc[df_row_pos])
            errored_fields = self.error_fields.iloc[df_row_pos]
            pc_pairs       = self.pc_errors.iloc[df_row_pos]  # set of (parent, child) tuples

            # Collect all columns that are part of a PC violation on this row
            pc_red_cols = set()
            for parent_col, child_col in pc_pairs:
                pc_red_cols.add(parent_col)
                pc_red_cols.add(child_col)

            has_pc_error = bool(pc_pairs)

            for c_idx, col in enumerate(headers, start=1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=row[col])
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if col in errored_fields:
                    # Field-level error takes priority
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT
                elif col in pc_red_cols:
                    # PC error – both parent and child cols highlighted red
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT
                elif has_error or has_pc_error:
                    cell.fill = ROW_ERROR_FILL
                else:
                    cell.fill = WHITE_FILL

        auto_width(ws)
        ws.row_dimensions[1].height = 30

    # ── Summary sheet ────────────────────────────────────

    def _write_summary_sheet(self):
        ws = self.wb.create_sheet("Summary")

        total_rows          = len(self.df)
        records_with_errors = len(self.error_df)

        # Also count rows with ONLY PC errors (no field errors) for the stats block
        pc_error_only_rows = sum(
            1 for idx in self.df.index
            if self.pc_errors.loc[idx] and not self.error_fields.loc[idx]
        )
        records_with_any_error = (
            self.df[
                (self.df["ERROR_COLUMNS"] != "") |
                (self.df["_PC_ERROR_PAIRS"].apply(bool))
            ].shape[0]
        )
        records_passing = total_rows - records_with_any_error

        # ── Field-level counts ────────────────────────────
        field_counts: dict[str, int] = defaultdict(int)
        for fields_set in self.error_fields:
            for f in fields_set:
                field_counts[f] += 1

        # ── PC-level counts (keyed by child_col) ─────────
        pc_counts: dict[str, int] = defaultdict(int)
        for pc_set in self.pc_errors:
            for parent_col, child_col in pc_set:
                pc_counts[child_col] += 1

        # ── Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="ProductHierarchy FG Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # ── Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 3

        # ── Per-field rows ──
        for field_num, (col_name, count) in enumerate(sorted(field_counts.items()), start=1):
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason     = FIELD_REASON_MAP.get(col_name, f"{col_name}: is blank for FERT/HAWA material types.")

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason)

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).font   = BODY_FONT
                ws.cell(row=row_num, column=c).border = THIN_BORDER
                align = Alignment(horizontal="center", vertical="center")
                if c == 7:
                    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=c).alignment = align

            row_num += 1

        # ── Per-PC-pair rows (appended after field rows) ──
        pc_field_offset = len(field_counts)
        for pc_num, (child_col, count) in enumerate(sorted(pc_counts.items()), start=1):
            # Find corresponding parent
            parent_col = next(
                (p for p, c in PARENT_CHILD_PAIRS if c == child_col), "Unknown"
            )
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason     = PC_REASON_MAP.get(child_col, f"{child_col}: mapped to multiple {parent_col} values.")

            ws.cell(row=row_num, column=1, value=pc_field_offset + pc_num)
            ws.cell(row=row_num, column=2, value=f"{child_col} → {parent_col}")
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason)

            for c in range(1, 8):
                ws.cell(row=row_num, column=c).font   = BODY_FONT
                ws.cell(row=row_num, column=c).border = THIN_BORDER
                align = Alignment(horizontal="center", vertical="center")
                if c == 7:
                    align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.cell(row=row_num, column=c).alignment = align

            row_num += 1

        # ── TOTAL row (field errors + PC errors combined) ──
        total_field_errors = sum(field_counts.values())
        total_pc_errors    = sum(pc_counts.values())
        total_errors       = total_field_errors + total_pc_errors

        total_field_record_count = total_rows * len(field_counts)
        total_pc_record_count    = total_rows * len(pc_counts)
        total_record_count       = total_field_record_count + total_pc_record_count

        total_pct_error  = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # ── Stats block ──
        for label, value in [
            ("Total Records:",              total_rows),
            ("Records with Errors:",        records_with_any_error),
            ("Records Passing:",            records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = BODY_FONT
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        auto_width(ws, min_w=8, max_w=70)

    # ── Rule_Set sheet ───────────────────────────────────

    def _write_ruleset_sheet(self):
        ws = self.wb.create_sheet("Rule_Set")

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="ProductHierarchy FG – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in RULES_CONTENT.items():
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1, value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2, value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(horizontal="center", vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 65

    # ── Per-field error sheets ───────────────────────────

    def _write_per_field_error_sheets(self):
        if self.error_df.empty:
            return

        field_rows: dict[str, list] = defaultdict(list)
        for idx, row in self.error_df.iterrows():
            for field in self.error_fields.loc[idx]:
                field_rows[field].append(idx)

        display_cols = [c for c in self.df.columns if c not in ("_ERROR_FIELDS", "_ERROR_DETAIL", "_PC_ERROR_PAIRS")]

        for field, row_indices in sorted(field_rows.items()):
            sheet_name = field[:31]
            existing   = [s.title for s in self.wb.worksheets]
            counter    = 1
            base_name  = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter   += 1

            ws = self.wb.create_sheet(sheet_name)

            subset = self.df.loc[row_indices, display_cols].copy()
            subset["ERROR_COLUMNS"] = FIELD_REASON_MAP.get(
                field, f"{field}: is blank for FERT/HAWA material types."
            )

            final_cols = list(subset.columns)
            ws.append(final_cols)
            style_header_row(ws, 1, len(final_cols))
            ws.freeze_panes = "A2"

            col_idx_map = {col: i for i, col in enumerate(final_cols, start=1)}

            for r_idx, (orig_idx, row) in enumerate(subset.iterrows(), start=2):
                for c_idx, col in enumerate(final_cols, start=1):
                    cell           = ws.cell(row=r_idx, column=c_idx, value=row[col])
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.fill      = ROW_ERROR_FILL

                if field in col_idx_map:
                    target_cell      = ws.cell(row=r_idx, column=col_idx_map[field])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            note_row = len(row_indices) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field}': {len(row_indices)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

            auto_width(ws, min_w=10, max_w=60)

    # ── Parent-child error sheets ────────────────────────

    def _write_parent_child_error_sheets(self):
        """
        One sheet per PC pair that has violations.
        Sheet name: PC_{child_col} (max 31 chars).
        Both parent and child columns are highlighted red.
        Overwrites ERROR_COLUMNS with the PC reason message.
        """
        display_cols = [c for c in self.df.columns if c not in ("_ERROR_FIELDS", "_ERROR_DETAIL", "_PC_ERROR_PAIRS")]

        for parent_col, child_col in PARENT_CHILD_PAIRS:
            # Collect row indices where this specific PC pair is violated
            row_indices = [
                idx for idx in self.df.index
                if (parent_col, child_col) in self.pc_errors.loc[idx]
            ]

            if not row_indices:
                continue

            raw_name   = f"PC_{child_col}"
            sheet_name = raw_name[:31]
            existing   = [s.title for s in self.wb.worksheets]
            counter    = 1
            base_name  = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base_name[:28]}_{counter}"
                counter   += 1

            ws = self.wb.create_sheet(sheet_name)

            subset = self.df.loc[row_indices, display_cols].copy()
            subset["ERROR_COLUMNS"] = PC_REASON_MAP.get(
                child_col,
                f"{child_col}: mapped to multiple {parent_col} values."
            )

            final_cols = list(subset.columns)
            ws.append(final_cols)
            style_header_row(ws, 1, len(final_cols))
            ws.freeze_panes = "A2"

            col_idx_map = {col: i for i, col in enumerate(final_cols, start=1)}

            for r_idx, (orig_idx, row) in enumerate(subset.iterrows(), start=2):
                for c_idx, col in enumerate(final_cols, start=1):
                    cell           = ws.cell(row=r_idx, column=c_idx, value=row[col])
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.fill      = ROW_ERROR_FILL

                # Highlight both parent and child columns red
                for highlight_col in (parent_col, child_col):
                    if highlight_col in col_idx_map:
                        target_cell      = ws.cell(row=r_idx, column=col_idx_map[highlight_col])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            note_row = len(row_indices) + 3
            ws.cell(
                row=note_row, column=1,
                value=(
                    f"Total error rows for '{child_col} → {parent_col}' mapping violation: "
                    f"{len(row_indices)}"
                ),
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

            auto_width(ws, min_w=10, max_w=60)


# ─────────────────────────────────────────────
#  PIPELINE CONTROLLER
# ─────────────────────────────────────────────

class ValidationPipeline:

    def __init__(self, input_path: str, output_path: str):
        self.input_path  = input_path
        self.output_path = output_path
        self.validator   = ProductHierarchyValidator()

    def run(self):
        print(f"📂  Reading: {self.input_path}")
        df = self._read_input()

        print(f"🔍  Validating {len(df)} rows …")
        df_validated = self.validator.validate_dataframe(df)

        error_count    = (df_validated["ERROR_COLUMNS"] != "").sum()
        pc_error_count = df_validated["_PC_ERROR_PAIRS"].apply(bool).sum()
        print(f"⚠️   Field errors found in {error_count} row(s)")
        print(f"⚠️   Parent-child mapping errors found in {pc_error_count} row(s)")

        print("📝  Building Excel report …")
        builder = ExcelReportBuilder(df_validated, self.output_path)
        builder.build()

    def _read_input(self) -> pd.DataFrame:
        ext = os.path.splitext(self.input_path)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return pd.read_excel(self.input_path, dtype=str)
        elif ext in (".csv", ".tab"):
            return pd.read_csv(
                self.input_path,
                dtype=str,
                sep="\t",
                encoding="latin-1",
            )
        else:
            raise ValueError(f"Unsupported file type: {ext}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    pipeline = ValidationPipeline(INPUT_FILE, OUTPUT_FILE)
    pipeline.run()
