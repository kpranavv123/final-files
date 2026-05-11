import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
SITE_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-11-1205.tab"
OUTPUT_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_Business_Rules.xlsx"


# ─────────────────────────────────────────────
#  BUSINESS RULE CONSTANTS
# ─────────────────────────────────────────────
VALID_NODE_TYPES = {"PL", "TBP", "3P", "CFA", "2P", "VENDING", "3P HYBRID"}

# Business rule identifiers
RULE1_KEY = "NODETYPESUPPLYCHAINNETWORK"
RULE2_KEY = "REGION_DESCRIPTION"

RULES_FIELDS_ORDERED  = [RULE1_KEY, RULE2_KEY]
ERROR_SHEET_PRIORITY  = [RULE1_KEY, RULE2_KEY]


# ─────────────────────────────────────────────
#  STYLING CONSTANTS  (identical to technical scripts)
# ─────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FF0000")
ROW_FILL   = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL   = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL  = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
STATS_FILL = PatternFill("solid", fgColor="EDEDED")

HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  REASON TEXT
# ─────────────────────────────────────────────
REASON_MAP = {
    RULE1_KEY: (
        "NODETYPESUPPLYCHAINNETWORK: Invalid or blank node type — "
        "must be one of PL / TBP / 3P / CFA / 2P / Vending / 3P Hybrid"
    ),
    RULE2_KEY: "REGION_DESCRIPTION: Field is blank — region description is mandatory",
}

RULES_CONTENT = {
    RULE1_KEY: [
        "Field must not be blank.",
        "Value must be one of the valid node types: PL, TBP, 3P, CFA, 2P, Vending, 3P Hybrid.",
        "Comparison is case-insensitive.",
        "Any other value (including free-text or typos) is treated as an error.",
    ],
    RULE2_KEY: [
        "Field must not be blank.",
        "A missing or empty REGION_DESCRIPTION is flagged as an error.",
    ],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class SiteBusinessRuleEngine:
    """
    Returns error_map:
        { row_index : { rule_key : { "reason": str, "highlight_cols": [col, ...] } } }
    """

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_node_type(self, row) -> tuple[bool, str]:
        val = row.get(RULE1_KEY)
        if self._is_blank(val):
            return False, REASON_MAP[RULE1_KEY]
        if str(val).strip().upper() not in VALID_NODE_TYPES:
            return False, REASON_MAP[RULE1_KEY]
        return True, ""

    def validate_region_description(self, row) -> tuple[bool, str]:
        if self._is_blank(row.get(RULE2_KEY)):
            return False, REASON_MAP[RULE2_KEY]
        return True, ""

    def get_rules(self) -> dict:
        return {
            RULE1_KEY: self.validate_node_type,
            RULE2_KEY: self.validate_region_description,
        }

    def run(self, df: pd.DataFrame) -> dict:
        rules     = self.get_rules()
        error_map = {}

        for idx, row in df.iterrows():
            for rule_key, rule_fn in rules.items():
                if rule_key not in df.columns:
                    continue
                try:
                    passed, reason = rule_fn(row)
                except Exception as e:
                    passed, reason = False, f"{rule_key}: Exception — {e}"

                if not passed:
                    error_map.setdefault(idx, {})[rule_key] = {
                        "reason":         reason,
                        "highlight_cols": [rule_key],
                    }

        return error_map


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class SiteBusinessValidator:

    def __init__(self, filepath: str):
        self.filepath  = filepath
        self.df        = pd.DataFrame()
        self.error_map = {}          # { row_idx : { rule_key : { reason, highlight_cols } } }

    def load(self):
        path = self.filepath.lower()
        if path.endswith(".csv"):
            self.df = pd.read_csv(self.filepath, dtype=str)
        elif path.endswith(".tab") or path.endswith(".tsv"):
            self.df = pd.read_csv(
                self.filepath,
                sep="\t",
                dtype=str,
                encoding="cp1252",
                engine="python",
            )
        elif path.endswith(".xlsx") or path.endswith(".xls"):
            self.df = pd.read_excel(self.filepath, dtype=str, engine="openpyxl")
        else:
            raise ValueError(f"Unsupported file format: {self.filepath}")

        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        engine         = SiteBusinessRuleEngine()
        self.error_map = engine.run(self.df)


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class SiteBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    def __init__(self, validator: SiteBusinessValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── Helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _auto_width(self, ws, min_w=10, max_w=60):
        for col in ws.columns:
            length = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(length + 3, min_w), max_w
            )

    # ── Summary Sheet ─────────────────────────
    def _write_summary_sheet(self, wb):
        ws         = wb.create_sheet(self.SHEET_SUMMARY)
        v          = self.validator
        total_rows = len(v.df)

        # Per-rule error counts
        col_error_counts = {r: 0 for r in RULES_FIELDS_ORDERED}
        for rule_dict in v.error_map.values():
            for rule_key in rule_dict:
                if rule_key in col_error_counts:
                    col_error_counts[rule_key] += 1

        # ── Title row ──
        ws.merge_cells("A1:G1")
        tc           = ws.cell(row=1, column=1,
                               value="Site Master – Business Rules Validation Summary")
        tc.font      = Font(name="Arial", bold=True, size=14)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Header row ──
        headers = ["#", "Rule / Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        # ── Data rows ──
        row_num = 3
        for rule_num, rule_key in enumerate(RULES_FIELDS_ORDERED, start=1):
            count      = col_error_counts.get(rule_key, 0)
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason     = REASON_MAP.get(rule_key, "") if count > 0 else ""

            ws.cell(row=row_num, column=1, value=rule_num)
            ws.cell(row=row_num, column=2, value=rule_key)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason)

            for c in range(1, 8):
                cell           = ws.cell(row=row_num, column=c)
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.fill      = WHITE_FILL
                cell.alignment = Alignment(
                    horizontal="left" if c == 7 else "center",
                    vertical="center",
                    wrap_text=(c == 7),
                )
            row_num += 1

        # ── TOTAL row ──
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(RULES_FIELDS_ORDERED)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c_idx, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""],
            start=1,
        ):
            cell           = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # ── Stats block ──
        records_with_errors = len(v.error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=2)
            lc           = ws.cell(row=row_num, column=1, value=label)
            lc.font      = Font(name="Arial", bold=True, size=10)
            lc.fill      = STATS_FILL
            lc.border    = THIN_BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")

            vc           = ws.cell(row=row_num, column=3, value=value)
            vc.font      = BODY_FONT
            vc.border    = THIN_BORDER
            vc.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        # ── Column widths ──
        col_widths = [6, 38, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Rules Sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        tc           = ws.cell(row=1, column=1,
                               value="Site Master – Business Validation Rules")
        tc.font      = Font(name="Arial", bold=True, size=13)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Rule / Field Name", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        current_row = 4
        for rule_num, (rule_key, rules_list) in enumerate(RULES_CONTENT.items(), start=1):
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                nc           = ws.cell(row=current_row, column=1,
                                       value=rule_num if r_idx == 0 else "")
                nc.fill      = RULE_FILL
                nc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                nc.border    = THIN_BORDER
                nc.alignment = Alignment(horizontal="center", vertical="center")

                fc           = ws.cell(row=current_row, column=2,
                                       value=rule_key if r_idx == 0 else "")
                fc.fill      = RULE_FILL
                fc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                fc.border    = THIN_BORDER
                fc.alignment = Alignment(horizontal="center", vertical="center",
                                         wrap_text=True)

                dc           = ws.cell(row=current_row, column=3, value=rule_text)
                dc.font      = BODY_FONT
                dc.border    = THIN_BORDER
                dc.alignment = Alignment(wrap_text=True, vertical="center",
                                         horizontal="left")
                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 70

    # ── Error Sheets ──────────────────────────
    def _write_error_sheets(self, wb):
        """
        One sheet per business rule.
        ALL columns from the input Site .tab file are shown.
        The failing column cell is highlighted red with white bold font.
        """
        v              = self.validator
        df             = v.df
        all_input_cols = list(df.columns)

        for rule_key in ERROR_SHEET_PRIORITY:
            row_indices = [
                idx for idx, rule_dict in v.error_map.items()
                if rule_key in rule_dict
            ]
            if not row_indices:
                continue

            subset = df.loc[row_indices, all_input_cols].copy()

            # Append reason column
            subset["ERROR_REASON"] = subset.index.map(
                lambda i: v.error_map.get(i, {}).get(rule_key, {}).get("reason", "")
            )

            # Safe sheet name (Excel max 31 chars)
            sheet_name = rule_key[:31]
            existing   = [s.title for s in wb.worksheets]
            counter    = 1
            base       = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base[:28]}_{counter}"
                counter   += 1

            ws = wb.create_sheet(sheet_name)
            self._write_header(ws, list(subset.columns))

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for r_idx, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for col, value in zip(subset.columns, row_data):
                    c_idx          = col_idx_map[col]
                    cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    cell.fill = ROW_FILL

                # Red highlight on the failing column
                if rule_key in col_idx_map:
                    tc      = ws.cell(row=r_idx, column=col_idx_map[rule_key])
                    tc.fill = RED_FILL
                    tc.font = ERR_FONT

            self._auto_width(ws, min_w=10, max_w=60)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{rule_key}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Orchestrate ───────────────────────────
    def write(self):
        v  = self.validator
        wb = Workbook()

        # Remove default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        self._write_summary_sheet(wb)
        self._write_rules_sheet(wb)
        self._write_error_sheets(wb)

        wb.save(self.output_path)

        rule1_errors = sum(1 for rd in v.error_map.values() if RULE1_KEY in rd)
        rule2_errors = sum(1 for rd in v.error_map.values() if RULE2_KEY in rd)

        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows                        : {len(v.df)}")
        print(f"   Rows with any error               : {len(v.error_map)}")
        print(f"   NODETYPESUPPLYCHAINNETWORK errors : {rule1_errors}")
        print(f"   REGION_DESCRIPTION errors         : {rule2_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class SiteBusinessProcessor:

    def __init__(self, input_path: str, output_path: str):
        self.validator = SiteBusinessValidator(input_path)
        self.writer    = SiteBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Columns detected : {list(self.validator.df.columns)}")
        print("🔍  Running business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = SiteBusinessProcessor(SITE_INPUT_FILE, OUTPUT_FILE)
    processor.run()
