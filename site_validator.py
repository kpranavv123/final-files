import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
SITE_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-11-1205.tab"
PART_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Part_Site_2026-05-14-1817.tab"
OUTPUT_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Validated_Site_Technical.xlsx"


# ─────────────────────────────────────────────
#  CONSOLIDATED PL LIST
# ─────────────────────────────────────────────
VALID_PLANTS = {
    "1555","1452","1204","1292","1253","1234","1731","1508","5011637","1100",
    "2091","2081","1436","1448","1438","1501","1248","1257","1623","1601",
    "1451","2092","1125","1649","1646","2088","1494","1137","1275","2084",
    "1437","5018849","1554","1213","1499","1575","1104","1104A","1145","1430",
    "1249","1509","1774","1295","1127","1146","1180","1432","1500","1742",
    "1642","1214","1426","1241","1463","1105","1423","1648","1733","1780",
    "1428","1495","1462","1738","5123531","1233","1158","1157","1482","1483",
    "1106","1485","1107","1225","5123742","1481","1203","2082","1416","2087",
    "1569","1647","1440","1645","1724","1503","1651","1421","1205","1563",
    "1640","1109","1471","1155","1110","1948","1473","1477","2089","1449",
    "1211","1788","1726","1578","1135","1491","1226","1650","1505","1433",
    "1739","2085","1659","1256","1653","1652","1111","1757","1559","1558",
    "1489","1657","4007430","1654","1112","1784","1579","1571","1754","2090",
    "1522","1771","1113","1229","1235","1156","1258","1176","1197","1265",
    "1445","1475","1740","2083","1725","1484","1114","1480","1506","5011407",
    "1520","5015073","1627","1758","1240","1478","1429","1656","1476","1298",
    "1801","1802","1661","1785","1507","5011308","1643","1638","1521","1118",
    "5013796","1511","1487","1488","1512","1281","1166","1184","1186","1218",
    "1223","1279","1442","1732","1472","2086","1439","1425","1208","1722",
    "1734","4011702","1525","1236","1568","1296","1589","1658","5123296",
    "1518","1455","1196","1441","1190","1593",
}

# ─────────────────────────────────────────────
#  FIELD CONFIG
# ─────────────────────────────────────────────
KEEP_COLS   = ["PLANT", "NAME", "POSTALCODE", "TCPL_PLANTTYPE", "REGIONID", "REGION_DESCRIPTION"]
FIELD_ORDER = ["PLANT", "NAME", "POSTALCODE", "TCPL_PLANTTYPE", "REGIONID", "REGION_DESCRIPTION"]

# ─────────────────────────────────────────────
#  Colours / Styles
# ─────────────────────────────────────────────
RED_FILL       = PatternFill("solid", fgColor="FF0000")
ROW_FILL       = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL       = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL      = PatternFill("solid", fgColor="E2EFDA")
TITLE_FILL     = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL     = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
STATS_FILL     = PatternFill("solid", fgColor="EDEDED")
PLANT_SUB_FILL = PatternFill("solid", fgColor="FFFFFF")
SUMM_HDR_FILL  = PatternFill("solid", fgColor="BDD7EE")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  Per-field reason shown in Summary sheet
# ─────────────────────────────────────────────
FIELD_REASON = {
    "NAME":               "NAME: Field is blank — site name is mandatory",
    "POSTALCODE":         "POSTALCODE: Field is blank — postal code is mandatory",
    "TCPL_PLANTTYPE":     "TCPL_PLANTTYPE: Field is blank",
    "REGIONID":           "REGIONID: Field is blank — region ID is mandatory",
    "REGION_DESCRIPTION": "REGION_DESCRIPTION: Field is blank — region description is mandatory",
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class SiteRuleEngine:

    def __init__(self, valid_plants: set, part_plants: set):
        self.valid_plants = {str(p).strip() for p in valid_plants}
        self.part_plants  = {str(p).strip() for p in part_plants}

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_plant(self, row) -> str:
        val = str(row.get("PLANT", "")).strip()
        if not val or val == "nan":
            return "PLANT: Field is blank — plant code is mandatory"
        if val not in self.valid_plants:
            return f"PLANT: '{val}' is not present in the Consolidated PL list"
        if val not in self.part_plants:
            return f"PLANT: '{val}' has no active Material-Plant combination in the Part master table"
        return ""

    def validate_name(self, row) -> str:
        if self._is_blank(row.get("NAME")):
            return "NAME: Field is blank — site name is mandatory"
        return ""

    def validate_postalcode(self, row) -> str:
        if self._is_blank(row.get("POSTALCODE")):
            return "POSTALCODE: Field is blank — postal code is mandatory"
        return ""

    def validate_tcpl_planttype(self, row) -> str:
        if self._is_blank(row.get("TCPL_PLANTTYPE")):
            return "TCPL_PLANTTYPE: Field is blank"
        return ""

    def validate_regionid(self, row) -> str:
        if self._is_blank(row.get("REGIONID")):
            return "REGIONID: Field is blank — region ID is mandatory"
        return ""

    def validate_region_description(self, row) -> str:
        if self._is_blank(row.get("REGION_DESCRIPTION")):
            return "REGION_DESCRIPTION: Field is blank — region description is mandatory"
        return ""

    def get_rules(self) -> dict:
        return {
            "PLANT":              self.validate_plant,
            "NAME":               self.validate_name,
            "POSTALCODE":         self.validate_postalcode,
            "TCPL_PLANTTYPE":     self.validate_tcpl_planttype,
            "REGIONID":           self.validate_regionid,
            "REGION_DESCRIPTION": self.validate_region_description,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class SiteTableValidator:

    def __init__(self, site_path: str, part_path: str, valid_plants: set):
        self.site_path    = site_path
        self.part_path    = part_path
        self.valid_plants = valid_plants
        self.df           = pd.DataFrame()
        self.part_plants  = set()
        self.error_map    = {}
        self.reason_map   = {}

    def load(self):
        self.df = pd.read_csv(self.site_path, sep="\t", dtype=str, encoding="latin1")
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        part_df = pd.read_csv(self.part_path, sep="\t", dtype=str, encoding="latin1")
        part_df.columns = [c.strip().upper() for c in part_df.columns]

        if "PLANT" not in part_df.columns:
            raise ValueError("PLANT column not found in Part table.")

        # Only consider FERT and HAWA rows from part master
        if "PRODUCTTYPE" in part_df.columns:
            part_df = part_df[
                part_df["PRODUCTTYPE"].str.strip().str.upper().isin({"FERT", "HAWA"})
            ]

        self.part_plants = set(part_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Part table plants loaded  : {len(self.part_plants)} unique values")

    def validate(self):
        engine = SiteRuleEngine(self.valid_plants, self.part_plants)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols    = []
            col_reason_map = {}

            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)
                except Exception:
                    reason = f"{col}: Unexpected validation error"

                if reason:
                    failed_cols.append(col)
                    col_reason_map[col] = reason

            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = col_reason_map

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors

    def get_plant_error_subcounts(self) -> dict:
        counts = {"blank": 0, "not_in_pl": 0, "no_part_site": 0}
        for idx, col_reason in self.reason_map.items():
            reason = col_reason.get("PLANT", "")
            if not reason:
                continue
            if "blank" in reason.lower():
                counts["blank"] += 1
            elif "consolidated pl list" in reason.lower():
                counts["not_in_pl"] += 1
            elif "material-plant combination" in reason.lower():
                counts["no_part_site"] += 1
        return counts


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class SiteReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Consolidated PL list.",
            "Must have an active Material-Plant combination in the Part master table (PLANT column).",
        ],
        "NAME":               ["Must not be blank."],
        "POSTALCODE":         ["Must not be blank."],
        "TCPL_PLANTTYPE":     ["Must not be blank."],
        "REGIONID":           ["Must not be blank."],
        "REGION_DESCRIPTION": ["Must not be blank."],
    }

    def __init__(self, validator: SiteTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── Helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    # ── Summary sheet ────────────────────────
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):

        # Title
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1, value="Site Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_HDR_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Column headers
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        # Per-field error counts
        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        plant_subcounts = self.validator.get_plant_error_subcounts()

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason_text = "" if col_name == "PLANT" else (
                FIELD_REASON.get(col_name, "") if has_errors else ""
            )

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_text)

            self._style_summary_data_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

            row_num  += 1

            # PLANT sub-rows
            if col_name == "PLANT" and has_errors:
                sub_definitions = [
                    (
                        "  ↳ Not in Consolidated PL List",
                        plant_subcounts["blank"] + plant_subcounts["not_in_pl"],
                        "PLANT: is not present in the Consolidated PL list",
                    ),
                    (
                        "  ↳ No Material-Plant Combination",
                        plant_subcounts["no_part_site"],
                        "PLANT: has no active Material-Plant combination in the Part master table",
                    ),
                ]
                for sub_label, sub_count, sub_reason in sub_definitions:
                    sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
                    sub_pct_health = round(100 - sub_pct_err, 2)

                    ws.cell(row=row_num, column=1, value="")
                    ws.cell(row=row_num, column=2, value=sub_label)
                    ws.cell(row=row_num, column=3, value=sub_count)
                    ws.cell(row=row_num, column=4, value=total_rows)
                    ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
                    ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
                    ws.cell(row=row_num, column=7, value=sub_reason)

                    self._style_summary_data_row(ws, row_num, fill=PLANT_SUB_FILL, italic=True)
                    ws.cell(row=row_num, column=2).alignment = Alignment(
                        horizontal="left", vertical="center", indent=1
                    )
                    ws.cell(row=row_num, column=7).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                    row_num += 1

            field_num += 1

        # TOTAL row
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c_idx, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""],
            start=1,
        ):
            cell           = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # Stats block
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        col_widths = [6, 30, 14, 16, 12, 12, 65]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Rules sheet ──────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1, value="Site Table – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 70

    # ── Field error sheets ───────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            subset    = df.loc[row_indices].copy()
            keep_here = [c for c in KEEP_COLS if c in subset.columns] + ["ERROR_COLUMNS"]
            subset    = subset[[c for c in keep_here if c in subset.columns]]

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i, fn=field_name: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = ROW_FILL
                    cell.border    = THIN_BORDER

                if field_name in col_idx_map:
                    target_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        keep_cols = [c for c in KEEP_COLS if c in df.columns] + ["ERROR_COLUMNS"]
        df        = df[keep_cols]

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(v.df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(v.df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class SiteTableProcessor:

    def __init__(self, site_path: str, part_path: str, output_path: str, valid_plants: set):
        self.validator = SiteTableValidator(site_path, part_path, valid_plants)
        self.writer    = SiteReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    Site columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = SiteTableProcessor(
        site_path    = SITE_INPUT_FILE,
        part_path    = PART_INPUT_FILE,
        output_path  = OUTPUT_FILE,
        valid_plants = VALID_PLANTS,
    )
    processor.run()
