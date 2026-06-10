import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
STO_INPUT_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt\STO\ScheduledReceipt_STO.tab"
SITE_INPUT_FILE    = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt\STO\Site_2026-04-09-1058.xlsx"
PART_FG_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt\STO\Part_FG.xlsx"
OUTPUT_FILE        = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt\STO\Validated_ScheduledReceipt_STO.xlsx"


# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
YYYYMMDD_PATTERN = re.compile(r'^\d{8}$')

FIELD_ORDER = [
    "PURCHASEORDER",
    "DESTINATIONPLANT",
    "PURCHASINGDOCUMENTTYPE",
    "PURCHASEORDERITEM",
    "VENDORSACCOUNTNUMBER",
    "SOURCEPLANT",
    "MATERIALNUMBER",
    "ACTUALDELIVEREDQUANTITYINBU",
    "PURCHASEORDERUNITOFMEASURE",
    "SCHEDULELINEDELIVERYDATE",
    "NETPRICEINPURCHASINGDOCUMENTI",
    "MATSTAGINGAVAILABILITYDATE",
    "TRANSITTIME",
]

KEEP_COLS = FIELD_ORDER   # same set, keep in same order

FIELDS_WITH_SUB_ROWS = {
    "DESTINATIONPLANT",
    "MATERIALNUMBER",
    "SCHEDULELINEDELIVERYDATE",
    "TRANSITTIME",
}

FIELD_REASON = {
    "PURCHASEORDER":                 "PURCHASEORDER: Field is blank",
    "DESTINATIONPLANT":              "",
    "PURCHASINGDOCUMENTTYPE":        "PURCHASINGDOCUMENTTYPE: Field is blank",
    "PURCHASEORDERITEM":             "PURCHASEORDERITEM: Field is blank",
    "VENDORSACCOUNTNUMBER":          "VENDORSACCOUNTNUMBER: Field is blank",
    "SOURCEPLANT":                   "SOURCEPLANT: Field is blank",
    "MATERIALNUMBER":                "",
    "ACTUALDELIVEREDQUANTITYINBU":   "ACTUALDELIVEREDQUANTITYINBU: Field is blank",
    "PURCHASEORDERUNITOFMEASURE":    "PURCHASEORDERUNITOFMEASURE: Field is blank",
    "SCHEDULELINEDELIVERYDATE":      "",
    "NETPRICEINPURCHASINGDOCUMENTI": "NETPRICEINPURCHASINGDOCUMENTI: Field is blank",
    "MATSTAGINGAVAILABILITYDATE":    "MATSTAGINGAVAILABILITYDATE: Field is blank",
    "TRANSITTIME":                   "",
}

RULES_CONTENT = {
    "PURCHASEORDER":                 ["Must not be blank."],
    "DESTINATIONPLANT":              ["Must not be blank.",
                                      "Must be present in the Site master (PLANT column)."],
    "PURCHASINGDOCUMENTTYPE":        ["Must not be blank."],
    "PURCHASEORDERITEM":             ["Must not be blank."],
    "VENDORSACCOUNTNUMBER":          ["Must not be blank."],
    "SOURCEPLANT":                   ["Must not be blank."],
    "MATERIALNUMBER":                ["Must not be blank.",
                                      "Must be present in the Part (FG) master."],
    "ACTUALDELIVEREDQUANTITYINBU":   ["Must not be blank."],
    "PURCHASEORDERUNITOFMEASURE":    ["Must not be blank."],
    "SCHEDULELINEDELIVERYDATE":      ["Must not be blank.",
                                      "Must follow the date format: YYYYMMDD (8 numeric digits)."],
    "NETPRICEINPURCHASINGDOCUMENTI": ["Must not be blank."],
    "MATSTAGINGAVAILABILITYDATE":    ["Must not be blank."],
    "TRANSITTIME":                   ["Must not be blank.",
                                      "Must follow the date format: YYYYMMDD (8 numeric digits)."],
}


# ─────────────────────────────────────────────
#  Colours / Styles
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUB_FILL        = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class STORuleEngine:

    def __init__(self, site_plants: set, part_fg_materials: set):
        self.site_plants       = set(str(p).strip() for p in site_plants)
        self.part_fg_materials = set(str(m).strip().upper() for m in part_fg_materials)

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _blank(self, row, field: str) -> str:
        return f"{field}: Field is blank" if self._is_blank(row.get(field)) else ""

    def validate_purchaseorder(self, row) -> str:
        return self._blank(row, "PURCHASEORDER")

    def validate_destinationplant(self, row) -> str:
        val = str(row.get("DESTINATIONPLANT", "")).strip()
        if not val or val == "nan":
            return "DESTINATIONPLANT: Field is blank"
        if val not in self.site_plants:
            return f"DESTINATIONPLANT: '{val}' is not present in the Site master"
        return ""

    def validate_purchasingdocumenttype(self, row) -> str:
        return self._blank(row, "PURCHASINGDOCUMENTTYPE")

    def validate_purchaseorderitem(self, row) -> str:
        return self._blank(row, "PURCHASEORDERITEM")

    def validate_vendorsaccountnumber(self, row) -> str:
        return self._blank(row, "VENDORSACCOUNTNUMBER")

    def validate_sourceplant(self, row) -> str:
        return self._blank(row, "SOURCEPLANT")

    def validate_materialnumber(self, row) -> str:
        val = row.get("MATERIALNUMBER", "")
        if self._is_blank(val):
            return "MATERIALNUMBER: Field is blank"
        if str(val).strip().upper() not in self.part_fg_materials:
            return f"MATERIALNUMBER: '{str(val).strip()}' is not present in the Part (FG) master"
        return ""

    def validate_actualdeliveredquantityinbu(self, row) -> str:
        return self._blank(row, "ACTUALDELIVEREDQUANTITYINBU")

    def validate_purchaseorderunitofmeasure(self, row) -> str:
        return self._blank(row, "PURCHASEORDERUNITOFMEASURE")

    def validate_schedulelinedeliverydate(self, row) -> str:
        val = row.get("SCHEDULELINEDELIVERYDATE", "")
        if self._is_blank(val):
            return "SCHEDULELINEDELIVERYDATE: Field is blank"
        if not YYYYMMDD_PATTERN.match(str(val).strip()):
            return f"SCHEDULELINEDELIVERYDATE: '{str(val).strip()}' does not follow the required format YYYYMMDD"
        return ""

    def validate_netpriceinpurchasingdocumenti(self, row) -> str:
        return self._blank(row, "NETPRICEINPURCHASINGDOCUMENTI")

    def validate_matstagingavailabilitydate(self, row) -> str:
        return self._blank(row, "MATSTAGINGAVAILABILITYDATE")

    def validate_transittime(self, row) -> str:
        val = row.get("TRANSITTIME", "")
        if self._is_blank(val):
            return "TRANSITTIME: Field is blank"
        if not YYYYMMDD_PATTERN.match(str(val).strip()):
            return f"TRANSITTIME: '{str(val).strip()}' does not follow the required format YYYYMMDD"
        return ""

    def get_rules(self) -> dict:
        return {
            "PURCHASEORDER":                 self.validate_purchaseorder,
            "DESTINATIONPLANT":              self.validate_destinationplant,
            "PURCHASINGDOCUMENTTYPE":        self.validate_purchasingdocumenttype,
            "PURCHASEORDERITEM":             self.validate_purchaseorderitem,
            "VENDORSACCOUNTNUMBER":          self.validate_vendorsaccountnumber,
            "SOURCEPLANT":                   self.validate_sourceplant,
            "MATERIALNUMBER":                self.validate_materialnumber,
            "ACTUALDELIVEREDQUANTITYINBU":   self.validate_actualdeliveredquantityinbu,
            "PURCHASEORDERUNITOFMEASURE":    self.validate_purchaseorderunitofmeasure,
            "SCHEDULELINEDELIVERYDATE":      self.validate_schedulelinedeliverydate,
            "NETPRICEINPURCHASINGDOCUMENTI": self.validate_netpriceinpurchasingdocumenti,
            "MATSTAGINGAVAILABILITYDATE":    self.validate_matstagingavailabilitydate,
            "TRANSITTIME":                   self.validate_transittime,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class STOValidator:

    def __init__(self, input_path: str, site_path: str, part_fg_path: str):
        self.input_path   = input_path
        self.site_path    = site_path
        self.part_fg_path = part_fg_path
        self.df           = pd.DataFrame()
        self.site_plants  = set()
        self.part_fg_mats = set()
        self.error_map    = {}
        self.reason_map   = {}

    def load(self):
        self.df = pd.read_csv(self.input_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        site_df = pd.read_excel(self.site_path, dtype=str, engine="openpyxl")
        site_df.columns = [c.strip().upper() for c in site_df.columns]
        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site master.")
        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site master plants loaded      : {len(self.site_plants)} unique values")

        part_fg_df = pd.read_excel(self.part_fg_path, dtype=str, engine="openpyxl")
        part_fg_df.columns = [c.strip().upper() for c in part_fg_df.columns]
        mat_col = next((c for c in part_fg_df.columns if "MATERIAL" in c), None)
        if mat_col is None:
            raise ValueError("No material number column found in Part (FG) master.")
        self.part_fg_mats = set(
            part_fg_df[mat_col].dropna().str.strip().str.upper().tolist()
        )
        print(f"    Part (FG) materials loaded     : {len(self.part_fg_mats)} unique values")

    def validate(self):
        engine = STORuleEngine(self.site_plants, self.part_fg_mats)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols    = []
            col_reason_map = {}
            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)
                except Exception:
                    reason = f"{col}: Unexpected validation error"
                if reason:
                    failed_cols.append(col)
                    col_reason_map[col] = reason
            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = col_reason_map

    def get_error_series(self) -> pd.Series:
        return pd.Series(
            {idx: " | ".join(cr.values()) for idx, cr in self.reason_map.items()},
            dtype=str,
        )

    def get_field_error_series(self, field_name: str) -> pd.Series:
        return pd.Series(
            {idx: cr[field_name] for idx, cr in self.reason_map.items() if field_name in cr},
            dtype=str,
        )

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors

    def _subcounts(self, field: str, blank_key: str, other_key: str) -> dict:
        counts = {blank_key: 0, other_key: 0}
        for cr in self.reason_map.values():
            reason = cr.get(field, "")
            if not reason:
                continue
            (counts[blank_key] if "blank" in reason.lower() else counts[other_key]).__class__  # noqa
            if "blank" in reason.lower():
                counts[blank_key] += 1
            else:
                counts[other_key] += 1
        return counts

    def get_destinationplant_error_subcounts(self) -> dict:
        return self._subcounts("DESTINATIONPLANT", "blank", "not_in_site")

    def get_materialnumber_error_subcounts(self) -> dict:
        return self._subcounts("MATERIALNUMBER", "blank", "not_in_part_fg")

    def get_schedulelinedeliverydate_error_subcounts(self) -> dict:
        return self._subcounts("SCHEDULELINEDELIVERYDATE", "blank", "invalid_format")

    def get_transittime_error_subcounts(self) -> dict:
        return self._subcounts("TRANSITTIME", "blank", "invalid_format")


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class STOReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    def __init__(self, validator: STOValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = WHITE_FILL if col_name == "ERROR_COLUMNS" else HDR_FILL
            cell.font      = (Font(bold=True, name="Arial", color="000000")
                               if col_name == "ERROR_COLUMNS" else HDR_FONT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_row(self, ws, row_num: int, num_cols: int = 7,
                           bold: bool = False, fill: PatternFill = None,
                           italic: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    def _write_sub_rows(self, ws, row_num: int, total_rows: int,
                        sub_definitions: list) -> int:
        for sub_label, sub_count, sub_reason in sub_definitions:
            sub_pct_err    = round((sub_count / total_rows) * 100, 2) if total_rows else 0
            sub_pct_health = round(100 - sub_pct_err, 2)

            ws.cell(row=row_num, column=1, value="")
            ws.cell(row=row_num, column=2, value=sub_label)
            ws.cell(row=row_num, column=3, value=sub_count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{sub_pct_err}%")
            ws.cell(row=row_num, column=7, value=sub_reason)

            self._style_summary_row(ws, row_num, fill=SUB_FILL, italic=True)
            ws.cell(row=row_num, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            row_num += 1
        return row_num

    # ══════════════════════════════════════════
    #  Summary sheet
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):

        ws.merge_cells("A1:G1")
        tc           = ws.cell(row=1, column=1,
                               value="ScheduledReceipt (STO) Validation Summary")
        tc.font      = Font(name="Arial", bold=True, size=14)
        tc.fill      = SUMM_TITLE_FILL
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        for c_idx, h in enumerate(
            ["#", "Field Name", "Error Count", "Record Count",
             "% Health", "% of Error", "Reason / Sub-Category"], start=1
        ):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        v        = self.validator
        dest_sub = v.get_destinationplant_error_subcounts()
        mat_sub  = v.get_materialnumber_error_subcounts()
        sld_sub  = v.get_schedulelinedeliverydate_error_subcounts()
        tt_sub   = v.get_transittime_error_subcounts()

        row_num   = 3
        field_num = 1

        for col_name in FIELD_ORDER:
            count      = col_error_counts.get(col_name, 0)
            has_errors = count > 0
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)
            reason_txt = ("" if col_name in FIELDS_WITH_SUB_ROWS
                          else (FIELD_REASON.get(col_name, "") if has_errors else ""))

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=col_name)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=reason_txt)

            self._style_summary_row(ws, row_num, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            row_num += 1

            if col_name == "DESTINATIONPLANT" and has_errors:
                row_num = self._write_sub_rows(ws, row_num, total_rows, [
                    ("  ↳ Blank Destination Plant",
                     dest_sub["blank"],
                     "DESTINATIONPLANT: Field is blank"),
                    ("  ↳ Not in Site Master",
                     dest_sub["not_in_site"],
                     "DESTINATIONPLANT: Plant code not found in the Site master"),
                ])
            elif col_name == "MATERIALNUMBER" and has_errors:
                row_num = self._write_sub_rows(ws, row_num, total_rows, [
                    ("  ↳ Blank Material Number",
                     mat_sub["blank"],
                     "MATERIALNUMBER: Field is blank"),
                    ("  ↳ Not in Part (FG) Master",
                     mat_sub["not_in_part_fg"],
                     "MATERIALNUMBER: Material number not found in the Part (FG) master"),
                ])
            elif col_name == "SCHEDULELINEDELIVERYDATE" and has_errors:
                row_num = self._write_sub_rows(ws, row_num, total_rows, [
                    ("  ↳ Blank Schedule Line Delivery Date",
                     sld_sub["blank"],
                     "SCHEDULELINEDELIVERYDATE: Field is blank"),
                    ("  ↳ Invalid Date Format (not YYYYMMDD)",
                     sld_sub["invalid_format"],
                     "SCHEDULELINEDELIVERYDATE: Does not follow required format YYYYMMDD"),
                ])
            elif col_name == "TRANSITTIME" and has_errors:
                row_num = self._write_sub_rows(ws, row_num, total_rows, [
                    ("  ↳ Blank Transit Time",
                     tt_sub["blank"],
                     "TRANSITTIME: Field is blank"),
                    ("  ↳ Invalid Date Format (not YYYYMMDD)",
                     tt_sub["invalid_format"],
                     "TRANSITTIME: Does not follow required format YYYYMMDD"),
                ])

            field_num += 1

        # ── TOTAL row ──
        total_errors       = sum(col_error_counts.values())
        total_record_count = total_rows * len(FIELD_ORDER)
        total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health   = round(100 - total_pct_error, 2)

        for c, val in enumerate(
            ["", "TOTAL", total_errors, total_record_count,
             f"{total_pct_health}%", f"{total_pct_error}%", ""], start=1
        ):
            cell           = ws.cell(row=row_num, column=c, value=val)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2

        # ── Stats block ──
        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", len(error_map)),
            ("Records Passing:",     total_rows - len(error_map)),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            lc           = ws.cell(row=row_num, column=1, value=label)
            lc.font      = Font(name="Arial", bold=True, size=10)
            lc.fill      = STATS_FILL
            lc.border    = THIN_BORDER
            lc.alignment = Alignment(horizontal="left", vertical="center")
            vc           = ws.cell(row=row_num, column=3, value=value)
            vc.font      = Font(name="Arial", size=10)
            vc.border    = THIN_BORDER
            vc.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        for c_idx, width in enumerate([6, 40, 14, 16, 12, 12, 70], start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            ws         = wb.create_sheet(field_name[:31].replace("/", "-").replace("\\", "-").replace("*", ""))
            subset     = df.loc[field_errors[field_name]].copy()
            keep_here  = [c for c in KEEP_COLS if c in subset.columns] + ["ERROR_COLUMNS"]
            subset     = subset[keep_here]

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(lambda i: field_err_series.get(i, ""))

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER
                if field_name in col_idx_map:
                    tc      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    tc.fill = RED_FILL
                    tc.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"
            ws.cell(
                row=len(subset) + 3, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        tc           = ws.cell(row=1, column=1, value="ScheduledReceipt (STO) – Validation Rules")
        tc.font      = Font(name="Arial", bold=True, size=13)
        tc.fill      = TITLE_FILL
        tc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                ws.cell(row=current_row, column=1,
                        value=rule_num if r_idx == 0 else "").fill = RULE_FILL
                ws.cell(row=current_row, column=1).font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                ws.cell(row=current_row, column=1).border    = THIN_BORDER
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row=current_row, column=2,
                        value=field if r_idx == 0 else "").fill = RULE_FILL
                ws.cell(row=current_row, column=2).font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                ws.cell(row=current_row, column=2).border    = THIN_BORDER
                ws.cell(row=current_row, column=2).alignment = Alignment(vertical="center")

                dc           = ws.cell(row=current_row, column=3, value=rule_text)
                dc.font      = BODY_FONT
                dc.border    = THIN_BORDER
                dc.alignment = Alignment(wrap_text=True, vertical="center")
                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 75

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        keep_cols = [c for c in KEEP_COLS if c in df.columns] + ["ERROR_COLUMNS"]
        df        = df[keep_cols]

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class STOProcessor:

    def __init__(self, input_path: str, site_path: str,
                 part_fg_path: str, output_path: str):
        self.validator = STOValidator(input_path, site_path, part_fg_path)
        self.writer    = STOReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    ScheduledReceipt (STO) columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = STOProcessor(
        input_path   = STO_INPUT_FILE,
        site_path    = SITE_INPUT_FILE,
        part_fg_path = PART_FG_INPUT_FILE,
        output_path  = OUTPUT_FILE,
    )
    processor.run()
