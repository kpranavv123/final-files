import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
PO_FROM_VENDOR_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt(PO_from Vendor)\ScheduledReceipt.tab"
PARTSOURCE_BUY_INPUT_FILE = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource(Buy)\PartSource_Buy.tab"
OUTPUT_FILE               = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt(PO_from Vendor)\Validated_ScheduledReceipt_Business.xlsx"


# ─────────────────────────────────────────────
#  MASTER FILE COLUMN NAMES  (PartSource(Buy) extract, after .strip().upper())
# ─────────────────────────────────────────────
MASTER_MATERIAL_COL = "MATERIALNUMBER(PART)"
MASTER_PLANT_COL    = "PLANT(SITE)"
MASTER_POLICY_COL   = "ORDERPOLICY"
MASTER_POLICY_VALUE = "BUY"

# ─────────────────────────────────────────────
#  PO_FROM_VENDOR COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
PO_MATERIAL_COL = "MATERIALNUMBER"
PO_PLANT_COL    = "DESTINATIONPLANT"
PO_QTY_COL      = "POQUANTITYINBU"


# ─────────────────────────────────────────────
#  Colours / Styles  — matched to existing Business Validator template
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUBROW_FILL     = PatternFill("solid", start_color="F7F7F7", end_color="F7F7F7")

HDR_FONT     = Font(bold=True, name="Arial")
BODY_FONT    = Font(name="Arial", size=10)
ERR_FONT     = Font(name="Arial", size=10, bold=True, color="FFFFFF")
SUBROW_FONT  = Font(name="Arial", size=10, italic=True)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  FIELD ORDER  (only fields with business rules)
# ─────────────────────────────────────────────
FIELD_ORDER = ["MATERIALNUMBER", "POQUANTITYINBU"]

# Sub-category reasons shown as indented rows under MATERIALNUMBER in Summary
SUB_CATEGORIES = ["Blank", "Not in Master"]

# Columns (per field) to highlight red in the per-field error sheets
FIELD_HIGHLIGHT_COLS = {
    "MATERIALNUMBER":  [PO_MATERIAL_COL, PO_PLANT_COL],
    "POQUANTITYINBU":  [PO_QTY_COL],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class ScheduledReceiptBusinessRuleEngine:
    """
    Rules:
      MATERIALNUMBER
        - The MATERIALNUMBER / DESTINATIONPLANT (Part-Site) combination on
          this row must exist in the PartSource(Buy) extract as a
          MATERIALNUMBER(PART) / PLANT(SITE) combination where
          ORDERPOLICY = 'BUY'.

      POQUANTITYINBU
        - The value must not be negative.
    """

    def __init__(self, valid_buy_combos: set):
        self.valid_buy_combos = valid_buy_combos

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_row(self, row) -> dict:
        """Returns {field_name: reason} for each field that fails its rule(s)."""
        reasons = {}

        # ── MATERIALNUMBER (Part-Site / BUY combo) ──
        mat_raw   = row.get(PO_MATERIAL_COL, "")
        plant_raw = row.get(PO_PLANT_COL, "")
        mat_blank   = self._is_blank(mat_raw)
        plant_blank = self._is_blank(plant_raw)

        if mat_blank or plant_blank:
            missing = []
            if mat_blank:
                missing.append(PO_MATERIAL_COL)
            if plant_blank:
                missing.append(PO_PLANT_COL)
            reasons["MATERIALNUMBER"] = (
                f"MATERIALNUMBER: Blank field(s) - {', '.join(missing)}"
            )
        else:
            mat   = str(mat_raw).strip()
            plant = str(plant_raw).strip()
            key   = f"{mat}|{plant}"

            if key not in self.valid_buy_combos:
                reasons["MATERIALNUMBER"] = (
                    f"MATERIALNUMBER: Part-Site combination '{mat}-{plant}' not found in "
                    f"PartSource(Buy) extract with ORDERPOLICY='{MASTER_POLICY_VALUE}'"
                )

        # ── POQUANTITYINBU (must not be negative) ──
        qty_raw = row.get(PO_QTY_COL, "")
        if not self._is_blank(qty_raw):
            try:
                qty_val = float(str(qty_raw).strip())
            except ValueError:
                qty_val = None

            if qty_val is not None and qty_val < 0:
                reasons["POQUANTITYINBU"] = (
                    f"POQUANTITYINBU: Value is negative ({str(qty_raw).strip()})"
                )

        return reasons


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class ScheduledReceiptBusinessTableValidator:

    def __init__(self, po_path: str, partsource_buy_path: str):
        self.po_path             = po_path
        self.partsource_buy_path = partsource_buy_path
        self.df                  = pd.DataFrame()
        self.valid_buy_combos    = set()
        self.error_map           = {}   # row_idx -> [failed field names]
        self.reason_map          = {}   # row_idx -> {field: reason}

        # sub-category counters for Summary sheet breakdown
        self.blank_count         = 0
        self.not_in_master_count = 0
        self.negative_qty_count  = 0

    def load(self):
        self.df = pd.read_csv(self.po_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        master_df = pd.read_csv(self.partsource_buy_path, sep="\t", dtype=str)
        master_df.columns = [c.strip().upper() for c in master_df.columns]

        for _, r in master_df.iterrows():
            policy = str(r.get(MASTER_POLICY_COL, "")).strip().upper()
            if policy != MASTER_POLICY_VALUE:
                continue
            mat   = str(r.get(MASTER_MATERIAL_COL, "")).strip()
            plant = str(r.get(MASTER_PLANT_COL, "")).strip()
            if mat == "" or plant == "":
                continue
            self.valid_buy_combos.add(f"{mat}|{plant}")

    def validate(self):
        engine = ScheduledReceiptBusinessRuleEngine(self.valid_buy_combos)

        for idx, row in self.df.iterrows():
            try:
                reasons = engine.validate_row(row)
            except Exception:
                reasons = {}

            if reasons:
                self.error_map[idx]  = list(reasons.keys())
                self.reason_map[idx] = reasons

                mat_reason = reasons.get("MATERIALNUMBER", "")
                if mat_reason.startswith("MATERIALNUMBER: Blank field(s)"):
                    self.blank_count += 1
                elif "not found in" in mat_reason:
                    self.not_in_master_count += 1

                if "POQUANTITYINBU" in reasons:
                    self.negative_qty_count += 1

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class ScheduledReceiptBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIALNUMBER": [
            "The MATERIALNUMBER / DESTINATIONPLANT (Part-Site) combination must be "
            "present in the PartSource(Buy) extract as MATERIALNUMBER(PART) / "
            "PLANT(SITE), where ORDERPOLICY = 'BUY'.",
        ],
        "POQUANTITYINBU": [
            "The value must not be negative.",
        ],
    }

    def __init__(self, validator: ScheduledReceiptBusinessTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None,
                                italic: bool = False, indent_col2: bool = False):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, italic=italic, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill
        if indent_col2:
            ws.cell(row=row_num, column=2).alignment = Alignment(
                horizontal="left", vertical="center", indent=2
            )

    # ══════════════════════════════════════════
    #  Summary sheet
    #  Row 1: MATERIALNUMBER + Blank / Not-in-Master sub-rows
    #  Row 2: POQUANTITYINBU (single rule, no sub-rows)
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):
        v = self.validator

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="ScheduledReceipt (PO from Vendor) Business Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        row_num  = 3
        field_no = 1

        # ── Field-level error counts (per field, from error_map) ──
        field_errors = v.get_errors_by_field()

        # ── MATERIALNUMBER main row ──
        mat_errors   = len(field_errors.get("MATERIALNUMBER", []))
        pct_error    = round((mat_errors / total_rows) * 100, 2) if total_rows else 0
        pct_health   = round(100 - pct_error, 2)

        ws.cell(row=row_num, column=1, value=field_no)
        ws.cell(row=row_num, column=2, value="MATERIALNUMBER")
        ws.cell(row=row_num, column=3, value=mat_errors)
        ws.cell(row=row_num, column=4, value=total_rows)
        ws.cell(row=row_num, column=5, value=f"{pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{pct_error}%")
        ws.cell(row=row_num, column=7,
                value="Part-Site combination not found in PartSource(Buy) extract (ORDERPOLICY='BUY')")
        self._style_summary_data_row(ws, row_num, bold=True, fill=WHITE_FILL)
        ws.cell(row=row_num, column=7).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        row_num  += 1
        field_no += 1

        # ── Sub-rows: Blank / Not in Master ──
        sub_counts = {
            "Blank":          v.blank_count,
            "Not in Master":  v.not_in_master_count,
        }
        for sub_label in SUB_CATEGORIES:
            sub_count      = sub_counts.get(sub_label, 0)
            sub_pct_error  = round((sub_count / total_rows) * 100, 2) if total_rows else 0
            sub_pct_health = round(100 - sub_pct_error, 2)

            ws.cell(row=row_num, column=1, value="")
            ws.cell(row=row_num, column=2, value=f"└─ {sub_label}")
            ws.cell(row=row_num, column=3, value=sub_count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{sub_pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{sub_pct_error}%")
            ws.cell(row=row_num, column=7, value="")
            self._style_summary_data_row(ws, row_num, italic=True, fill=SUBROW_FILL,
                                          indent_col2=True)
            row_num += 1

        # ── POQUANTITYINBU main row (single rule, no sub-rows) ──
        qty_errors     = len(field_errors.get("POQUANTITYINBU", []))
        qty_pct_error  = round((qty_errors / total_rows) * 100, 2) if total_rows else 0
        qty_pct_health = round(100 - qty_pct_error, 2)

        ws.cell(row=row_num, column=1, value=field_no)
        ws.cell(row=row_num, column=2, value="POQUANTITYINBU")
        ws.cell(row=row_num, column=3, value=qty_errors)
        ws.cell(row=row_num, column=4, value=total_rows)
        ws.cell(row=row_num, column=5, value=f"{qty_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{qty_pct_error}%")
        ws.cell(row=row_num, column=7, value="Value is negative")
        self._style_summary_data_row(ws, row_num, bold=True, fill=WHITE_FILL)
        ws.cell(row=row_num, column=7).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        row_num  += 1
        field_no += 1

        # ── TOTAL row ──
        total_errors        = mat_errors + qty_errors
        total_record_count  = total_rows * len(FIELD_ORDER)
        total_pct_error     = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health    = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2   # blank spacer

        # ── Quick-glance stats block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 42, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets  (ALL source columns) ──
    def _write_field_error_sheets(self, wb, df: pd.DataFrame, all_cols: list):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            keep_here = all_cols + ["ERROR_COLUMNS"]
            subset    = df.loc[row_indices, keep_here].copy()

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER

                # Highlight only the column(s) relevant to this field's rule.
                for involved_col in FIELD_HIGHLIGHT_COLS.get(field_name, []):
                    if involved_col in col_idx_map:
                        target_cell      = ws.cell(row=excel_row, column=col_idx_map[involved_col])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="ScheduledReceipt (PO from Vendor) – Business Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 90

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        # All source columns, in original order — used for per-field error sheets
        all_cols = list(df.columns)

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df, all_cols)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows        : {len(df)}")
        print(f"   Error rows        : {len(v.error_map)}")
        print(f"     - Blank              : {v.blank_count}")
        print(f"     - Not in Master      : {v.not_in_master_count}")
        print(f"     - Negative Quantity  : {v.negative_qty_count}")
        print(f"   Field sheets       : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class ScheduledReceiptBusinessTableProcessor:

    def __init__(self, po_path: str, partsource_buy_path: str, output_path: str):
        self.validator = ScheduledReceiptBusinessTableValidator(po_path, partsource_buy_path)
        self.writer    = ScheduledReceiptBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    ScheduledReceipt columns detected : {list(self.validator.df.columns)}")
        print(f"    Valid BUY Part-Site combinations  : {len(self.validator.valid_buy_combos)}")
        print("🔍  Validating business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = ScheduledReceiptBusinessTableProcessor(
        po_path             = PO_FROM_VENDOR_INPUT_FILE,
        partsource_buy_path = PARTSOURCE_BUY_INPUT_FILE,
        output_path         = OUTPUT_FILE,
    )
    processor.run()
