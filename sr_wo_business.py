import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS
# ─────────────────────────────────────────────
WO_INPUT_FILE          = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt(WO)\ScheduledReceipt_WO.tab"
PART_SOURCE_MAKE_FILE  = r"C:\Users\SW526XH\Downloads\Go Live-2\PartSource(Make)\PartSource_Make.tab"
OUTPUT_FILE            = r"C:\Users\SW526XH\Downloads\Go Live-2\ScheduledReceipt(WO)\Validated_ScheduledReceipt_WO_Business.xlsx"


# ─────────────────────────────────────────────
#  SR(WO) EXTRACT COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
WO_PARTNAME_COL          = "PARTNAME"
WO_PARTSITE_COL          = "PARTSITE"
WO_PENDINGQUANTITY_COL   = "PENDINGQUANTITY"

# ─────────────────────────────────────────────
#  PartSource(Make) LOOKUP COLUMN NAMES  (after .strip().upper())
# ─────────────────────────────────────────────
PSM_MATERIAL_COL = "MATERIAL"
PSM_PLANT_COL    = "PLANT"


# ─────────────────────────────────────────────
#  Colours / Styles  — matched to existing Business Validator template
# ─────────────────────────────────────────────
RED_FILL        = PatternFill("solid", start_color="FF0000", end_color="FF0000")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL       = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL      = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
SUMM_HDR_FILL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
SUMM_TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")

HDR_FONT     = Font(bold=True, name="Arial")
BODY_FONT    = Font(name="Arial", size=10)
ERR_FONT     = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ─────────────────────────────────────────────
#  FIELD ORDER  (only fields with business rules)
# ─────────────────────────────────────────────
FIELD_ORDER = ["PARTNAME", "PENDINGQUANTITY"]

# Columns (per field) to highlight red in the per-field error sheets
FIELD_HIGHLIGHT_COLS = {
    "PARTNAME":        [WO_PARTNAME_COL, WO_PARTSITE_COL],
    "PENDINGQUANTITY": [WO_PENDINGQUANTITY_COL],
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class ScheduledReceiptWOBusinessRuleEngine:
    """
    Rules:
      PARTNAME (composite with PARTSITE)
        - The PARTNAME-PARTSITE combination must exist as a
          MATERIAL-PLANT combination in the PartSource(Make) extract.

      PENDINGQUANTITY
        - The value must not be negative.
    """

    def __init__(self, valid_combinations: set):
        # valid_combinations: set of (MATERIAL, PLANT) tuples from PartSource(Make)
        self.valid_combinations = valid_combinations

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def validate_row(self, row) -> dict:
        """Returns {field_name: reason} for any failed rules on this row."""
        reasons = {}

        # ── Rule 1: PARTNAME-PARTSITE must exist in PartSource(Make) ──
        part_name = row.get(WO_PARTNAME_COL, "")
        part_site = row.get(WO_PARTSITE_COL, "")

        if not self._is_blank(part_name) and not self._is_blank(part_site):
            key = (str(part_name).strip(), str(part_site).strip())
            if key not in self.valid_combinations:
                reasons["PARTNAME"] = (
                    f"PARTNAME-PARTSITE combination ({key[0]}-{key[1]}) not found "
                    f"as MATERIAL-PLANT combination in PartSource(Make)"
                )

        # ── Rule 2: PENDINGQUANTITY must not be negative ──
        qty_raw = row.get(WO_PENDINGQUANTITY_COL, "")
        if not self._is_blank(qty_raw):
            try:
                qty_val = float(str(qty_raw).strip())
            except ValueError:
                qty_val = None

            if qty_val is not None and qty_val < 0:
                reasons["PENDINGQUANTITY"] = (
                    f"PENDINGQUANTITY: Value is negative ({str(qty_raw).strip()})"
                )

        return reasons


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class ScheduledReceiptWOBusinessTableValidator:

    def __init__(self, wo_path: str, part_source_make_path: str):
        self.wo_path                = wo_path
        self.part_source_make_path  = part_source_make_path
        self.df                     = pd.DataFrame()
        self.valid_combinations     = set()
        self.error_map              = {}   # row_idx -> [failed field names]
        self.reason_map             = {}   # row_idx -> {field: reason}

        # counters for Summary sheet
        self.partname_mismatch_count = 0
        self.negative_qty_count      = 0

    def load(self):
        self.df = pd.read_csv(self.wo_path, sep="\t", dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def load_lookup(self):
        lookup_df = pd.read_csv(self.part_source_make_path, sep="\t", dtype=str)
        lookup_df.columns = [c.strip().upper() for c in lookup_df.columns]

        combos = set()
        for _, row in lookup_df.iterrows():
            material = row.get(PSM_MATERIAL_COL, "")
            plant    = row.get(PSM_PLANT_COL, "")

            if pd.isna(material) or pd.isna(plant):
                continue

            material = str(material).strip()
            plant    = str(plant).strip()

            if material == "" or plant == "":
                continue

            combos.add((material, plant))

        self.valid_combinations = combos

    def validate(self):
        engine = ScheduledReceiptWOBusinessRuleEngine(self.valid_combinations)

        for idx, row in self.df.iterrows():
            try:
                reasons = engine.validate_row(row)
            except Exception:
                reasons = {}

            if reasons:
                self.error_map[idx]  = list(reasons.keys())
                self.reason_map[idx] = reasons

                if "PARTNAME" in reasons:
                    self.partname_mismatch_count += 1
                if "PENDINGQUANTITY" in reasons:
                    self.negative_qty_count += 1

    def get_error_series(self) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            result[idx] = " | ".join(col_reason.values())
        return pd.Series(result, dtype=str)

    def get_field_error_series(self, field_name: str) -> pd.Series:
        result = {}
        for idx, col_reason in self.reason_map.items():
            if field_name in col_reason:
                result[idx] = col_reason[field_name]
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class ScheduledReceiptWOBusinessReportWriter:

    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "PARTNAME": [
            "The PARTNAME-PARTSITE combination must exist as a MATERIAL-PLANT "
            "combination in the PartSource(Make) extract.",
        ],
        "PENDINGQUANTITY": [
            "The value must not be negative.",
        ],
    }

    def __init__(self, validator: ScheduledReceiptWOBusinessTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            if col_name == "ERROR_COLUMNS":
                cell.fill = WHITE_FILL
                cell.font = Font(bold=True, name="Arial", color="000000")
            else:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    def _style_summary_data_row(self, ws, row_num: int, num_cols: int = 7,
                                bold: bool = False, fill: PatternFill = None):
        for c in range(1, num_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=bold, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    # ══════════════════════════════════════════
    #  Summary sheet
    # ══════════════════════════════════════════
    def _write_summary_sheet_into(self, ws, error_map: dict, total_rows: int):
        v = self.validator

        # ── Row 1 : Title ──
        ws.merge_cells("A1:G1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="ScheduledReceipt (WO) Business Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = SUMM_TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # ── Row 2 : Column headers ──
        headers = ["#", "Field Name", "Error Count", "Record Count",
                   "% Health", "% of Error", "Reason / Sub-Category"]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=2, column=c_idx, value=h)
            cell.fill      = SUMM_HDR_FILL
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 30

        row_num = 3

        field_errors = v.get_errors_by_field()
        field_counts = {
            "PARTNAME":        len(field_errors.get("PARTNAME", [])),
            "PENDINGQUANTITY": len(field_errors.get("PENDINGQUANTITY", [])),
        }

        field_reasons = {
            "PARTNAME":        "PARTNAME-PARTSITE combination not found in PartSource(Make)",
            "PENDINGQUANTITY": "Value is negative",
        }

        sorted_fields = sorted(FIELD_ORDER, key=lambda f: field_counts.get(f, 0), reverse=True)

        for field_num, field in enumerate(sorted_fields, start=1):
            count      = field_counts.get(field, 0)
            pct_error  = round((count / total_rows) * 100, 2) if total_rows else 0
            pct_health = round(100 - pct_error, 2)

            ws.cell(row=row_num, column=1, value=field_num)
            ws.cell(row=row_num, column=2, value=field)
            ws.cell(row=row_num, column=3, value=count)
            ws.cell(row=row_num, column=4, value=total_rows)
            ws.cell(row=row_num, column=5, value=f"{pct_health}%")
            ws.cell(row=row_num, column=6, value=f"{pct_error}%")
            ws.cell(row=row_num, column=7, value=field_reasons.get(field, "") if count else "")
            self._style_summary_data_row(ws, row_num, bold=True, fill=WHITE_FILL)
            ws.cell(row=row_num, column=7).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            row_num += 1

        # ── TOTAL row ──
        total_errors        = sum(field_counts.values())
        total_record_count  = total_rows * len(FIELD_ORDER)
        total_pct_error     = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
        total_pct_health    = round(100 - total_pct_error, 2)

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=total_errors)
        ws.cell(row=row_num, column=4, value=total_record_count)
        ws.cell(row=row_num, column=5, value=f"{total_pct_health}%")
        ws.cell(row=row_num, column=6, value=f"{total_pct_error}%")
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            ws.cell(row=row_num, column=c).font      = Font(name="Arial", bold=True, size=10)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center", vertical="center")

        row_num += 2   # blank spacer

        # ── Quick-glance stats block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 42, 14, 16, 12, 12, 70]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets  (ALL source columns) ──
    def _write_field_error_sheets(self, wb, df: pd.DataFrame, all_cols: list):
        field_errors = self.validator.get_errors_by_field()

        for field_name in FIELD_ORDER:
            if field_name not in field_errors:
                continue

            row_indices = field_errors[field_name]
            sheet_name  = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws          = wb.create_sheet(sheet_name)

            keep_here = all_cols + ["ERROR_COLUMNS"]
            subset    = df.loc[row_indices, keep_here].copy()

            field_err_series        = self.validator.get_field_error_series(field_name)
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i: field_err_series.get(i, "")
            )

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, (col, value) in enumerate(zip(subset.columns, row_data), start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = WHITE_FILL
                    cell.border    = THIN_BORDER

                # Highlight only the column(s) relevant to this field's rule.
                for involved_col in FIELD_HIGHLIGHT_COLS.get(field_name, []):
                    if involved_col in col_idx_map:
                        target_cell      = ws.cell(row=excel_row, column=col_idx_map[involved_col])
                        target_cell.fill = RED_FILL
                        target_cell.font = ERR_FONT

            self._set_widths(ws)
            ws.freeze_panes = "A2"

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        ws.merge_cells("A1:C1")
        title_cell           = ws.cell(row=1, column=1,
                                        value="ScheduledReceipt (WO) – Business Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field in FIELD_ORDER:
            rules_list = self.RULES_CONTENT.get(field, [])
            num_rules  = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1,
                                             value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2,
                                               value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 90

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        # All source columns, in original order — used for per-field error sheets
        all_cols = list(df.columns)

        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        wb               = Workbook()
        ws_summary       = wb.active
        ws_summary.title = self.SHEET_SUMMARY
        self._write_summary_sheet_into(ws_summary, v.error_map, total_rows=len(df))

        self._write_rules_sheet(wb)
        self._write_field_error_sheets(wb, df, all_cols)

        wb.save(self.output_path)

        fields_with_errors = [f for f in FIELD_ORDER if f in v.get_errors_by_field()]
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows              : {len(df)}")
        print(f"   Error rows              : {len(v.error_map)}")
        print(f"     - PARTNAME mismatch    : {v.partname_mismatch_count}")
        print(f"     - Negative PendingQty  : {v.negative_qty_count}")
        print(f"   Field sheets             : {fields_with_errors}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class ScheduledReceiptWOBusinessTableProcessor:

    def __init__(self, wo_path: str, part_source_make_path: str, output_path: str):
        self.validator = ScheduledReceiptWOBusinessTableValidator(wo_path, part_source_make_path)
        self.writer    = ScheduledReceiptWOBusinessReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        self.validator.load_lookup()
        print(f"    ScheduledReceipt(WO) columns detected : {list(self.validator.df.columns)}")
        print(f"    PartSource(Make) valid combinations   : {len(self.validator.valid_combinations)}")
        print("🔍  Validating business rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = ScheduledReceiptWOBusinessTableProcessor(
        wo_path               = WO_INPUT_FILE,
        part_source_make_path = PART_SOURCE_MAKE_FILE,
        output_path           = OUTPUT_FILE,
    )
    processor.run()
