import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURATION - UPDATE THIS DAILY!
# ─────────────────────────────────────────────
TARGET_DATE = "17 June"  # Change to "18 June", "19 June", etc.

# ─────────────────────────────────────────────
#  FILE PATHS FOR MODULES
# ─────────────────────────────────────────────
BASE_DIR = os.path.join(r"C:\Users\SW526XH\Downloads\Go Live-2\Validated Outputs")

INPUT_FILES = {
    "Allocation":                       os.path.join(BASE_DIR, "Validated_Allocation.xlsx"),
    "Batch":                            os.path.join(BASE_DIR, "Validated_Batch.xlsx"),
    "BillOfMaterial":                   os.path.join(BASE_DIR, "Validated_BillOfMaterial.xlsx"),
    "Constraints_WC":                   os.path.join(BASE_DIR, "Validated_Constraints_WC.xlsx"),
    "HistoricalSupplyActuals_Buy":      os.path.join(BASE_DIR, "Validated_HistoricalSupplyActuals_Buy.xlsx"),
    "HistoricalSupplyActuals_Make":     os.path.join(BASE_DIR, "Validated_HistoricalSupplyActuals_Make.xlsx"),
    "HistoricalSupplyActuals_Transfer": os.path.join(BASE_DIR, "Validated_HistoricalSupplyActuals_Transfer.xlsx"),
    "OnHand":                           os.path.join(BASE_DIR, "Validated_OnHand.xlsx"),
    "Part_RM":                          os.path.join(BASE_DIR, "Validated_Part_RM.xlsx"),
    "PartSource_Buy":                   os.path.join(BASE_DIR, "Validated_PartSource_Buy.xlsx"),
    "PartSource_Make":                  os.path.join(BASE_DIR, "Validated_PartSource_Make.xlsx"),
    "PartSource_Transfer":              os.path.join(BASE_DIR, "Validated_PartSource_Transfer.xlsx"),
    "ScheduleReceipt_PO_from_Vend":     os.path.join(BASE_DIR, "Validated_ScheduleReceipt_PO_from_Vend.xlsx"),
    "ScheduleReceipt_STO":              os.path.join(BASE_DIR, "Validated_ScheduleReceipt_STO.xlsx"),
    "ScheduleReceipt_SubCon_PO":        os.path.join(BASE_DIR, "Validated_ScheduleReceipt_SubCon_PO.xlsx"),
    "ScheduleReceipt_WO":               os.path.join(BASE_DIR, "Validated_ScheduledReceipt_WO.xlsx"),
    "Source_Constraint":                os.path.join(BASE_DIR, "Validated_SourceConstraint.xlsx"),
}

OUTPUT_WORKBOOK = os.path.join(
    r"C:\Users\SW526XH\Downloads\Go Live-2\Summary",
    f"GL2_Summary_{TARGET_DATE}.xlsx"
)

# ─────────────────────────────────────────────
#  INTERFACE MAPPING
#  type: "Master" | "Transaction"
# ─────────────────────────────────────────────
INTERFACE_MAPPING = {
    "Allocation":                       {"type": "Transaction", "name": "Allocation"},
    "Batch":                            {"type": "Transaction", "name": "Batch"},
    "BillOfMaterial":                   {"type": "Master",      "name": "Bill Of Material"},
    "Constraints_WC":                   {"type": "Master",      "name": "Constraints (WC)"},
    "HistoricalSupplyActuals_Buy":      {"type": "Transaction", "name": "Historical Supply Actuals (Buy)"},
    "HistoricalSupplyActuals_Make":     {"type": "Transaction", "name": "Historical Supply Actuals (Make)"},
    "HistoricalSupplyActuals_Transfer": {"type": "Transaction", "name": "Historical Supply Actuals (Transfer)"},
    "OnHand":                           {"type": "Transaction", "name": "On Hand"},
    "Part_RM":                          {"type": "Master",      "name": "Part (RM)"},
    "PartSource_Buy":                   {"type": "Master",      "name": "Part Source (Buy)"},
    "PartSource_Make":                  {"type": "Master",      "name": "Part Source (Make)"},
    "PartSource_Transfer":              {"type": "Master",      "name": "Part Source (Transfer)"},
    "ScheduleReceipt_PO_from_Vend":    {"type": "Transaction", "name": "Schedule Receipt (PO from Vend)"},
    "ScheduleReceipt_STO":              {"type": "Transaction", "name": "Schedule Receipt (STO)"},
    "ScheduleReceipt_SubCon_PO":        {"type": "Transaction", "name": "Schedule Receipt (Sub-Con PO)"},
    "ScheduleReceipt_WO":               {"type": "Transaction", "name": "Schedule Receipt (WO)"},
    "Source_Constraint":                {"type": "Master",      "name": "Source Constraint"},
}

# ─────────────────────────────────────────────
#  COLOURS & STYLES
# ─────────────────────────────────────────────
TITLE_FILL      = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FILL        = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
TOTAL_FILL      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL      = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
MASTER_FILL     = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TXN_FILL        = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
OVERALL_FILL    = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")

HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=9)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Data Extractor
# ══════════════════════════════════════════════
def get_na_placeholder_data():
    return {
        'fields': [{
            'num':       1,
            'name':      'N/A - File Missing',
            'err_count': 'N/A',
            'rec_count': 'N/A',
            'health_pct': 'N/A',
            'err_pct':   'N/A',
            'reason':    'Input file not found',
        }],
        'total': {
            'err_count':  'N/A',
            'rec_count':  'N/A',
            'health_pct': 'N/A',
            'err_pct':    'N/A',
        },
        'stats': {
            'total_records':       'N/A',
            'records_with_errors': 'N/A',
            'records_passing':     'N/A',
        },
    }


def extract_summary_data(filepath):
    """
    Reads the 'Summary' sheet from a validator output file.
    Extracts field rows, TOTAL row, and the Stats block.
    Returns N/A placeholder if the file is missing or unreadable.
    """
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return get_na_placeholder_data()

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [ERROR] Cannot open {filepath}: {e}")
        return get_na_placeholder_data()

    summary_sheet_name = next(
        (name for name in wb.sheetnames if "summary" in name.lower()), None
    )
    if not summary_sheet_name:
        print(f"  [WARN] No Summary sheet found in {filepath}")
        return get_na_placeholder_data()

    ws          = wb[summary_sheet_name]
    fields_data = []
    total_row   = None
    stats       = {}
    in_data     = False

    for row in ws.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            continue

        label = str(row[0]).strip() if row[0] else ""

        # Stats block
        if label == "Total Records:":
            stats['total_records'] = row[2]
            continue
        elif label == "Records with Errors:":
            stats['records_with_errors'] = row[2]
            continue
        elif label == "Records Passing:":
            stats['records_passing'] = row[2]
            continue

        # Header row
        if row[0] == "#" and row[1] == "Field Name":
            in_data = True
            continue

        # TOTAL row
        if in_data and row[1] == "TOTAL":
            total_row = {
                'err_count':  row[2],
                'rec_count':  row[3],
                'health_pct': row[4],
                'err_pct':    row[5],
            }
            in_data = False
            continue

        # Field-level rows
        if in_data and row[1]:
            fields_data.append({
                'num':       row[0],
                'name':      row[1],
                'err_count': row[2],
                'rec_count': row[3],
                'health_pct': row[4],
                'err_pct':   row[5],
                'reason':    row[6] if len(row) > 6 else "",
            })

    wb.close()
    return {'fields': fields_data, 'total': total_row, 'stats': stats}


# ══════════════════════════════════════════════
#  Per-Module Detail Sheet
# ══════════════════════════════════════════════
def apply_module_sheet(wb, sheet_name, data):
    """
    Writes the extracted data into a new sheet replicating
    the validator Summary sheet format.
    """
    ws = wb.create_sheet(sheet_name)

    # Title
    display_name = INTERFACE_MAPPING.get(sheet_name, {}).get("name", sheet_name.replace("_", " "))
    title_cell = ws.cell(row=1, column=1, value=f"{display_name} Validation Summary")
    title_cell.font      = Font(name="Arial", bold=True, size=14)
    title_cell.fill      = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers    = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
    col_widths = [6, 38, 16, 16, 14, 14, 70]

    for c_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=3, column=c_idx, value=h)
        cell.fill      = TITLE_FILL
        cell.font      = Font(name="Arial", bold=True)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 4
    for field in data['fields']:
        ws.cell(row=row_num, column=1, value=field['num'])
        ws.cell(row=row_num, column=2, value=field['name'])
        ws.cell(row=row_num, column=3, value=field['err_count'])
        ws.cell(row=row_num, column=4, value=field['rec_count'])
        ws.cell(row=row_num, column=5, value=field['health_pct'])
        ws.cell(row=row_num, column=6, value=field['err_pct'])
        ws.cell(row=row_num, column=7, value=field['reason'])

        for c in range(1, 8):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left" if c in (2, 7) else "center",
                wrap_text=(c == 7),
            )

        row_num += 1

    # TOTAL row
    if data['total']:
        ws.cell(row=row_num, column=2, value="TOTAL")
        ws.cell(row=row_num, column=3, value=data['total']['err_count'])
        ws.cell(row=row_num, column=4, value=data['total']['rec_count'])
        ws.cell(row=row_num, column=5, value=data['total']['health_pct'])
        ws.cell(row=row_num, column=6, value=data['total']['err_pct'])
        ws.cell(row=row_num, column=7, value="")

        for c in range(1, 8):
            cell           = ws.cell(row=row_num, column=c)
            cell.font      = Font(name="Arial", bold=True)
            cell.fill      = TOTAL_FILL
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    row_num += 2

    # Stats block
    for label, key in [
        ("Total Records:",       'total_records'),
        ("Records with Errors:", 'records_with_errors'),
        ("Records Passing:",     'records_passing'),
    ]:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        lc           = ws.cell(row=row_num, column=1, value=label)
        lc.font      = Font(name="Arial", bold=True, size=10)
        lc.fill      = STATS_FILL
        lc.border    = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")

        vc           = ws.cell(row=row_num, column=3, value=data['stats'].get(key, ''))
        vc.font      = Font(name="Arial", size=10)
        vc.border    = THIN_BORDER
        vc.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    for c_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width


# ══════════════════════════════════════════════
#  Overall Summary Sheet
# ══════════════════════════════════════════════
def create_master_summary_sheet(wb, module_data_dict):
    ws = wb.create_sheet("GL2 Summary", 0)

    current_date  = datetime.now().strftime("%d-%b-%Y")
    health_header = f"Health % (as on {current_date})"

    # ── Helpers ──
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    def _style(cell, font=None, fill=None, bold=False, border=True, align=None):
        cell.font      = font or Font(name="Arial", bold=bold, size=10)
        cell.alignment = align or center
        if fill:
            cell.fill = fill
        if border:
            cell.border = THIN_BORDER

    def _hdr(ws, row, cols_values, fill=HDR_FILL):
        for c, val in enumerate(cols_values, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            _style(cell, font=Font(name="Arial", bold=True, size=10), fill=fill)

    # ── Build records from extracted data ──
    records = []
    for mod_key, info in INTERFACE_MAPPING.items():
        data = module_data_dict.get(mod_key)
        if not data:
            continue

        is_na = (data['stats'].get('total_records') == 'N/A')

        try:
            records_with_errors = int(data['stats'].get('records_with_errors', 0)) if not is_na else 0
        except (TypeError, ValueError):
            records_with_errors = 0

        try:
            total_records = int(data['stats'].get('total_records', 0)) if not is_na else 0
        except (TypeError, ValueError):
            total_records = 0

        # Health % = 1 - (records_with_errors / total_records)
        if total_records > 0:
            health_pct = 1 - (records_with_errors / total_records)
        else:
            health_pct = None

        records.append({
            "type":                  info["type"],
            "name":                  info["name"],
            "records_with_errors":   records_with_errors,
            "total_records":         total_records,
            "health_pct":            health_pct,
            "is_na":                 is_na,
        })

    # ── Aggregate totals ──
    def _aggregate(recs):
        err   = sum(r["records_with_errors"] for r in recs if not r["is_na"])
        total = sum(r["total_records"]        for r in recs if not r["is_na"])
        hp    = (1 - err / total) if total > 0 else None
        return err, total, hp

    master_recs  = [r for r in records if r["type"] == "Master"]
    txn_recs     = [r for r in records if r["type"] == "Transaction"]
    all_recs     = records

    m_err, m_total, m_health   = _aggregate(master_recs)
    t_err, t_total, t_health   = _aggregate(txn_recs)
    o_err, o_total, o_health   = _aggregate(all_recs)

    # ════════════════════════════════
    #  TABLE 1 — High-level roll-up
    # ════════════════════════════════
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value=f"GL-2 Data Validation Summary  |  {current_date}")
    title.font      = Font(name="Arial", bold=True, size=14)
    title.fill      = OVERALL_FILL
    title.alignment = left

    _hdr(ws, 3, ["Data", "Records with Errors", "Total Records", health_header])

    def _write_agg_row(r, label, err, total, health, fill=None):
        vals = [label, err, total, health]
        for c, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            bold = (label == "Overall")
            _style(cell, font=Font(name="Arial", bold=bold, size=10), fill=fill)
            if c == 4 and isinstance(val, float):
                cell.number_format = '0.00%'

    _write_agg_row(4, "Master",      m_err, m_total, m_health, fill=MASTER_FILL)
    _write_agg_row(5, "Transaction", t_err, t_total, t_health, fill=TXN_FILL)
    _write_agg_row(6, "Overall",     o_err, o_total, o_health, fill=OVERALL_FILL)

    # ════════════════════════════════
    #  TABLE 2 — Master breakdown
    # ════════════════════════════════
    r = 9
    ws.cell(row=r-1, column=1, value="Master Data").font = Font(name="Arial", bold=True, size=11)

    _hdr(ws, r, ["Data", "Interface", "Records with Errors", "Total Records", health_header],
         fill=MASTER_FILL)
    r += 1

    for rec in master_recs:
        health_val = rec["health_pct"] if not rec["is_na"] else "N/A"
        row_vals   = ["Master", rec["name"], rec["records_with_errors"],
                      rec["total_records"], health_val]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            _style(cell, font=Font(name="Arial", size=10),
                   align=left if c == 2 else center)
            if c == 5 and isinstance(val, float):
                cell.number_format = '0.00%'
        r += 1

    # Master sub-total
    for c, val in enumerate(["", "Sub-Total", m_err, m_total, m_health], start=1):
        cell = ws.cell(row=r, column=c, value=val)
        _style(cell, font=Font(name="Arial", bold=True, size=10), fill=TOTAL_FILL)
        if c == 5 and isinstance(val, float):
            cell.number_format = '0.00%'
    r += 2

    # ════════════════════════════════
    #  TABLE 3 — Transaction breakdown
    # ════════════════════════════════
    ws.cell(row=r-1, column=1, value="Transaction Data").font = Font(name="Arial", bold=True, size=11)

    _hdr(ws, r, ["Data", "Interface", "Records with Errors", "Total Records", health_header],
         fill=TXN_FILL)
    r += 1

    for rec in txn_recs:
        health_val = rec["health_pct"] if not rec["is_na"] else "N/A"
        row_vals   = ["Transaction", rec["name"], rec["records_with_errors"],
                      rec["total_records"], health_val]
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            _style(cell, font=Font(name="Arial", size=10),
                   align=left if c == 2 else center)
            if c == 5 and isinstance(val, float):
                cell.number_format = '0.00%'
        r += 1

    # Transaction sub-total
    for c, val in enumerate(["", "Sub-Total", t_err, t_total, t_health], start=1):
        cell = ws.cell(row=r, column=c, value=val)
        _style(cell, font=Font(name="Arial", bold=True, size=10), fill=TOTAL_FILL)
        if c == 5 and isinstance(val, float):
            cell.number_format = '0.00%'
    r += 2

    # ── Overall row ──
    for c, val in enumerate(["Overall", "", o_err, o_total, o_health], start=1):
        cell = ws.cell(row=r, column=c, value=val)
        _style(cell, font=Font(name="Arial", bold=True, size=10), fill=OVERALL_FILL)
        if c == 5 and isinstance(val, float):
            cell.number_format = '0.00%'

    # ── Column widths ──
    for col_letter, width in zip(["A", "B", "C", "D", "E"], [18, 35, 22, 18, 25]):
        ws.column_dimensions[col_letter].width = width


# ══════════════════════════════════════════════
#  Execution Pipeline
# ══════════════════════════════════════════════
def compile_gl2_summary():
    print("=" * 60)
    print("  GL-2 Summary Compiler")
    print("=" * 60)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    all_module_data = {}

    for module_name, filepath in INPUT_FILES.items():
        print(f"\nProcessing: {module_name}")
        data = extract_summary_data(filepath)
        apply_module_sheet(wb, module_name, data)
        all_module_data[module_name] = data
        print(f"  [OK] Sheet written.")

    create_master_summary_sheet(wb, all_module_data)

    os.makedirs(os.path.dirname(OUTPUT_WORKBOOK), exist_ok=True)
    print(f"\n[SAVE] Saving to:\n  {OUTPUT_WORKBOOK}")
    wb.save(OUTPUT_WORKBOOK)
    print("\nDone ✅")


if __name__ == "__main__":
    compile_gl2_summary()
