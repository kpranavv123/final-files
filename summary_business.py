import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURATION - UPDATE THIS DAILY!
# ─────────────────────────────────────────────
TARGET_DATE = "27-May"  # Change to "18 apr", "20 apr", etc.

# ─────────────────────────────────────────────
#  FILE PATHS FOR MODULES
# ─────────────────────────────────────────────
BASE_DIR = os.path.join(r"C:\Users\SW526XH\Downloads\Go Live-1\Business")

INPUT_FILES = {
    "Part": os.path.join(BASE_DIR, "Validated_Part_Business_Rulesets_with_IBPStatus.xlsx"),
    "Site": os.path.join(BASE_DIR, "Validated_Site_business.xlsx"), 
    "Customer": os.path.join(BASE_DIR, "Validated_Customer_Business.xlsx"),
    "Part_UOM": os.path.join(BASE_DIR, "Validated_PartUOM_Business.xlsx"),
    "Product_Hierarchy": os.path.join(BASE_DIR, "Validated_ProductHierarchy_Business.xlsx"),
    "ID": os.path.join(BASE_DIR, "Validated_IndependentDemand_Business.xlsx"),
    "HDA": os.path.join(BASE_DIR, "Validated_HDA_Business.xlsx"),
    "HDA_Secondary": os.path.join(BASE_DIR, "Validated_HDA_Secondary_Business.xlsx")
}

OUTPUT_WORKBOOK = os.path.join(r"C:\Users\SW526XH\Downloads\Go Live-1\Summary\outputs\Business", f"summary2_{TARGET_DATE}.xlsx")

# ─────────────────────────────────────────────
#  COLOURS & STYLES (CUSTOMER TEMPLATE)
# ─────────────────────────────────────────────
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL  = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

HDR_FONT    = Font(bold=True, name="Arial", size=9)
BODY_FONT   = Font(name="Arial", size=9)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

# ══════════════════════════════════════════════
#  Data Extractor
# ══════════════════════════════════════════════
def get_na_placeholder_data():
    """
    Returns a placeholder data structure with N/A values for missing files.
    """
    return {
        'fields': [{
            'num': 1,
            'name': 'N/A - File Missing',
            'err_count': 'N/A',
            'rec_count': 'N/A',
            'health_pct': 'N/A',
            'err_pct': 'N/A',
            'reason': 'Input file not found'
        }],
        'total': {
            'err_count': 'N/A',
            'rec_count': 'N/A',
            'health_pct': 'N/A',
            'err_pct': 'N/A'
        },
        'stats': {
            'total_records': 'N/A',
            'records_with_errors': 'N/A',
            'records_passing': 'N/A'
        }
    }

def extract_summary_data(filepath):
    """
    Reads the 'Summary' sheet from a generated validator output file,
    extracting the field error rows, the TOTAL row, and the Stats block.
    Returns a dictionary of data to be identically re-generated.
    If the file is missing, returns a placeholder with N/A values.
    """
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return get_na_placeholder_data()

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [ERROR] Cannot open {filepath}: {e}")
        return get_na_placeholder_data()

    summary_sheet_name = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            summary_sheet_name = name
            break
            
    if not summary_sheet_name:
        print(f"  [WARN] No Summary sheet found in {filepath}")
        return get_na_placeholder_data()
        
    ws = wb[summary_sheet_name]
    
    fields_data = []
    total_row   = None
    stats       = {}
    
    in_data = False
    for row in ws.iter_rows(values_only=True):

        # 1️⃣ Skip empty rows
        if all(cell is None for cell in row):
            continue

        # 2️⃣ Read stats block (comes AFTER TOTAL)
        label = str(row[0]).strip() if row[0] else ""

        if label == "Total Records:":
            stats['total_records'] = row[2]
            continue
        elif label == "Records with Errors:":
            stats['records_with_errors'] = row[2]
            continue
        elif label == "Records Passing:":
            stats['records_passing'] = row[2]
            continue

        # 3️⃣ Detect header row
        if row[1] and str(row[1]).strip() in ["Field Name", "Rule Name"]:
            in_data = True
            continue

        # 4️⃣ Detect TOTAL row (end of table)
        if in_data and row[1] == "TOTAL":
            total_row = {
                'err_count': row[2],
                'rec_count': row[3],
                'health_pct': row[4],
                'err_pct': row[5]
            }
            in_data = False
            continue

        # 5️⃣ Read field-level summary rows
        if in_data and row[1]:
            fields_data.append({
                'num': row[0],
                'name': row[1],
                'err_count': row[2],
                'rec_count': row[3],
                'health_pct': row[4],
                'err_pct': row[5],
                'reason': row[6] if len(row) > 6 else ""
            })



    wb.close()    
    
# ✅ Debug warning if TOTAL row is missing
    
# ✅ Debug checks
    if not fields_data:
        print(f"  [WARN] No data rows found in {filepath}")

    if total_row is None:
     print(f"  [WARN] TOTAL row missing in {filepath}")
     total_row = {
        'err_count': 'N/A',
        'rec_count': 'N/A',
        'health_pct': 'N/A',
        'err_pct': 'N/A'
    }

# ✅ FINAL SAFE RETURN (always executes)
    return {
    'fields': fields_data if fields_data else get_na_placeholder_data()['fields'],
    'total': total_row,
    'stats': stats
}

   
    # return {'fields': fields_data, 'total': total_row, 'stats': stats}

# ══════════════════════════════════════════════
#  Writer (CUSTOMER FORMAT)
# ══════════════════════════════════════════════
def apply_customer_template_sheet(wb, sheet_name, data):
    """
    Writes the extracted data into a new sheet exactly mimicking
    the Customer validation template styles.
    """
    ws = wb.create_sheet(sheet_name)
    
    # ── Title ──
    title_text = f"{sheet_name.replace('_', ' ')} Validation Summary"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Arial", bold=True, size=14)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 24
    
    # ── Headers ──
    headers = ["#", "Field Name", "Error Count", "Record Count", "% Health", "% of Error", "Reason"]
    col_widths = [6, 35, 16, 16, 16, 16, 70]
    
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.fill = TITLE_FILL
        cell.font = Font(name="Arial", bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    row_num = 4
    for field in data['fields']:
        ws.cell(row=row_num, column=1, value=field['num']).font = BODY_FONT
        ws.cell(row=row_num, column=2, value=field['name']).font = BODY_FONT
        ws.cell(row=row_num, column=3, value=field['err_count']).font = BODY_FONT
        ws.cell(row=row_num, column=4, value=field['rec_count']).font = BODY_FONT
        
        cell_health = ws.cell(row=row_num, column=5, value=field['health_pct'])
        cell_health.font = BODY_FONT
        if isinstance(field['health_pct'], (int, float)):
            cell_health.number_format = '0.00%'
            
        cell_err = ws.cell(row=row_num, column=6, value=field['err_pct'])
        cell_err.font = BODY_FONT
        if isinstance(field['err_pct'], (int, float)):
            cell_err.number_format = '0.00%'
            
        ws.cell(row=row_num, column=7, value=field['reason']).font = BODY_FONT
        
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).border = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center" if c != 7 else "left", wrap_text=(c == 7))
            
        row_num += 1
        
    # ── TOTAL row ──
    if data['total']:
        ws.cell(row=row_num, column=2, value="TOTAL").font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=data['total']['err_count']).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=data['total']['rec_count']).font = Font(name="Arial", bold=True)
        
        c5 = ws.cell(row=row_num, column=5, value=data['total']['health_pct'])
        c5.font = Font(name="Arial", bold=True)
        if isinstance(data['total']['health_pct'], (int, float)):
             c5.number_format = '0.00%'
             
        c6 = ws.cell(row=row_num, column=6, value=data['total']['err_pct'])
        c6.font = Font(name="Arial", bold=True)
        if isinstance(data['total']['err_pct'], (int, float)):
             c6.number_format = '0.00%'
             
        ws.cell(row=row_num, column=7, value="").font = Font(name="Arial", bold=True)
        
        for c in range(1, 8):
            ws.cell(row=row_num, column=c).fill = TOTAL_FILL
            ws.cell(row=row_num, column=c).border = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center" if c != 7 else "left")
            
    row_num += 2
    
    # ── Stats Block ──
    stats_map = [
        ("Total Records:", data['stats'].get('total_records', '')),
        ("Records with Errors:", data['stats'].get('records_with_errors', '')),
        ("Records Passing:", data['stats'].get('records_passing', ''))
    ]
    
    for label, value in stats_map:
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = Font(name="Arial", bold=True, size=10)
        label_cell.fill = STATS_FILL
        label_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        
        value_cell = ws.cell(row=row_num, column=3, value=value)
        value_cell.font = Font(name="Arial", size=10)
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    # Apply specific column widths
    for c_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width


# ══════════════════════════════════════════════
#  Master Summary Sheet
# ══════════════════════════════════════════════
def create_master_summary_sheet(wb, module_data_dict):
    """
    Creates the overarching 'Summary' sheet combining all modules metrics.
    """
    ws = wb.create_sheet("Summary", 0)
    
    # ── Title ──
    title_cell = ws.cell(row=1, column=1, value="Summary")
    title_cell.font = Font(name="Arial", bold=True, italic=True, size=14)
    
    current_date = datetime.now().strftime("%d-%b")
    health_header = f"Health% as on {current_date}"
    
    INTERFACE_MAPPING = {
        "Part": {"type": "Master", "name": "Part"},
        "Site": {"type": "Master", "name": "Site"},
        "Customer": {"type": "Master", "name": "Customer"},
        "Part_UOM": {"type": "Master", "name": "PartUOMConversion"},
        "Product_Hierarchy": {"type": "Master", "name": "Product Hierarchy"},
        "HDA": {"type": "Transaction", "name": "HDA"},
        "HDA_Secondary": {"type": "Transaction", "name": "HDA Secondary Sales"},
        "ID": {"type": "Transaction", "name": "Independent demand"}
    }
    
    records = []
    for mod_key, info in INTERFACE_MAPPING.items():
        data = module_data_dict.get(mod_key)
        if data:
            try:
                err_count = float(data['total']['err_count']) if data['total']['err_count'] != 'N/A' else 0
            except:
                err_count = 0
            
            try:
                total_rec = float(data['total']['rec_count']) if data['total']['rec_count'] != 'N/A' else 0
            except:
                total_rec = 0

            records.append({
                "type": info["type"],
                "name": info["name"],
                "err": err_count,
                "rec": total_rec,
                "is_na": data['total']['err_count'] == 'N/A'
            })
            
    master_err = sum(r["err"] for r in records if r["type"] == "Master" and not r.get("is_na", False))
    master_rec = sum(r["rec"] for r in records if r["type"] == "Master" and not r.get("is_na", False))
    txn_err = sum(r["err"] for r in records if r["type"] == "Transaction" and not r.get("is_na", False))
    txn_rec = sum(r["rec"] for r in records if r["type"] == "Transaction" and not r.get("is_na", False))
    overall_err = master_err + txn_err
    overall_rec = master_rec + txn_rec
    
    def calc_health(err, rec):
        if rec == 0: return 0
        h = 1 - (err / rec)
        return max(0, h)

    header_fill = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")
    center_align = Alignment(horizontal="center", vertical="center")
    
    def style_cell(cell, font_style, is_header=False):
        cell.font = font_style
        cell.border = THIN_BORDER
        cell.alignment = center_align
        if is_header:
            cell.fill = header_fill

    # TABLE 1
    ws.merge_cells("A3:B3")
    style_cell(ws.cell(row=3, column=1, value="Data"), header_font, is_header=True)
    style_cell(ws.cell(row=3, column=2), header_font, is_header=True)
    style_cell(ws.cell(row=3, column=3, value="Total Error Count"), header_font, is_header=True)
    style_cell(ws.cell(row=3, column=4, value="Total records"), header_font, is_header=True)
    style_cell(ws.cell(row=3, column=5, value=health_header), header_font, is_header=True)

    def write_t1_row(r_num, name, err, rec, is_bold=False):
        ws.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=2)
        f = Font(name="Arial", bold=True) if is_bold else body_font
        style_cell(ws.cell(row=r_num, column=1, value=name), f)
        style_cell(ws.cell(row=r_num, column=2), f)
        style_cell(ws.cell(row=r_num, column=3, value=err), f)
        style_cell(ws.cell(row=r_num, column=4, value=rec), f)
        hc = ws.cell(row=r_num, column=5, value=calc_health(err, rec))
        style_cell(hc, f)
        hc.number_format = '0.00%'

    write_t1_row(4, "Master", master_err, master_rec)
    write_t1_row(5, "Transaction", txn_err, txn_rec)
    write_t1_row(6, "Overall", overall_err, overall_rec, is_bold=True)

    # TABLE 2
    r_idx = 8
    headers2 = ["Data", "Interface", "Total Error Count", "Total records", health_header]
    for c_idx, h in enumerate(headers2, start=1):
        style_cell(ws.cell(row=r_idx, column=c_idx, value=h), header_font, is_header=True)

    r_idx += 1
    for rec in (x for x in records if x["type"] == "Master"):
        style_cell(ws.cell(row=r_idx, column=1, value=rec["type"]), body_font)
        style_cell(ws.cell(row=r_idx, column=2, value=rec["name"]), body_font)
        style_cell(ws.cell(row=r_idx, column=3, value=rec["err"] if not rec.get("is_na", False) else "N/A"), body_font)
        style_cell(ws.cell(row=r_idx, column=4, value=rec["rec"] if not rec.get("is_na", False) else "N/A"), body_font)
        hc = ws.cell(row=r_idx, column=5, value="N/A" if rec.get("is_na", False) else calc_health(rec["err"], rec["rec"]))
        style_cell(hc, body_font)
        if not rec.get("is_na", False) and isinstance(hc.value, (int, float)):
            hc.number_format = '0.00%'
        r_idx += 1

    # TABLE 3
    r_idx += 1
    for c_idx, h in enumerate(headers2, start=1):
        style_cell(ws.cell(row=r_idx, column=c_idx, value=h), header_font, is_header=True)

    r_idx += 1
    for rec in (x for x in records if x["type"] == "Transaction"):
        style_cell(ws.cell(row=r_idx, column=1, value=rec["type"]), body_font)
        style_cell(ws.cell(row=r_idx, column=2, value=rec["name"]), body_font)
        style_cell(ws.cell(row=r_idx, column=3, value=rec["err"] if not rec.get("is_na", False) else "N/A"), body_font)
        style_cell(ws.cell(row=r_idx, column=4, value=rec["rec"] if not rec.get("is_na", False) else "N/A"), body_font)
        hc = ws.cell(row=r_idx, column=5, value="N/A" if rec.get("is_na", False) else calc_health(rec["err"], rec["rec"]))
        style_cell(hc, body_font)
        if not rec.get("is_na", False) and isinstance(hc.value, (int, float)):
            hc.number_format = '0.00%'
        r_idx += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25

# ══════════════════════════════════════════════
#  Execution Pipeline
# ══════════════════════════════════════════════
def compile_gl1_summary():
    print("="*60)
    print(" GL1 Summary Compiler (Customer Template Format)")
    print("="*60)
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    processed_count = 0
    all_module_data = {}
    
    for module_name, filepath in INPUT_FILES.items():
        print(f"\nProcessing Module: {module_name}")
        data = extract_summary_data(filepath)
        apply_customer_template_sheet(wb, module_name, data)
        all_module_data[module_name] = data
        print(f"  [SUCCESS] Extracted summary metrics and painted template.")
        processed_count += 1

    if processed_count > 0:
        create_master_summary_sheet(wb, all_module_data)
        print(f"\n[SAVE] Saving GL1 Workbook to:\n  {OUTPUT_WORKBOOK}")
        wb.save(OUTPUT_WORKBOOK)
        print("Done.")
    else:
        print("\n[INFO] No modules were processed. Workbook not saved.")

if __name__ == "__main__":
    compile_gl1_summary()
