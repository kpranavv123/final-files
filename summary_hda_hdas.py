import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ======================================================
# File paths
# ======================================================
HDA_PRIMARY_FILE   = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\BillingDocument(HDA)_2026-05-22-1152.tab"
HDA_SECONDARY_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA_Secondary\HDA(SecSales)2026-05-06-1606.tab"
 
PART_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Part\Part_Site_2026-05-21-1510.tab"
CUSTOMER_FILE = r"C:\Users\SW526XH\Downloads\Go Live-1\Customer\Cutomer_2026-05-20-1205.tab"
SITE_FILE     = r"C:\Users\SW526XH\Downloads\Go Live-1\Site\Site_2026-05-20-1153.tab"
 
OUTPUT_EXCEL = r"C:\Users\SW526XH\Downloads\Go Live-1\HDA\Technical_Summary3.xlsx"
 
# ======================================================
# Constants
# ======================================================
# NOTE: CHUNK_SIZE removed — full-file load is required for duplicate detection.
# Chunked processing cannot detect cross-chunk duplicates, so both HDA Primary
# and HDA Secondary are now loaded entirely into memory before validation.
date_pattern = re.compile(r"^\d{8}$|^\d{4}-\d{2}-\d{2}$")
 
# ======================================================
# Styling
# ======================================================
TITLE_FILL  = PatternFill("solid", fgColor="BDD7EE")
HDR_FILL    = PatternFill("solid", fgColor="D9E1F2")
RULE_FILL   = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL  = PatternFill("solid", fgColor="F2F2F2")
STATS_FILL  = PatternFill("solid", fgColor="EDEDED")
 
BOLD_FONT   = Font(name="Arial", bold=True)
TITLE_FONT  = Font(name="Arial", bold=True, size=14)
BODY_FONT   = Font(name="Arial", size=10)
 
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
 
# ======================================================
# Helper: Universal loader
# ======================================================
def read_input(path: str) -> pd.DataFrame:
    ext = path.lower()
    if ext.endswith((".csv", ".tab", ".txt")):
        return pd.read_csv(path, dtype=str, sep="\t")
    elif ext.endswith((".xlsx", ".xls")):
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported file type: {path}")
 
 
# ======================================================
# Load master/reference data
# ======================================================
print("📂 Loading reference files...")
 
part_df = read_input(PART_FILE)
part_df.columns = part_df.columns.str.strip().str.upper()
if "MATERIALNUMBER_PLANT" not in part_df.columns:
    part_df["MATERIALNUMBER_PLANT"] = (
        part_df["MATERIALNUMBER"].astype(str).str.strip() + "_" +
        part_df["PLANT"].astype(str).str.strip()
    )
part_plant_set = set(part_df["MATERIALNUMBER_PLANT"].dropna().str.strip())
part_set       = set(part_df["MATERIALNUMBER"].dropna().str.strip())
 
customer_df = read_input(CUSTOMER_FILE)
customer_df.columns = customer_df.columns.str.strip().str.upper()
if "SUPPLYINGPLANT_CUSTOMER" not in customer_df.columns:
    customer_df["SUPPLYINGPLANT_CUSTOMER"] = (
        customer_df["SUPPLYINGPLANT"].astype(str).str.strip() + "_" +
        customer_df["CUSTOMER"].astype(str).str.strip()
    )
customer_plant_set = set(customer_df["SUPPLYINGPLANT_CUSTOMER"].dropna().str.strip())
customer_set       = set(customer_df["CUSTOMER"].dropna().str.strip())
 
site_df = read_input(SITE_FILE)
site_df.columns = site_df.columns.str.strip().str.upper()
site_set = set(site_df["PLANT"].dropna().str.strip())
 
 
# ══════════════════════════════════════════════════════
#  HDA PRIMARY — rules & counting
#  Order: 1) MATERIAL_PLANT  2) PLANT  3) PLANT_SOLDTOPARTY
#         4) BILLING_DATE    5) DUPLICATE_CHECK
# ══════════════════════════════════════════════════════
PRIMARY_RULES = [
    ("MATERIAL_PLANT",     "ERROR_MATERIAL_PLANT",     "Material-Plant combination not present in the Part master."),
    ("PLANT",              "ERROR_PLANT",              "Plant is not present in site master."),
    ("PLANT_SOLDTOPARTY",  "ERROR_PLANT_SOLDTOPARTY",  "Plant-Soldtoparty combination is not present in customer master."),
    ("BILLING_DATE",       "ERROR_BILLING_DATE",       "Must not be blank and must be in YYYYMMDD format."),
    ("DUPLICATE_CHECK",    "ERROR_DUPLICATE_CHECK",    "Duplicate record: MATERIAL + PLANT + SOLDTOPARTY + BILLING_DATE combination already exists in the extract."),
]
 
PRIMARY_RULESET_DESC = {
    "MATERIAL_PLANT": [
        "Field should not be blank.",
        "Material-Plant combination must exist in Part master.",
    ],
    "PLANT": [
        "Must not be blank.",
        "Must exist in Site master.",
    ],
    "PLANT_SOLDTOPARTY": [
        "Field should not be blank.",
        "Plant-Soldtoparty combination must exist in Customer master.",
    ],
    "BILLING_DATE": [
        "Must not be blank.",
        "Must strictly be in YYYYMMDD format.",
    ],
    "DUPLICATE_CHECK": [
        "There should be no duplicate records in the extract.",
        "A duplicate is identified when MATERIAL + PLANT + SOLDTOPARTY + BILLING_DATE are all identical across two or more rows.",
        "All rows involved in a duplicate combination are flagged as errors.",
    ],
}
 
# ── Load full HDA Primary file ──────────────────────────────────────────────
print("\n📂 Loading HDA Primary file (full load for duplicate detection)...")
primary_df = pd.read_csv(HDA_PRIMARY_FILE, sep="\t", dtype=str)
primary_df = primary_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
primary_df.columns = primary_df.columns.str.strip().str.upper()
 
p_total_records = len(primary_df)
print(f"   Rows loaded: {p_total_records:,}")
 
# ── Build composite keys ─────────────────────────────────────────────────────
if "MATERIAL_PLANT" not in primary_df.columns:
    primary_df["MATERIAL_PLANT"] = (
        primary_df["MATERIAL"].astype(str).str.strip() + "_" +
        primary_df["PLANT"].astype(str).str.strip()
    )
if "PLANT_SOLDTOPARTY" not in primary_df.columns:
    primary_df["PLANT_SOLDTOPARTY"] = (
        primary_df["PLANT"].astype(str).str.strip() + "_" +
        primary_df["SOLDTOPARTY"].astype(str).str.strip()
    )
 
# ── Validation rules ─────────────────────────────────────────────────────────
primary_df["ERROR_MATERIAL_PLANT"] = primary_df["MATERIAL_PLANT"].apply(
    lambda x: "Yes" if pd.isna(x) or x not in part_plant_set else "")
primary_df["ERROR_PLANT"] = primary_df["PLANT"].apply(
    lambda x: "Yes" if pd.isna(x) or x not in site_set else "")
primary_df["ERROR_PLANT_SOLDTOPARTY"] = primary_df["PLANT_SOLDTOPARTY"].apply(
    lambda x: "Yes" if pd.isna(x) or x not in customer_plant_set else "")
primary_df["ERROR_BILLING_DATE"] = primary_df["BILLING_DATE"].apply(
    lambda x: "Yes" if pd.isna(x) or not date_pattern.match(str(x)) else "")
 
# Duplicate check — keep=False flags ALL occurrences
PRIMARY_DUP_COLS = ["MATERIAL", "PLANT", "SOLDTOPARTY", "BILLING_DATE"]
primary_df["ERROR_DUPLICATE_CHECK"] = primary_df.duplicated(
    subset=PRIMARY_DUP_COLS, keep=False
).map({True: "Yes", False: ""})
 
# ── Count errors per rule ────────────────────────────────────────────────────
print("🔍 Counting errors for HDA Primary...")
p_error_counts = {}
for field, col, _ in PRIMARY_RULES:
    p_error_counts[field] = int((primary_df[col] == "Yes").sum())
 
error_cols_primary    = [col for _, col, _ in PRIMARY_RULES]
p_records_with_errors = int((primary_df[error_cols_primary] == "Yes").any(axis=1).sum())
 
print(f"   ✅ Primary: {p_total_records:,} records scanned")
 
 
# ══════════════════════════════════════════════════════
#  HDA SECONDARY — rules & counting
#  Order: 1) DISTRIBUTOR_CODE  2) PLANT  3) INVOICE_DATE
#         4) CSKU              5) DUPLICATE_CHECK
# ══════════════════════════════════════════════════════
SECONDARY_RULES = [
    ("DISTRIBUTOR_CODE", "ERROR_DISTRIBUTOR_CODE", "DISTRIBUTOR_CODE: Distributor code is blank or missing in Customer master"),
    ("PLANT",            "ERROR_PLANT",            "PLANT: Plant does not exist in Site master or is blank"),
    ("INVOICE_DATE",     "ERROR_INVOICE_DATE",     "INVOICE_DATE: Invoice week start is blank or not in YYYYMMDD format"),
    ("CSKU",             "ERROR_CSKU",             "CSKU: CSKU missing in Part master"),
    ("DUPLICATE_CHECK",  "ERROR_DUPLICATE_CHECK",  "Duplicate record: DISTRIBUTOR_CODE + PLANT + INVOICE_DATE + CSKU combination already exists in the extract."),
]
 
SECONDARY_RULESET_DESC = {
    "DISTRIBUTOR_CODE": [
        "Must not be blank.",
        "Must exist as CUSTOMER in Customer master.",
    ],
    "PLANT": [
        "Must not be blank.",
        "Must exist in Site master.",
    ],
    "INVOICE_DATE": [
        "Must not be blank.",
        "Must strictly be in YYYYMMDD format.",
    ],
    "CSKU": [
        "Must not be blank.",
        "Must exist as MATERIALNUMBER in Part master.",
    ],
    "DUPLICATE_CHECK": [
        "There should be no duplicate records in the extract.",
        "A duplicate is identified when DISTRIBUTOR_CODE + PLANT + INVOICE_DATE + CSKU are all identical across two or more rows.",
        "All rows involved in a duplicate combination are flagged as errors.",
    ],
}
 
# ── Load full HDA Secondary file ─────────────────────────────────────────────
print("\n📂 Loading HDA Secondary file (full load for duplicate detection)...")
secondary_df = pd.read_csv(HDA_SECONDARY_FILE, sep="\t", dtype=str)
secondary_df = secondary_df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
secondary_df.columns = secondary_df.columns.str.strip().str.upper()
 
s_total_records = len(secondary_df)
print(f"   Rows loaded: {s_total_records:,}")
 
# ── Validation rules ─────────────────────────────────────────────────────────
secondary_df["ERROR_DISTRIBUTOR_CODE"] = secondary_df["DISTRIBUTOR_CODE"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or x not in customer_set else "")
secondary_df["ERROR_PLANT"] = secondary_df["PLANT"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or x not in site_set else "")
secondary_df["ERROR_INVOICE_DATE"] = secondary_df["INVOICE_DATE"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or not date_pattern.fullmatch(str(x)) else "")
secondary_df["ERROR_CSKU"] = secondary_df["CSKU"].apply(
    lambda x: "Yes" if pd.isna(x) or x == "" or x not in part_set else "")
 
# Duplicate check — keep=False flags ALL occurrences
SECONDARY_DUP_COLS = ["DISTRIBUTOR_CODE", "PLANT", "INVOICE_DATE", "CSKU"]
secondary_df["ERROR_DUPLICATE_CHECK"] = secondary_df.duplicated(
    subset=SECONDARY_DUP_COLS, keep=False
).map({True: "Yes", False: ""})
 
# ── Count errors per rule ────────────────────────────────────────────────────
print("🔍 Counting errors for HDA Secondary...")
s_error_counts = {}
for field, col, _ in SECONDARY_RULES:
    s_error_counts[field] = int((secondary_df[col] == "Yes").sum())
 
error_cols_secondary   = [col for _, col, _ in SECONDARY_RULES]
s_records_with_errors  = int((secondary_df[error_cols_secondary] == "Yes").any(axis=1).sum())
 
print(f"   ✅ Secondary: {s_total_records:,} records scanned")
 
 
# ══════════════════════════════════════════════════════
#  BUILD EXCEL — Summary + Ruleset only (no error rows)
# ══════════════════════════════════════════════════════
print("\n📝 Writing Summary & Ruleset sheets...")
wb = Workbook()
wb.remove(wb.active)  # remove default "Sheet"
 
 
def write_summary(wb, sheet_title, title_text, rules, error_counts,
                  total_records, records_with_errors):
    """Create a styled Summary sheet."""
    ws = wb.create_sheet(sheet_title)
 
    # ── Title row ──
    ws.merge_cells("A1:G1")
    tc           = ws.cell(row=1, column=1, value=title_text)
    tc.font      = TITLE_FONT
    tc.fill      = TITLE_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 24
 
    # ── Header row ──
    headers = ["#", "Field Name", "Error Count", "Record Count",
               "% Health", "% of Error", "Reason"]
    for ci, h in enumerate(headers, 1):
        c           = ws.cell(row=2, column=ci, value=h)
        c.fill      = TITLE_FILL
        c.font      = BOLD_FONT
        c.border    = THIN_BORDER
        c.alignment = CENTER
 
    # ── Data rows (order driven entirely by the rules list passed in) ──
    row = 3
    for idx, (field, _, reason) in enumerate(rules, 1):
        cnt        = error_counts[field]
        pct_error  = round((cnt / total_records) * 100, 2) if total_records else 0
        pct_health = round(100 - pct_error, 2)
        display_reason = reason if cnt > 0 else ""
 
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=field)
        ws.cell(row=row, column=3, value=cnt)
        ws.cell(row=row, column=4, value=total_records)
        ws.cell(row=row, column=5, value=f"{pct_health}%")
        ws.cell(row=row, column=6, value=f"{pct_error}%")
        ws.cell(row=row, column=7, value=display_reason)
 
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.font   = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = (Alignment(horizontal="left", vertical="center", wrap_text=True)
                           if col == 7 else CENTER)
        row += 1
 
    # ── TOTAL row ──
    total_errors       = sum(error_counts[f] for f, _, _ in rules)
    total_record_count = total_records * len(rules)
    total_pct_error    = round((total_errors / total_record_count) * 100, 2) if total_record_count else 0
    total_pct_health   = round(100 - total_pct_error, 2)
 
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="TOTAL")
    ws.cell(row=row, column=3, value=total_errors)
    ws.cell(row=row, column=4, value=total_record_count)
    ws.cell(row=row, column=5, value=f"{total_pct_health}%")
    ws.cell(row=row, column=6, value=f"{total_pct_error}%")
    ws.cell(row=row, column=7, value="")
 
    for col in range(1, 8):
        c        = ws.cell(row=row, column=col)
        c.font   = BOLD_FONT
        c.fill   = TOTAL_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER
    row += 2
 
    # ── Stats block ──
    records_passing = total_records - records_with_errors
    for label, value in [("Total Records:",        total_records),
                         ("Records with Errors:",  records_with_errors),
                         ("Records Passing:",       records_passing)]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        lc           = ws.cell(row=row, column=1, value=label)
        lc.font      = Font(name="Arial", bold=True, size=10)
        lc.fill      = STATS_FILL
        lc.border    = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center")
 
        vc           = ws.cell(row=row, column=3, value=value)
        vc.font      = BODY_FONT
        vc.border    = THIN_BORDER
        vc.alignment = CENTER
        row += 1
 
    # ── Autofit columns ──
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 3, 10), 60)
 
 
def write_ruleset(wb, sheet_title, title_text, ruleset_desc):
    """Create a styled Ruleset sheet.
    Row order matches the insertion order of ruleset_desc (Python 3.7+ dicts are ordered).
    """
    wsr = wb.create_sheet(sheet_title)
 
    wsr.merge_cells("A1:C1")
    tc           = wsr.cell(row=1, column=1, value=title_text)
    tc.font      = Font(name="Arial", bold=True, size=13)
    tc.fill      = TITLE_FILL
    tc.alignment = CENTER
    wsr.row_dimensions[1].height = 22
 
    for ci, h in enumerate(["#", "Field", "Rule Description"], 1):
        c           = wsr.cell(row=3, column=ci, value=h)
        c.fill      = HDR_FILL
        c.font      = BOLD_FONT
        c.border    = THIN_BORDER
        c.alignment = CENTER
 
    current_row = 4
    rule_num    = 1
 
    for field, rules_list in ruleset_desc.items():
        num_rules = len(rules_list)
 
        for r_idx, rule_text in enumerate(rules_list):
            nc           = wsr.cell(row=current_row, column=1, value=rule_num if r_idx == 0 else "")
            nc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
            nc.fill      = RULE_FILL
            nc.border    = THIN_BORDER
            nc.alignment = CENTER
 
            fc           = wsr.cell(row=current_row, column=2, value=field if r_idx == 0 else "")
            fc.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
            fc.fill      = RULE_FILL
            fc.border    = THIN_BORDER
            fc.alignment = CENTER
 
            dc           = wsr.cell(row=current_row, column=3, value=rule_text)
            dc.font      = BODY_FONT
            dc.border    = THIN_BORDER
            dc.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
 
            current_row += 1
 
        if num_rules > 1:
            s = current_row - num_rules
            e = current_row - 1
            wsr.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
            wsr.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)
 
        rule_num += 1
 
    wsr.column_dimensions["A"].width = 6
    wsr.column_dimensions["B"].width = 30
    wsr.column_dimensions["C"].width = 65
 
 
# ── Write Primary sheets ──
write_summary(wb, "HDA_Summary",  "HDA Validation Summary",
              PRIMARY_RULES, p_error_counts, p_total_records, p_records_with_errors)
write_ruleset(wb, "HDA_Rulesets", "HDA – Validation Rules", PRIMARY_RULESET_DESC)
 
# ── Write Secondary sheets ──
write_summary(wb, "HDA_Sec_Summary",  "HDA Secondary Validation Summary",
              SECONDARY_RULES, s_error_counts, s_total_records, s_records_with_errors)
write_ruleset(wb, "HDA_Sec_Rulesets", "HDA(Secondary) – Validation Rules", SECONDARY_RULESET_DESC)
 
wb.save(OUTPUT_EXCEL)
print(f"\n✅ Summary report saved → {OUTPUT_EXCEL}")
print("   Sheets: HDA_Summary | HDA_Rulesets | HDA_Sec_Summary | HDA_Sec_Rulesets")
 
